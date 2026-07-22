import os
import re
import sys
import uuid
import json
import shutil
import base64
import logging
import platform
import subprocess
import tempfile
import textwrap
import threading
import time
import zipfile
import unicodedata
import xml.etree.ElementTree as ET
from difflib import SequenceMatcher

_VLM_SEMAPHORE = threading.Semaphore(int(os.environ.get("VLM_CONCURRENCY", "5")))
_SOFFICE_LOCK = threading.Lock()
RENDER_DPI = int(os.environ.get("RENDER_DPI", "300"))

# VLM 호출 상한 노브(기본값은 기존 동작과 동일 — 정상 페이지엔 영향 없음)
VLM_TIMEOUT = int(os.environ.get("VLM_TIMEOUT", "300"))                   # 구조추출 VLM 호출 타임아웃(초)
VLM_MAX_ATTEMPTS = max(1, int(os.environ.get("VLM_MAX_ATTEMPTS", "3")))   # 한 페이지 추출의 최대 시도 수
VLM_MAX_TOKENS = max(512, int(os.environ.get("VLM_MAX_TOKENS", "16384"))) # 한 페이지 구조추출 응답 토큰 상한
VLM_STREAM_ABORT = os.environ.get("VLM_STREAM_ABORT", "1") == "1"         # 스트리밍 중 반복·과다 생성 조기 중단(0으로 비활성화)
VLM_STREAM_ABORT_MIN_CHARS = max(2048, int(os.environ.get("VLM_STREAM_ABORT_MIN_CHARS", "4096")))
VLM_STREAM_ABORT_INPUT_MIN_CHARS = max(128, int(os.environ.get("VLM_STREAM_ABORT_INPUT_MIN_CHARS", "256")))
VLM_STREAM_ABORT_BASE_CHARS = max(4096, int(os.environ.get("VLM_STREAM_ABORT_BASE_CHARS", "8000")))
VLM_STREAM_ABORT_INPUT_RATIO = max(2.0, float(os.environ.get("VLM_STREAM_ABORT_INPUT_RATIO", "8")))
VLM_COVERAGE_VERIFY = os.environ.get("VLM_COVERAGE_VERIFY", "1") == "1"   # 낮은 coverage 페이지를 이미지 기준 VLM 판정으로 재검증
VLM_COVERAGE_VERIFY_TIMEOUT = int(os.environ.get("VLM_COVERAGE_VERIFY_TIMEOUT", "180"))
VLM_COVERAGE_VERIFY_MAX_TOKENS = max(
    256, int(os.environ.get("VLM_COVERAGE_VERIFY_MAX_TOKENS", "16384"))
)
VLM_COVERAGE_VERIFY_IMG_MAXW = max(768, int(os.environ.get("VLM_COVERAGE_VERIFY_IMG_MAXW", "1600")))
VLM_COVERAGE_REPAIR = os.environ.get("VLM_COVERAGE_REPAIR", "1") == "1"   # verifier 실패 시 이미지 기준 재추출 1회
VLM_COVERAGE_REPAIR_TIMEOUT = int(os.environ.get("VLM_COVERAGE_REPAIR_TIMEOUT", "300"))
VLM_COVERAGE_REPAIR_MAX_TOKENS = max(512, int(os.environ.get("VLM_COVERAGE_REPAIR_MAX_TOKENS", str(VLM_MAX_TOKENS))))
VLM_COVERAGE_REPAIR_IMG_MAXW = max(768, int(os.environ.get("VLM_COVERAGE_REPAIR_IMG_MAXW", "1600")))
VLM_COMPACT_RETRY = os.environ.get("VLM_COMPACT_RETRY", "1") == "1"       # 반복/타임아웃 페이지를 간결 모드로 1회 구조화
VLM_COMPACT_RETRY_TIMEOUT = int(os.environ.get("VLM_COMPACT_RETRY_TIMEOUT", "240"))
VLM_COMPACT_RETRY_MAX_TOKENS = max(
    512, int(os.environ.get("VLM_COMPACT_RETRY_MAX_TOKENS", "16384"))
)
VLM_COMPACT_RETRY_IMG_MAXW = max(768, int(os.environ.get("VLM_COMPACT_RETRY_IMG_MAXW", "1400")))
DOC_VLM_BUDGET_SEC = float(os.environ.get("DOC_VLM_BUDGET_SEC", "0"))     # 문서 전체 VLM 시간 상한(초). 0=비활성
VLM_PAGE_CONCURRENCY = max(1, int(os.environ.get("VLM_PAGE_CONCURRENCY", "16")))  # 페이지 동시 추출 수(vLLM 배칭으로 가속)
# VLM 입력 이미지 폭 상한(px). 기본 2464(28의 배수) — 300DPI 렌더와 짝, 나란히 표 분리에 필요.
VLM_IMG_MAXW = max(512, int(os.environ.get("VLM_IMG_MAXW", "2464")))
# 재시도 시 낮출 폴백 해상도.
VLM_IMG_MAXW_FALLBACK = max(512, int(os.environ.get("VLM_IMG_MAXW_FALLBACK", "1024")))
# repetition_penalty(기본/강).
VLM_REP_PENALTY = float(os.environ.get("VLM_REP_PENALTY", "1.05"))
VLM_REP_PENALTY_HI = float(os.environ.get("VLM_REP_PENALTY_HI", "1.18"))
VLM_METADATA_TIMEOUT = max(30, int(os.environ.get("VLM_METADATA_TIMEOUT", "120")))
VLM_METADATA_ATTEMPTS = max(1, int(os.environ.get("VLM_METADATA_ATTEMPTS", "2")))
# PII 마스킹(사용자 지정). 콤마목록으로 켤 타입 지정. 기본 빈값=마스킹 안 함(충실 추출).
# 지원 타입: rrn(주민번호) bizno(사업자등록번호) email(이메일) phone(전화) account(계좌) card(카드)
PII_MASK_TYPES = {t.strip().lower() for t in os.environ.get("PII_MASK", "").split(",") if t.strip()}

# 타입별 PII 패턴(한국 양식 우선). 캡처는 형식 보존 마스킹.
_PII_PATTERNS = {
    "rrn":     re.compile(r'(?<!\d)(\d{6})[-\s]?([1-4]\d{6})(?!\d)'),            # 주민등록번호
    "bizno":   re.compile(r'(?<!\d)(\d{3})[-\s]?(\d{2})[-\s]?(\d{5})(?!\d)'),    # 사업자등록번호
    "card":    re.compile(r'(?<!\d)(\d{4})[-\s]?(\d{4})[-\s]?(\d{4})[-\s]?(\d{4})(?!\d)'),
    "account": re.compile(r'(?<!\d)\d{2,6}[-\s]\d{2,6}[-\s]\d{2,7}(?:[-\s]\d{1,6})?(?!\d)'),
    "phone":   re.compile(r'(?<!\d)(01[016789])[-\s]?(\d{3,4})[-\s]?(\d{4})(?!\d)'),
    "email":   re.compile(r'\b([A-Za-z0-9._%+\-]+)@([A-Za-z0-9.\-]+\.[A-Za-z]{2,})\b'),
}


def _mask_one(kind, m):
    """매칭된 PII 를 형식 보존하며 마스킹."""
    if kind == "email":
        loc = m.group(1)
        return (loc[0] + "***" if loc else "***") + "@" + m.group(2)
    if kind == "rrn":
        return f"{m.group(1)}-*******"
    if kind == "bizno":
        return f"{m.group(1)}-**-*****"
    if kind == "phone":
        return f"{m.group(1)}-****-{m.group(3)}"
    if kind == "card":
        return f"{m.group(1)}-****-****-{m.group(4)}"
    return re.sub(r"\d", "*", m.group(0))   # account 등: 숫자만 가림


def mask_pii(text, types=None):
    """설정된 타입의 PII 만 마스킹(끄면 원문 그대로). structured/청크 공용."""
    types = PII_MASK_TYPES if types is None else types
    if not text or not types:
        return text
    out = text
    for kind in ("card", "account", "rrn", "bizno", "phone", "email"):   # 긴 패턴 우선
        if kind in types and kind in _PII_PATTERNS:
            out = _PII_PATTERNS[kind].sub(lambda m, k=kind: _mask_one(k, m), out)
    return out


# Continuous columns are extracted independently to make visual order explicit.
VLM_MULTICOL = os.environ.get("VLM_MULTICOL", "1") == "1"

def _detect_column_split(png_path):
    """2단 본문이면 가운데 거터 x비율(0~1) 반환, 아니면 None."""
    try:
        from PIL import Image
        with Image.open(png_path) as im:
            rgb = im.convert("RGB")
            W, H = rgb.size
            sw = 500
            if W > sw:
                rgb = rgb.resize((sw, max(1, int(H * sw / W))))
                W, H = rgb.size
            g = rgb.convert("L")
            gray_pixels = g.tobytes()
            rgb_values = iter(rgb.tobytes())
            total_pixels = max(1, W * H)

            # Cropping is only useful for text-dominant pages. Screenshot-heavy
            # manuals and richly coloured layouts can also have a central gap,
            # but their intended order is not necessarily left-column first.
            ink_fraction = sum(value < 245 for value in gray_pixels) / total_pixels
            colour_fraction = sum(
                max(red, green, blue) - min(red, green, blue) > 20 and gray < 245
                for (red, green, blue), gray in zip(
                    zip(rgb_values, rgb_values, rgb_values), gray_pixels
                )
            ) / total_pixels
            if ink_fraction > 0.28 or colour_fraction > 0.04:
                return None

            px = g.load()
            colfill = [sum(1 for y in range(H) if px[x, y] < 190) / H for x in range(W)]  # 열별 잉크 비율
            # 표 오탐 가드: 중앙대역 세로룰(>0.5) 있으면 표로 보고 미발동.
            if max(colfill[int(W * 0.35):int(W * 0.65)] or [0]) > 0.5:
                return None
            # 얇은 표 선은 축소 과정에서 세로룰 비율이 낮아질 수 있다. 중앙을
            # 가로지르는 독립적인 긴 수평선이 반복되면 표/패널 레이아웃이다.
            central_lo, central_hi = int(W * 0.35), int(W * 0.65)
            min_rule_width = (central_hi - central_lo) * 0.60
            long_rule_rows = []
            for y in range(H):
                longest = current = 0
                for x in range(central_lo, central_hi):
                    if px[x, y] < 190:
                        current += 1
                        longest = max(longest, current)
                    else:
                        current = 0
                long_rule_rows.append(longest >= min_rule_width)
            rule_groups = sum(
                is_rule and (index == 0 or not long_rule_rows[index - 1])
                for index, is_rule in enumerate(long_rule_rows)
            )
            if rule_groups >= 4:
                return None
            lo, hi = int(W * 0.40), int(W * 0.60)
            band = [x for x in range(lo, hi) if colfill[x] < 0.04]                          # 중앙의 거의 빈 열
            if len(band) < W * 0.02:
                return None
            left_content = sum(1 for x in range(0, int(W * 0.35)) if colfill[x] > 0.10)
            right_content = sum(1 for x in range(int(W * 0.65), W) if colfill[x] > 0.10)
            if left_content > W * 0.05 and right_content > W * 0.05:                        # 좌·우 양쪽에 본문 존재
                return (sum(band) / len(band)) / W
    except Exception:
        pass
    return None


def _image_has_ink(png_path, white=245, min_frac=0.004):
    """페이지에 비백색 픽셀(잉크)이 있는가."""
    try:
        from PIL import Image
        with Image.open(png_path) as im:
            g = im.convert("L")
            g.thumbnail((400, 400))
            px = list(g.getdata())
            if not px:
                return False
            dark = sum(1 for p in px if p < white)
            return (dark / len(px)) >= min_frac
    except Exception:
        return False


def _korean_ratio(text):
    """한글:라틴 글자 비율(0~1). 숫자·기호·공백 무시. 문서/설명 언어 판정용."""
    if not text:
        return 0.0
    kr = sum(1 for ch in text if "가" <= ch <= "힣")
    en = sum(1 for ch in text if "a" <= ch.lower() <= "z")
    tot = kr + en
    return (kr / tot) if tot else 0.0


def _semantic_text_length(text):
    """Count content while excluding JSON and HTML serialization overhead."""
    value = re.sub(r"<[^>]*>", "", text or "")
    value = re.sub(
        r'"(?:page_number|elements|type|content|caption|description)"\s*:',
        "",
        value,
    )
    return sum(ch.isalnum() for ch in value)


def _stream_output_excessive(generated, source_text):
    """Detect non-repeating overgeneration when a useful text layer exists."""
    source_chars = _semantic_text_length(source_text)
    if source_chars < VLM_STREAM_ABORT_INPUT_MIN_CHARS:
        return False
    limit = max(
        VLM_STREAM_ABORT_BASE_CHARS,
        int(source_chars * VLM_STREAM_ABORT_INPUT_RATIO),
    )
    return _semantic_text_length(generated) > limit


def _collapse_compound_repeat_text(text):
    """Collapse short adjacent duplicate runs such as '참고참고1717' -> '참고17'."""
    if not isinstance(text, str):
        return text
    leading = text[:len(text) - len(text.lstrip())]
    trailing = text[len(text.rstrip()):]
    core = text.strip()
    if not core or len(core) > 40:
        return text

    def _collapse_token(token):
        # Whole-string exact repetition: AAAAAA -> A.
        for n in range(1, min(20, len(token) // 2) + 1):
            if len(token) % n == 0:
                unit = token[:n]
                if unit and unit * (len(token) // n) == token:
                    return unit
        # Two adjacent duplicate runs: A A B B -> A B.
        max_part = min(16, len(token) // 2)
        for a_len in range(1, max_part + 1):
            a = token[:a_len]
            if not token.startswith(a * 2):
                continue
            rest = token[a_len * 2:]
            for b_len in range(1, min(16, len(rest) // 2) + 1):
                b = rest[:b_len]
                if b and len(rest) % b_len == 0 and b * (len(rest) // b_len) == rest:
                    return a + b
        return token

    if not re.search(r"\s", core):
        collapsed = _collapse_token(core)
        return leading + collapsed + trailing if collapsed != core else text

    # OCR/VLM can insert a space between duplicated runs: "참고참고 1717".
    # Require at least two changed tokens so ordinary words such as "하하 결과" stay intact.
    pieces = re.split(r"(\s+)", core)
    changed = 0
    for index in range(0, len(pieces), 2):
        collapsed = _collapse_token(pieces[index])
        if collapsed != pieces[index]:
            pieces[index] = collapsed
            changed += 1
    if changed >= 2:
        return leading + "".join(pieces) + trailing
    return text


def _dedupe_exact_text_elements(elements):
    """Drop exact duplicate prose while preserving repeated structural elements."""
    seen_text = set()
    result = []
    for element in elements or []:
        if not isinstance(element, dict):
            result.append(element)
            continue
        element_type = element.get("type", "text")
        content = re.sub(r"\s+", " ", (element.get("content") or "")).strip()
        if element_type in {"text", "footnote"} and content:
            if content in seen_text:
                continue
            seen_text.add(content)
        result.append(element)
    return result


def _element_visible_text(element):
    """Return user-visible element text without serialization markup."""
    if not isinstance(element, dict):
        return ""
    content = str(element.get("content") or "")
    if element.get("type") == "table":
        try:
            return BeautifulSoup(content, "html.parser").get_text(" ", strip=True)
        except Exception:
            return re.sub(r"<[^>]+>", " ", content)
    return content


def _compact_visible_text(text):
    value = unicodedata.normalize("NFKC", str(text or "")).casefold()
    return "".join(char for char in value if char.isalnum())


def _paginate_plain_text(text, width=84, lines_per_page=88):
    """Wrap trusted plain text into deterministic, non-empty display pages."""
    wrapped = []
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    for raw_line in normalized.split("\n"):
        if not raw_line.strip():
            if wrapped and wrapped[-1] != "":
                wrapped.append("")
            continue
        wrapped.extend(
            textwrap.wrap(
                raw_line,
                width=max(20, int(width)),
                break_long_words=True,
                break_on_hyphens=False,
                replace_whitespace=False,
                drop_whitespace=True,
            )
            or [raw_line]
        )
    while wrapped and not wrapped[-1]:
        wrapped.pop()
    if not wrapped:
        return []
    page_size = max(10, int(lines_per_page))
    return [
        "\n".join(wrapped[start:start + page_size]).strip()
        for start in range(0, len(wrapped), page_size)
        if any(line.strip() for line in wrapped[start:start + page_size])
    ]


def _dedupe_comparison_text(text):
    """Normalize formatting while retaining address and URL identity."""
    return _compact_visible_text(text)


def _compact_text_with_offsets(text):
    """Normalize prose while retaining offsets into the original string."""
    chars = []
    offsets = []
    for index, char in enumerate(str(text or "")):
        normalized = unicodedata.normalize("NFKC", char).casefold()
        for normalized_char in normalized:
            if normalized_char.isalnum():
                chars.append(normalized_char)
                offsets.append(index)
    return "".join(chars), offsets


def _resolve_flattened_table_duplicates(elements):
    """Split a page-wide prose duplicate around exact structured table blocks.

    Some VLM responses contain the whole page once as flat prose and then emit
    the same table and trailing blocks structurally. Reconstruct reading order
    only when a complete, substantial table is an exact normalized substring;
    shorter phrase overlap alone is never enough to trigger this transform.
    """
    source = list(elements or [])
    consumed = set()
    result = []

    for index, element in enumerate(source):
        if index in consumed:
            continue
        if not isinstance(element, dict) or element.get("type", "text") != "text":
            result.append(element)
            continue

        raw = str(element.get("content") or "")
        container, offsets = _compact_text_with_offsets(raw)
        if len(container) < 160:
            result.append(element)
            continue

        candidates = []
        for child_index in range(index + 1, len(source)):
            child = source[child_index]
            if child_index in consumed or not isinstance(child, dict):
                continue
            child_type = child.get("type", "text")
            if child_type not in {
                "table", "text", "footnote", "heading_1", "heading_2", "heading_3"
            }:
                continue
            child_text = _compact_visible_text(_element_visible_text(child))
            minimum = 32 if child_type == "table" else 12
            if len(child_text) < minimum:
                continue
            start = container.find(child_text)
            while start >= 0:
                candidates.append(
                    {
                        "index": child_index,
                        "type": child_type,
                        "start": start,
                        "end": start + len(child_text),
                        "length": len(child_text),
                    }
                )
                start = container.find(child_text, start + 1)

        table_candidates = [item for item in candidates if item["type"] == "table"]
        if not table_candidates:
            result.append(element)
            continue

        exact_children = {item["index"] for item in candidates}
        for child_index in range(index + 1, len(source)):
            if child_index in exact_children or child_index in consumed:
                continue
            child = source[child_index]
            if not isinstance(child, dict) or child.get("type", "text") not in {
                "text", "footnote", "heading_1", "heading_2", "heading_3"
            }:
                continue
            child_text = _compact_visible_text(_element_visible_text(child))
            if len(child_text) < 12:
                continue
            matcher = SequenceMatcher(None, container, child_text, autojunk=False)
            blocks = [block for block in matcher.get_matching_blocks() if block.size]
            if not blocks:
                continue
            first, last = blocks[0], blocks[-1]
            start = max(0, first.a - first.b)
            end = min(
                len(container),
                last.a + last.size + len(child_text) - (last.b + last.size),
            )
            span_length = end - start
            if not (len(child_text) * 0.75 <= span_length <= len(child_text) * 1.35):
                continue
            window_matcher = SequenceMatcher(
                None, container[start:end], child_text, autojunk=False
            )
            matched_chars = sum(block.size for block in window_matcher.get_matching_blocks())
            if (
                window_matcher.ratio() >= 0.90
                and matched_chars / len(child_text) >= 0.85
            ):
                candidates.append(
                    {
                        "index": child_index,
                        "type": child.get("type", "text"),
                        "start": start,
                        "end": end,
                        "length": len(child_text),
                    }
                )

        # Assign each structured child once. Tables win overlapping spans, then
        # longer exact blocks win over their own short labels.
        selected = []
        used_children = set()
        for item in sorted(
            candidates,
            key=lambda value: (
                value["type"] != "table",
                -value["length"],
                value["start"],
            ),
        ):
            if item["index"] in used_children:
                continue
            if any(
                item["start"] < chosen["end"] and chosen["start"] < item["end"]
                for chosen in selected
            ):
                continue
            selected.append(item)
            used_children.add(item["index"])

        selected.sort(key=lambda value: value["start"])
        if not any(item["type"] == "table" for item in selected):
            result.append(element)
            continue

        unmatched_text = []
        for child_index in range(index + 1, len(source)):
            if child_index in used_children or child_index in consumed:
                continue
            child = source[child_index]
            if not isinstance(child, dict) or child.get("type", "text") not in {"text", "footnote"}:
                continue
            value = _compact_visible_text(child.get("content"))
            if len(value) >= 12:
                unmatched_text.append(value)

        def _append_residual(start, end):
            if start >= end or not offsets:
                return
            raw_start = 0 if start == 0 else offsets[start - 1] + 1
            raw_end = len(raw) if end >= len(offsets) else offsets[end]
            fragment = raw[raw_start:raw_end].strip()
            normalized = _compact_visible_text(fragment)
            if len(normalized) < 2:
                return
            # Prefer an already separated block when the residual is merely an
            # OCR/VLM spelling variant of it.
            for other in unmatched_text:
                if max(len(normalized), len(other)) > min(len(normalized), len(other)) * 1.25:
                    continue
                matcher = SequenceMatcher(None, normalized, other)
                if matcher.ratio() >= 0.94 and matcher.find_longest_match().size >= 12:
                    return
            residual = dict(element)
            residual["content"] = fragment
            result.append(residual)

        cursor = 0
        for item in selected:
            _append_residual(cursor, item["start"])
            result.append(source[item["index"]])
            consumed.add(item["index"])
            cursor = item["end"]
        _append_residual(cursor, len(container))

    return result


def _drop_prose_duplicated_by_nearby_tables(elements):
    """Drop nearby prose that exactly repeats a visible table cell/value."""
    source = list(elements or [])
    tables = []
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type") != "table":
            continue
        content = str(element.get("content") or "")
        try:
            soup = BeautifulSoup(content, "html.parser")
            cells = {
                _compact_visible_text(cell.get_text(" ", strip=True))
                for cell in soup.find_all(["td", "th"])
            }
            visible = _compact_visible_text(soup.get_text(" ", strip=True))
        except Exception:
            cells = set()
            visible = _compact_visible_text(re.sub(r"<[^>]+>", " ", content))
        tables.append((index, visible, {value for value in cells if value}))

    cleaned = []
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type", "text") != "text":
            cleaned.append(element)
            continue
        value = _compact_visible_text(element.get("content"))
        duplicate = False
        if len(value) >= 4:
            for table_index, table_text, cells in tables:
                if abs(index - table_index) > 3:
                    continue
                if value in cells or (len(value) >= 12 and value in table_text):
                    duplicate = True
                    break
        if not duplicate:
            cleaned.append(element)
    return cleaned


def _drop_trailing_duplicate_heading_cluster(elements):
    """Drop a heading-only tail when every heading already appeared on the page."""
    result = list(elements or [])
    heading_end = len(result)
    if heading_end:
        suffix = result[-1]
        suffix_text = (
            str(suffix.get("content") or "").strip()
            if isinstance(suffix, dict) else ""
        )
        if (
            isinstance(suffix, dict)
            and suffix.get("type", "text") in {"text", "footnote"}
            and re.fullmatch(r"\d{4}", suffix_text)
        ):
            heading_end -= 1
    start = heading_end
    while start > 0:
        element = result[start - 1]
        if not isinstance(element, dict) or not str(element.get("type") or "").startswith("heading_"):
            break
        start -= 1
    if heading_end - start < 2:
        return result

    prior = {
        _compact_visible_text(element.get("content"))
        for element in result[:start]
        if isinstance(element, dict)
        and str(element.get("type") or "").startswith("heading_")
    }
    tail = [
        _compact_visible_text(element.get("content"))
        for element in result[start:heading_end]
    ]
    if tail and all(value and value in prior for value in tail):
        return result[:start] + result[heading_end:]
    return result


def _dedupe_figures_supported_by_text_layer(elements, source_text):
    """Remove exact duplicate figures only when the text layer supports fewer copies."""
    source = list(elements or [])
    normalized_source = _compact_visible_text(source_text)
    if not normalized_source:
        return source

    groups = {}
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type") != "figure":
            continue
        signature = tuple(
            re.sub(r"\s+", " ", str(element.get(key) or "")).strip()
            for key in ("content", "caption", "description")
        )
        if len(_compact_visible_text(signature[0])) < 32:
            continue
        groups.setdefault(signature, []).append(index)

    remove = set()
    for indices in groups.values():
        if len(indices) < 2:
            continue
        first_index, last_index = min(indices), max(indices)
        intervening = [
            source[index]
            for index in range(first_index, last_index + 1)
            if index not in indices
        ]
        if any(
            not isinstance(element, dict)
            or not str(element.get("type") or "").startswith("heading_")
            for element in intervening
        ):
            continue
        figure = source[indices[0]]
        anchors = []
        content = str(figure.get("content") or "")
        try:
            soup = BeautifulSoup(content, "html.parser")
            if soup.find(["td", "th"]):
                anchors.extend(
                    cell.get_text(" ", strip=True)
                    for cell in soup.find_all(["td", "th"])
                )
            else:
                anchors.extend(soup.get_text("\n", strip=True).splitlines())
        except Exception:
            anchors.extend(content.splitlines())
        anchors.append(str(figure.get("caption") or ""))

        normalized_anchors = sorted(
            {
                value
                for item in anchors
                if len(value := _compact_visible_text(item)) >= 10
            },
            key=len,
            reverse=True,
        )[:8]
        occurrence_evidence = [
            (anchor, count)
            for anchor in normalized_anchors
            if (count := normalized_source.count(anchor)) > 0
        ]

        def immediate_heading_source_order(index):
            if index <= 0:
                return False
            previous = source[index - 1]
            if not isinstance(previous, dict) or not str(
                previous.get("type") or ""
            ).startswith("heading_"):
                return False
            heading = _compact_visible_text(previous.get("content"))
            heading_position = normalized_source.find(heading)
            return bool(
                len(heading) >= 4
                and heading_position >= 0
                and any(
                    normalized_source.find(anchor, heading_position + len(heading))
                    >= 0
                    for anchor in normalized_anchors
                )
            )

        strong_occurrence_evidence = (
            len(occurrence_evidence) >= 2
            or any(len(anchor) >= 24 for anchor, _ in occurrence_evidence)
        )
        ordered_heading_evidence = any(
            immediate_heading_source_order(index) for index in indices
        )
        if not strong_occurrence_evidence and not (
            ordered_heading_evidence
            and any(len(anchor) >= 12 for anchor, _ in occurrence_evidence)
        ):
            continue
        supported_copies = min(count for _, count in occurrence_evidence)
        if supported_copies >= len(indices):
            continue

        caption = _compact_visible_text(figure.get("caption"))

        def context_score(index):
            score = 1 if immediate_heading_source_order(index) else 0
            for distance in (1, 2):
                previous_index = index - distance
                if previous_index < 0:
                    continue
                previous = source[previous_index]
                if not isinstance(previous, dict) or not str(
                    previous.get("type") or ""
                ).startswith("heading_"):
                    continue
                heading = _compact_visible_text(previous.get("content"))
                if heading and caption:
                    matcher = SequenceMatcher(
                        None, heading, caption, autojunk=False
                    )
                    if (
                        heading in caption
                        or caption in heading
                        or (
                            matcher.ratio() >= 0.65
                            and matcher.find_longest_match().size >= 6
                        )
                    ):
                        score = max(score, 3 - distance)
            return score

        context_scores = {index: context_score(index) for index in indices}
        if intervening and max(context_scores.values(), default=0) <= 0:
            continue
        keep_count = max(1, supported_copies)
        ranked = sorted(
            indices, key=lambda index: (-context_scores[index], index)
        )
        if intervening and keep_count < len(ranked):
            cutoff = context_scores[ranked[keep_count - 1]]
            if context_scores[ranked[keep_count]] == cutoff:
                continue
        keep = set(ranked[:keep_count])
        remove.update(index for index in indices if index not in keep)

    return [element for index, element in enumerate(source) if index not in remove]


def _labels_relate(left, right):
    left = _compact_visible_text(left)
    right = _compact_visible_text(right)
    if min(len(left), len(right)) < 4:
        return False
    if left in right or right in left:
        return True
    matcher = SequenceMatcher(None, left, right, autojunk=False)
    return matcher.ratio() >= 0.65 and matcher.find_longest_match().size >= 6


def _element_source_bounds(element, normalized_source, allow_short=False):
    """Locate an element from unique text-layer anchors without trusting its order."""
    if not isinstance(element, dict) or not normalized_source:
        return None
    element_type = str(element.get("type") or "")
    anchors = []
    if element_type == "table":
        try:
            soup = BeautifulSoup(element.get("content") or "", "html.parser")
            anchors.extend(
                _compact_visible_text(cell.get_text(" ", strip=True))
                for cell in soup.find_all(["td", "th"])
            )
        except Exception:
            pass
    else:
        content = _compact_visible_text(element.get("content"))
        if content:
            anchors.append(content)
            if len(content) >= 48:
                anchors.extend((content[:40], content[-40:]))
    caption = _compact_visible_text(element.get("caption"))
    if caption:
        anchors.append(caption)
    minimum = 4 if allow_short else 10
    matches = [
        (position, position + len(anchor), len(anchor))
        for anchor in dict.fromkeys(anchors)
        if len(anchor) >= minimum
        and normalized_source.count(anchor) == 1
        and (position := normalized_source.find(anchor)) >= 0
    ]
    if not matches:
        return None
    if not allow_short and len(matches) < 2 and max(item[2] for item in matches) < 24:
        return None
    return min(item[0] for item in matches), max(item[1] for item in matches)


def _restore_leading_source_blocks(elements, source_text):
    """Move a short misplaced suffix back before a leading heading when proven."""
    source = list(elements or [])
    if not source or not str(source[0].get("type") or "").startswith("heading_"):
        return source
    normalized_source = _compact_visible_text(source_text)
    heading_bounds = _element_source_bounds(
        source[0], normalized_source, allow_short=True
    )
    if heading_bounds is None:
        return source
    suffix_start = len(source)
    bounds = []
    while suffix_start > 1 and len(source) - suffix_start < 4:
        candidate = source[suffix_start - 1]
        if str(candidate.get("type") or "") not in {"text", "footnote", "table"}:
            break
        candidate_bounds = _element_source_bounds(candidate, normalized_source)
        if candidate_bounds is None or candidate_bounds[1] > heading_bounds[0]:
            break
        suffix_start -= 1
        bounds.append(candidate_bounds)
    if suffix_start == len(source):
        return source
    bounds.reverse()
    if any(left[0] >= right[0] for left, right in zip(bounds, bounds[1:])):
        return source
    return source[suffix_start:] + source[:suffix_start]


def _starts_with_lowercase_fragment(element):
    if not isinstance(element, dict) or element.get("type") != "text":
        return False
    value = str(element.get("content") or "").lstrip(" \t\r\n\"'“‘([{")
    return bool(value and "a" <= value[0] <= "z" and len(value) <= 500)


def _ends_without_sentence_boundary(element):
    if not isinstance(element, dict) or element.get("type") != "text":
        return False
    value = str(element.get("content") or "").rstrip()
    while value and value[-1] in "\"'”’)]}":
        value = value[:-1].rstrip()
    return bool(value and value[-1] not in ".!?。！？:")


def _normalize_roman_section_subheadings(elements):
    result = []
    within_roman_section = False
    major = re.compile(r"^(?:[IVXLCDM]+|[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+)[.)]\s+", re.I)
    for element in elements or []:
        current = element
        if isinstance(element, dict) and str(element.get("type") or "").startswith(
            "heading_"
        ):
            content = str(element.get("content") or "").strip()
            if major.match(content):
                within_roman_section = True
            elif within_roman_section and element.get("type") == "heading_1":
                current = {**element, "type": "heading_2"}
        result.append(current)
    return result


def _drop_clipped_multicolumn_header_fragments(columns, source_text):
    """Drop a short header suffix created by cropping a full-width page header."""
    cleaned = [list(column or []) for column in columns or []]
    if len(cleaned) < 2:
        return cleaned
    source_lines = [
        (line.strip(), _compact_visible_text(line))
        for line in str(source_text or "").splitlines()[:16]
        if line.strip()
    ]
    for column_index in range(1, len(cleaned)):
        remove = set()
        for element_index, element in enumerate(cleaned[column_index][:3]):
            if not isinstance(element, dict) or not str(
                element.get("type") or ""
            ).startswith("heading_"):
                continue
            raw = str(element.get("content") or "").strip()
            fragment = _compact_visible_text(raw)
            if not (4 <= len(fragment) <= 32):
                continue
            matches = []
            for _, full in source_lines:
                if (
                    len(full) >= len(fragment) + 5
                    and full.endswith(fragment)
                    and len(fragment) * 5 <= len(full) * 3
                ):
                    matches.append(full)
            if len(matches) != 1:
                continue
            first_alpha = next((char for char in raw if char.isalpha()), "")
            if first_alpha and first_alpha.islower():
                remove.add(element_index)
        if remove:
            cleaned[column_index] = [
                element
                for index, element in enumerate(cleaned[column_index])
                if index not in remove
            ]
    return cleaned


def _merge_multicolumn_elements(columns):
    """Merge independently extracted columns and repair a proven carry-over line."""
    if len(columns or []) != 2:
        return [element for column in columns or [] for element in column]
    left, right = (list(column or []) for column in columns)
    footer_start = len(left)
    while footer_start and left[footer_start - 1].get("type") == "footnote":
        footer_start -= 1
    left_body = left[:footer_start]
    candidates = [
        index
        for index, element in enumerate(right)
        if _starts_with_lowercase_fragment(element)
        and any(
            str(previous.get("type") or "").startswith("heading_")
            for previous in right[:index]
            if isinstance(previous, dict)
        )
    ]
    if (
        len(candidates) == 1
        and left_body
        and _ends_without_sentence_boundary(left_body[-1])
    ):
        continuation = right.pop(candidates[0])
        merged = left_body + [continuation] + left[footer_start:] + right
    else:
        merged = left + right
    return _normalize_roman_section_subheadings(merged)


def _table_text_layer_copy_support(element, source_text):
    """Return the copy count supported by stable table anchors, or ``None``."""
    normalized_source = _compact_visible_text(source_text)
    if not normalized_source or not isinstance(element, dict):
        return None
    try:
        soup = BeautifulSoup(element.get("content") or "", "html.parser")
        anchors = [
            _compact_visible_text(cell.get_text(" ", strip=True))
            for cell in soup.find_all(["td", "th"])
        ]
    except Exception:
        anchors = []
    anchors.append(_compact_visible_text(element.get("caption")))
    evidence = [
        (anchor, count)
        for anchor in sorted(
            {anchor for anchor in anchors if len(anchor) >= 10},
            key=len,
            reverse=True,
        )[:8]
        if (count := normalized_source.count(anchor)) > 0
    ]
    if len(evidence) < 2 and not any(len(anchor) >= 24 for anchor, _ in evidence):
        return None
    return min(count for _, count in evidence)


def _dedupe_tables_supported_by_text_layer(elements, source_text):
    """Drop a repeated table only when source anchors support fewer copies."""
    source = list(elements or [])
    normalized_source = _compact_visible_text(source_text)
    if not normalized_source:
        return source

    groups = {}
    table_anchors = {}
    table_context_labels = {}
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type") != "table":
            continue
        try:
            soup = BeautifulSoup(element.get("content") or "", "html.parser")
            visible = _compact_visible_text(soup.get_text(" ", strip=True))
            anchors = [
                _compact_visible_text(cell.get_text(" ", strip=True))
                for cell in soup.find_all(["td", "th"])
            ]
        except Exception:
            visible = _compact_visible_text(element.get("content"))
            anchors = []
        if len(visible) < 32:
            continue
        caption = _compact_visible_text(element.get("caption"))
        signature = (visible, caption)
        groups.setdefault(signature, []).append(index)
        table_anchors[index] = sorted(
            {anchor for anchor in anchors + [caption] if len(anchor) >= 10},
            key=len,
            reverse=True,
        )[:8]
        table_context_labels[index] = [
            label
            for label in dict.fromkeys(anchors + [caption])
            if 4 <= len(label) <= 80
        ][:32]

    remove = set()
    for (_, caption), indices in groups.items():
        if len(indices) < 2:
            continue
        intervening = [
            source[index]
            for index in range(min(indices), max(indices) + 1)
            if index not in indices
        ]
        if any(
            not isinstance(element, dict)
            or not str(element.get("type") or "").startswith("heading_")
            for element in intervening
        ):
            continue
        evidence = [
            (anchor, count)
            for anchor in table_anchors.get(indices[0], [])
            if (count := normalized_source.count(anchor)) > 0
        ]
        if len(evidence) < 2 and not any(len(anchor) >= 24 for anchor, _ in evidence):
            continue
        supported_copies = min(count for _, count in evidence)
        if supported_copies >= len(indices):
            continue

        unique_anchor_ranges = [
            (position, position + len(anchor))
            for anchor, count in evidence
            if anchor != caption
            and count == 1
            and (position := normalized_source.find(anchor)) >= 0
        ]
        if not unique_anchor_ranges:
            unique_anchor_ranges = [
                (position, position + len(anchor))
                for anchor, count in evidence
                if count == 1
                and (position := normalized_source.find(anchor)) >= 0
            ]
        anchor_start = (
            min(start for start, _ in unique_anchor_ranges)
            if unique_anchor_ranges
            else None
        )
        anchor_end = (
            max(end for _, end in unique_anchor_ranges)
            if unique_anchor_ranges
            else None
        )

        def nearby_heading(index, direction):
            for distance in range(1, 4):
                candidate_index = index + direction * distance
                if not (0 <= candidate_index < len(source)):
                    break
                candidate = source[candidate_index]
                if isinstance(candidate, dict) and str(
                    candidate.get("type") or ""
                ).startswith("heading_"):
                    return candidate
            return None

        def context_score(index):
            score = 0
            for distance in (1, 2):
                previous_index = index - distance
                if previous_index < 0:
                    continue
                previous = source[previous_index]
                if (
                    isinstance(previous, dict)
                    and str(previous.get("type") or "").startswith("heading_")
                    and any(
                        _labels_relate(previous.get("content"), label)
                        for label in table_context_labels.get(indices[0], [])
                    )
                ):
                    score = max(score, 4 - distance)
            if anchor_start is None or anchor_end is None:
                return score
            previous = nearby_heading(index, -1)
            following = nearby_heading(index, 1)
            previous_before = False
            following_after = False
            if previous is not None:
                heading = _compact_visible_text(previous.get("content"))
                positions = [
                    match.start()
                    for match in re.finditer(re.escape(heading), normalized_source)
                ] if heading else []
                if positions:
                    previous_before = any(
                        position + len(heading) <= anchor_start
                        for position in positions
                    )
                    score += 1 if previous_before else -2
            if following is not None:
                heading = _compact_visible_text(following.get("content"))
                positions = [
                    match.start()
                    for match in re.finditer(re.escape(heading), normalized_source)
                ] if heading else []
                if positions:
                    following_after = any(position >= anchor_end for position in positions)
                    score += 1 if following_after else -1
            if previous_before and following_after:
                score += 4
            return score

        scores = {index: context_score(index) for index in indices}
        if intervening and max(scores.values(), default=0) <= 0:
            continue
        keep_count = max(1, supported_copies)
        ranked = sorted(indices, key=lambda index: (-scores[index], index))
        if intervening and keep_count < len(ranked):
            cutoff = scores[ranked[keep_count - 1]]
            if scores[ranked[keep_count]] == cutoff:
                continue
        keep = set(ranked[:keep_count])
        remove.update(index for index in indices if index not in keep)

    return [element for index, element in enumerate(source) if index not in remove]


def _dedupe_headings_supported_by_text_layer(elements, source_text):
    """Resolve duplicate headings using source counts and adjacent content."""
    source = list(elements or [])
    normalized_source = _compact_visible_text(source_text)
    if not normalized_source:
        return source
    groups = {}
    for index, element in enumerate(source):
        if not isinstance(element, dict) or not str(
            element.get("type") or ""
        ).startswith("heading_"):
            continue
        signature = _compact_visible_text(element.get("content"))
        if len(signature) >= 5:
            groups.setdefault(signature, []).append(index)

    remove = set()

    normalized_lines = [
        value
        for line in str(source_text or "").splitlines()
        if (value := _compact_visible_text(line))
    ]

    def standalone_occurrences(signature):
        count = 0
        for line_index in range(len(normalized_lines)):
            combined = ""
            for width in range(1, 4):
                if line_index + width > len(normalized_lines):
                    break
                combined += normalized_lines[line_index + width - 1]
                if combined == signature:
                    count += 1
                    break
                if len(combined) >= len(signature):
                    break
        return count

    for signature, indices in groups.items():
        if len(indices) < 2:
            continue
        supported_copies = standalone_occurrences(signature)
        if supported_copies < 1 or supported_copies >= len(indices):
            continue

        def context_score(index):
            score = 0
            heading = source[index].get("content")
            for distance in (1, 2):
                following_index = index + distance
                if following_index >= len(source):
                    continue
                following = source[following_index]
                if not isinstance(following, dict):
                    continue
                following_type = str(following.get("type") or "")
                if following_type in {"table", "figure"} and _labels_relate(
                    heading, following.get("caption")
                ):
                    score = max(score, 6 - distance)
                elif (
                    following_type in {"text", "footnote"}
                    and len(_compact_visible_text(following.get("content"))) >= 24
                ):
                    score = max(score, 3 - distance)
            return score

        scores = {index: context_score(index) for index in indices}
        ranked = sorted(indices, key=lambda index: (-scores[index], index))
        keep_count = max(1, supported_copies)
        if keep_count >= len(ranked):
            continue
        if scores[ranked[keep_count - 1]] <= scores[ranked[keep_count]]:
            continue
        keep = set(ranked[:keep_count])
        remove.update(index for index in indices if index not in keep)
    return [element for index, element in enumerate(source) if index not in remove]


def _drop_prose_duplicated_by_nearby_figures(elements, source_text):
    """Keep screenshot text once when a nearby figure already contains it."""
    source = list(elements or [])
    normalized_source = _compact_visible_text(source_text)
    figures = []
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type") != "figure":
            continue
        content = str(element.get("content") or "")
        if "<" in content and ">" in content:
            try:
                content = BeautifulSoup(content, "html.parser").get_text(
                    " ", strip=True
                )
            except Exception:
                pass
        visible = _compact_visible_text(
            " ".join(
                (
                    content,
                    str(element.get("caption") or ""),
                    str(element.get("description") or ""),
                )
            )
        )
        if visible:
            figures.append((index, visible))

    matches_by_figure = {}
    for index, element in enumerate(source):
        if not isinstance(element, dict) or element.get("type", "text") != "text":
            continue
        value = _compact_visible_text(element.get("content"))
        raw_value = str(element.get("content") or "").lstrip()
        numbered = bool(
            re.match(r"^(?:[①②③④⑤⑥⑦⑧⑨⑩]|\d+[.)])", raw_value)
        )
        if len(value) < (4 if numbered else 16) or normalized_source.count(value) > 1:
            continue
        matching_figures = [
            figure_index
            for figure_index, figure_text in figures
            if abs(index - figure_index) <= 8 and value in figure_text
        ]
        if len(matching_figures) == 1:
            matches_by_figure.setdefault(matching_figures[0], []).append(index)

    remove = set()
    for figure_index, text_indices in matches_by_figure.items():
        if len(text_indices) == 1:
            remove.add(text_indices[0])
            continue
        figure = source[figure_index]
        soup = BeautifulSoup(str(figure.get("content") or ""), "html.parser")
        table = soup.find("table")
        if table is None:
            continue
        row_texts = [
            _compact_visible_text(row.get_text(" ", strip=True))
            for row in table.find_all("tr")
        ]
        assignments = []
        for text_index in text_indices:
            value = _compact_visible_text(source[text_index].get("content"))
            matching_rows = [
                row_index
                for row_index, row_text in enumerate(row_texts)
                if value and value in row_text
            ]
            if len(matching_rows) != 1:
                assignments = []
                break
            assignments.append(matching_rows[0])
        if assignments and len(set(assignments)) == len(assignments):
            remove.update(text_indices)
    return [element for index, element in enumerate(source) if index not in remove]


def _mark_element(e, source=None, confidence=None, issues=None):
    """Attach lightweight provenance/confidence without changing public content fields."""
    if not isinstance(e, dict):
        return e
    if source and not e.get("_source"):
        e["_source"] = source
    if confidence is not None and e.get("_confidence") is None:
        try:
            e["_confidence"] = round(float(confidence), 3)
        except (TypeError, ValueError):
            pass
    if issues:
        cur = e.get("_issues")
        if not isinstance(cur, list):
            cur = []
        norm_cur = []
        for issue in cur:
            if isinstance(issue, (list, tuple)):
                issue = ":".join(str(x) for x in issue)
            elif issue is not None:
                issue = str(issue)
            if issue and issue not in norm_cur:
                norm_cur.append(issue)
        cur = norm_cur
        for issue in issues:
            if isinstance(issue, (list, tuple)):
                issue = ":".join(str(x) for x in issue)
            elif issue is not None:
                issue = str(issue)
            if issue and issue not in cur:
                cur.append(issue)
        if cur:
            e["_issues"] = cur
    return e


def _collect_provenance_summary(elements):
    summary = {"sources": {}, "issues": {}, "low_confidence_elements": 0}
    for e in elements or []:
        if not isinstance(e, dict):
            continue
        src = e.get("_source") or "unknown"
        summary["sources"][src] = summary["sources"].get(src, 0) + 1
        try:
            if float(e.get("_confidence", 1.0)) < 0.75:
                summary["low_confidence_elements"] += 1
        except (TypeError, ValueError):
            pass
        for issue in e.get("_issues") or []:
            if isinstance(issue, (list, tuple)):
                issue = ":".join(str(x) for x in issue)
            elif issue is not None:
                issue = str(issue)
            if not issue:
                continue
            summary["issues"][issue] = summary["issues"].get(issue, 0) + 1
    return summary


def _looks_page_artifact_text(text, page_no=None, index=0, total=0, has_page_marker=False):
    """Detect repeated page headers/footers that should not enter RAG text."""
    s = re.sub(r"\s+", " ", (text or "")).strip()
    if not s:
        return False
    edge = index <= 1 or index >= max(0, total - 3)
    # Report footers such as "- / 51 -", "- / 40 -", "- 18-".
    if re.fullmatch(r"[-–—]?\s*/\s*\d{1,4}\s*[-–—]?", s):
        return True
    if re.fullmatch(r"[-–—]\s*\d{1,4}\s*[-–—]?", s):
        return True
    # Standalone small page numbers are ambiguous. Drop them only when they match
    # the parsed page number or another strong footer marker exists on this page.
    if edge and re.fullmatch(r"\d{1,3}", s):
        try:
            n = int(s)
            if n <= 200:
                matches_page = page_no is not None and abs(n - int(page_no)) <= 3
                if matches_page or (has_page_marker and index >= max(0, total - 3)):
                    return True
        except (TypeError, ValueError):
            pass
    return False


def _strip_attached_page_counter(text, page_no=None, index=0, total=0):
    """Strip an edge-attached printed counter only when it matches this page."""
    value = str(text or "").strip()
    edge = index <= 1 or index >= max(0, total - 3)
    if not value or not edge or page_no is None:
        return value
    match = re.fullmatch(
        r"(?P<body>.+?)\s*[-–—]\s*/?\s*"
        r"(?P<digits>(?:\d\s*){1,4})[-–—]\s*",
        value,
    )
    if not match:
        return value
    try:
        printed_page = int(re.sub(r"\s+", "", match.group("digits")))
        parsed_page = int(page_no)
    except (TypeError, ValueError):
        return value
    body = match.group("body").strip()
    if printed_page != parsed_page or sum(char.isalnum() for char in body) < 4:
        return value
    return body


def _drop_page_artifact_elements(elements, page_no=None):
    """Remove page-number/header/footer noise from text-like elements only."""
    if not elements:
        return elements
    total = len(elements)
    has_page_marker = any(
        isinstance(e, dict)
        and e.get("type", "text") in {"text", "footnote"}
        and _looks_page_artifact_text(e.get("content"), None, idx, total)
        for idx, e in enumerate(elements)
    )
    cleaned = []
    for idx, e in enumerate(elements):
        if not isinstance(e, dict):
            cleaned.append(e)
            continue
        et = e.get("type", "text")
        if et in {"text", "footnote", "heading_1", "heading_2", "heading_3"}:
            stripped = _strip_attached_page_counter(
                e.get("content"), page_no, idx, total
            )
            if stripped != str(e.get("content") or "").strip():
                e = {**e, "content": stripped}
        if et in {"text", "footnote"} and _looks_page_artifact_text(
            e.get("content"), page_no, idx, total, has_page_marker
        ):
            continue
        cleaned.append(e)
    return cleaned


def _backfill_table_captions(elements):
    """Attach nearby visible section labels to following captionless tables."""
    if not elements:
        return elements

    def _candidate_title(e):
        if not isinstance(e, dict):
            return ""
        if not str(e.get("type", "")).startswith(("heading", "text")):
            return ""
        s = re.sub(r"\s+", " ", (e.get("content") or "")).strip()
        if not s or len(s) > 90:
            return ""
        if re.fullmatch(r"[-–—]?\s*/?\s*\d{1,4}\s*[-–—]?", s):
            return ""
        strong = (
            str(e.get("type", "")).startswith("heading") or
            any(k in s for k in ("표", "등급", "현황", "내역", "태도", "산정", "평가", "위험", "요약"))
        )
        return s if strong else ""

    last_title = ""
    for e in elements:
        if not isinstance(e, dict):
            continue
        if e.get("type") == "table":
            if not (e.get("caption") or "").strip():
                if last_title:
                    e["caption"] = last_title
            last_title = ""
            continue
        title = _candidate_title(e)
        if title:
            last_title = title
        elif e.get("type") not in {"footnote"}:
            # Only an immediately adjacent visible label may caption a table.
            last_title = ""
    return elements


def _standalone_table_unit_marker(value):
    marker = re.sub(r"\s+", " ", str(value or "")).strip()
    if not marker or len(marker) > 48:
        return ""

    bracketed = len(marker) >= 3 and (
        (marker[0] == "(" and marker[-1] == ")")
        or (marker[0] == "[" and marker[-1] == "]")
    )
    body = marker[1:-1].strip() if bracketed else marker
    if re.match(r"^(?:units?|단위)\s*[:：]", body, re.I):
        return marker
    if not bracketed:
        return ""

    if re.search(r"\b(?:USD|KRW|EUR|JPY|CNY|RMB)\b|[$€£¥₩]", body, re.I):
        return marker
    if re.fullmatch(
        r"(?:억|백만|천|만)?\s*(?:원|달러|유로|엔)"
        r"|(?:millions?|thousands?|billions?)"
        r"|(?:%|percent|kg|g|t|tonnes?|m[23²³]?|㎡|㎥)",
        body,
        re.I,
    ):
        return marker
    return ""


def _attach_adjacent_table_unit_captions(elements):
    """Move a trailing standalone unit label into its adjacent table caption."""
    if not elements:
        return elements

    repaired = []
    index = 0
    while index < len(elements):
        element = elements[index]
        following = elements[index + 1] if index + 1 < len(elements) else None
        marker = ""
        if (
            isinstance(element, dict)
            and element.get("type") == "table"
            and isinstance(following, dict)
            and following.get("type", "text") in {"text", "footnote"}
        ):
            marker = _standalone_table_unit_marker(following.get("content"))

        if not marker:
            repaired.append(element)
            index += 1
            continue

        table = dict(element)
        caption = re.sub(r"\s+", " ", str(table.get("caption") or "")).strip()
        if marker not in caption:
            table["caption"] = f"{caption} {marker}".strip()

        # A VLM can append the same visible unit to nearby prose as well as
        # emit it after the table. Remove only an exact terminal duplicate.
        for previous_index in range(len(repaired) - 1, max(-1, len(repaired) - 5), -1):
            previous = repaired[previous_index]
            if not isinstance(previous, dict):
                continue
            previous_type = previous.get("type", "text")
            if previous_type == "table" or previous_type == "figure" or str(previous_type).startswith("heading"):
                break
            if previous_type not in {"text", "footnote"}:
                continue
            content = str(previous.get("content") or "").rstrip()
            if not content.endswith(marker):
                continue
            content = content[:-len(marker)].rstrip()
            if content:
                repaired[previous_index] = {**previous, "content": content}
            else:
                repaired.pop(previous_index)
            break

        repaired.append(table)
        index += 2

    return repaired


def _reposition_figures_by_anchor(elements, anchors):
    """figure 를 IR 앵커의 직후 텍스트 앞으로 이동(그림만 이동, 텍스트·표 불변).
    anchors: [{'prev':앞문단, 'next':뒷문단}] 문서순. 매칭 실패 그림은 제자리."""
    def _nm(s):
        return re.sub(r"[^0-9A-Za-z가-힣]", "", (s or "")).lower()
    figs = [(i, e) for i, e in enumerate(elements) if e.get("type") == "figure"]
    if not figs or not anchors:
        return elements
    used = [False] * len(anchors)
    base = [e for e in elements if e.get("type") != "figure"]
    bnorm = [_nm((e.get("content") or "") + (e.get("caption") or "")) for e in base]
    inserts = []
    for orig_i, f in figs:
        fk = ""
        for k in ("caption", "content", "description"):
            v = _nm(f.get(k) or "")
            if len(v) >= 4:
                fk = v; break
        target = None
        for ai, a in enumerate(anchors):
            if used[ai]:
                continue
            ap = _nm(a.get("prev"))
            if fk and ap and (fk in ap or ap in fk):
                used[ai] = True
                an = _nm(a.get("next"))[:16]
                if an:
                    for bi, bn in enumerate(bnorm):
                        if an in bn:
                            target = bi; break
                break
        if target is None:          # 매칭 실패 → 원래 위치(앞 비-그림 요소 수) 유지
            target = sum(1 for e in elements[:orig_i] if e.get("type") != "figure")
        inserts.append((target, f))
    out = list(base)
    for target, f in sorted(inserts, key=lambda x: -x[0]):
        out.insert(max(0, min(target, len(out))), f)
    return out


def _correct_figure_captions(elements, anchors):
    """figure 캡션 교정: 직후/페이지내 텍스트 ↔ IR.next 로 앵커 매칭 후, IR.prev 가
    제목형·현재페이지부재·현재캡션과 상이일 때 IR.prev 로 교정. anchors:[{'prev','next'}]."""
    if not anchors:
        return elements
    def _nm(s):
        return re.sub(r"[^0-9A-Za-z가-힣]", "", (s or "")).lower()
    def _title_like(s):
        s = (s or "").strip()
        if not s or len(s) > 30 or s[0] in "□◦○●*-–·•":
            return False
        if re.match(r"^[①-⑳]", s) or s.endswith((".", "다", "음", "함", "임", "됨")):
            return False
        return True
    def _plen(a, b):
        m = 0
        for x, y in zip(a, b):
            if x != y:
                break
            m += 1
        return m
    # 페이지에 이미 존재하는 텍스트(본문 content + 모든 figure caption).
    present = [p for e in elements for p in (_nm(e.get("content")), _nm(e.get("caption"))) if p]
    txt = lambda x: str(x.get("type", "")).startswith(("text", "heading")) and (x.get("content") or "").strip()
    for k, e in enumerate(elements):
        if e.get("type") != "figure":
            continue
        foll = next((elements[t].get("content") for t in range(k + 1, len(elements)) if txt(elements[t])), "")
        fn = _nm(foll)
        match = None
        for a in anchors:
            an = _nm(a.get("next"))
            if len(an) >= 6 and fn and _plen(fn, an) >= 6:
                match = a; break
        if match is None:                       # 직후 불일치(위치까지 어긋난 그림) → 페이지 내 탐색
            for a in anchors:
                an = _nm(a.get("next"))
                if len(an) >= 6 and any(_plen(_nm(x.get("content")), an) >= 6 for x in elements if x is not e and txt(x)):
                    match = a; break
        if match is None:
            continue
        prev = (match.get("prev") or "").strip()
        if not _title_like(prev):
            continue
        # If the anchor's "next" is already a heading on this page, the
        # anchor likely spans a page/section boundary. In that case "prev"
        # belongs to the preceding section/figure and must not be copied as
        # this figure's caption.
        nn = _nm(match.get("next"))
        if nn and any(
            str(x.get("type", "")).startswith("heading") and _plen(_nm(x.get("content")), nn) >= 6
            for x in elements if x is not e
        ):
            continue
        np = _nm(prev)
        if any(np in p or p in np for p in present):
            continue                            # 진짜 제목이 이미 페이지에 있음 → 분리 아님
        cap = (e.get("caption") or "").strip()
        if cap and (_nm(cap) in np or np in _nm(cap)):
            continue                            # 이미 일치
        desc = e.get("description") or ""
        if cap and desc.startswith(cap):        # 설명 첫 구절이 옛 캡션을 반복하면 함께 교정
            e["description"] = prev + desc[len(cap):]
        e["caption"] = prev
    return elements


def _vlm_finish_reason(response):
    """choices[0].finish_reason 안전 추출(백엔드마다 없을 수 있음)."""
    try:
        return getattr(response.choices[0], "finish_reason", None)
    except Exception:
        return None


def _is_runaway_repeat(text, min_repeats=16, max_unit=512, tail=8192):
    """출력 꼬리가 같은 조각의 연속 반복(주기 ≤max_unit)으로 끝나면 True(반복 생성 지문).
    단일 반복뿐 아니라 A,B,A,B 같은 순환도 'AB' 단위로 잡는다(주기성→위상 무관).
    정상 문서·표는 동일 조각을 십수 번 연속 반복하며 끝나지 않으므로 고정밀."""
    if not text:
        return False
    t = text[-tail:].rstrip()
    n = len(t)
    if n < 32:
        return False
    upper = min(max_unit, n // min_repeats)
    for unit in range(2, upper + 1):
        seg = t[n - unit:]
        if not seg.strip():
            continue
        reps, pos = 1, n - 2 * unit
        while pos >= 0 and t[pos:pos + unit] == seg:
            reps += 1
            pos -= unit
        if reps >= min_repeats:
            return True
    return False
# HWP/HWPX 렌더러로 rhwp 사용(기본 ON). rhwp 미설치/렌더 실패 시 LibreOffice 로 자동 폴백.
USE_RHWP = os.environ.get("USE_RHWP", "1") == "1"
_RHWP_MOD = None
_RHWP_INIT_LOCK = threading.Lock()
_RHWP_FT_CANDIDATES = ["/lib/x86_64-linux-gnu/libfreetype.so.6", "/usr/lib/x86_64-linux-gnu/libfreetype.so.6"]

def _ensure_min_fontconfig():
    # rhwp(Skia)는 글자마다 시스템 폰트 전체를 폴백 스캔하므로 폰트가 수천 개면 페이지당 수십 초가 걸린다.
    # 한글 + 기본 Latin 폰트만 담은 최소 fontconfig 를 만들어 FONTCONFIG_FILE 로 지정(렌더 속도 ~100배).
    # RHWP_FONTCONFIG 로 직접 지정하거나 FONTCONFIG_FILE 가 이미 있으면 그대로 둔다.
    if os.environ.get("FONTCONFIG_FILE"):
        return
    cfg = os.environ.get("RHWP_FONTCONFIG")
    if cfg and os.path.exists(cfg):
        os.environ["FONTCONFIG_FILE"] = cfg
        return
    import glob, tempfile
    pats = ["HCR*", "NotoSansCJK*", "NotoSerifCJK*", "NanumGothic*", "NanumMyeongjo*",
            "malgun*", "*Batang*", "*Gulim*", "DejaVuSans*", "DejaVuSerif*",
            "LiberationSans*", "LiberationSerif*"]
    dirs = [os.path.expanduser("~/.local/share/fonts"),
            "/usr/share/fonts/truetype/nanum", "/usr/share/fonts/opentype/noto",
            "/usr/share/fonts/truetype/noto", "/usr/share/fonts/truetype/dejavu",
            "/usr/share/fonts/truetype/liberation", "/usr/share/fonts", "/usr/local/share/fonts"]
    found = []
    for d in dirs:
        if not os.path.isdir(d):
            continue
        for p in pats:
            found += glob.glob(os.path.join(d, "**", p), recursive=True)
    found = sorted({f for f in found if os.path.isfile(f)})[:80]
    if not found:
        return
    base = os.path.join(tempfile.gettempdir(), "rhwp_fonts")
    os.makedirs(base, exist_ok=True)
    for f in found:
        ln = os.path.join(base, os.path.basename(f))
        try:
            if not os.path.lexists(ln):
                os.symlink(f, ln)
        except OSError:
            pass
    cachedir = os.path.join(tempfile.gettempdir(), "rhwp_fc_cache")
    os.makedirs(cachedir, exist_ok=True)
    conf = os.path.join(tempfile.gettempdir(), "rhwp_min_fonts.conf")
    with open(conf, "w", encoding="utf-8") as fp:
        fp.write('<?xml version="1.0"?>\n<!DOCTYPE fontconfig SYSTEM "fonts.dtd">\n'
                 '<fontconfig>\n  <dir>%s</dir>\n  <cachedir>%s</cachedir>\n</fontconfig>\n' % (base, cachedir))
    os.environ["FONTCONFIG_FILE"] = conf

def _load_rhwp():
    global _RHWP_MOD
    if _RHWP_MOD is None:
        with _RHWP_INIT_LOCK:
            if _RHWP_MOD is None:
                _ensure_min_fontconfig()
                import ctypes
                for _p in _RHWP_FT_CANDIDATES:
                    try:
                        ctypes.CDLL(_p, mode=ctypes.RTLD_GLOBAL)
                        break
                    except OSError:
                        continue
                # 참고: rhwp 코어는 레이아웃 진단(LAYOUT_OVERFLOW 등)을 stderr 로 출력한다.
                # stderr 는 앱 로거와 공유라 인프로세스에서 안전하게 끌 수 없으므로(스레드 공유+로그 손실),
                # 노이즈가 거슬리면 프로세스 stderr 를 배포 레벨에서 리다이렉트할 것.
                import rhwp as _r
                _RHWP_MOD = _r
    return _RHWP_MOD

import fitz
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from bs4 import BeautifulSoup, NavigableString

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


class HWPTextExtractor:

    @staticmethod
    def text_preprocessing(text, threshold=0.5):
        if not text: return ""
        special_chars = re.sub(r'[가-힣a-zA-Z0-9\s]', '', text)
        if len(text) > 0 and (len(special_chars) / len(text)) >= threshold: return ""
        
        text = unicodedata.normalize("NFC", text)
        text = re.sub(r'[\uE000-\uF8FF]', "", text)
        text = re.sub("\u00A0", " ", text)
        text = re.sub(r'\n+', '\n', text)
        text = re.sub(r' +', ' ', text)
        return text.strip()

    @classmethod
    def hwp_process_children(cls, tag):
        content = ""
        for child in tag.children:
            if isinstance(child, NavigableString):
                content += str(child)
            elif child.name == 'table':
                content += cls.hwp_parse_table(child)
            elif child.name == 'p':
                inner_html = cls.hwp_process_children(child)
                content += f"<p>{inner_html}</p>"
            elif child.name == 'img':
                src = child.get('src', '')
                content += f'<img src="{src}" alt="이미지"/>'
            elif hasattr(child, 'prettify'):
                content += child.prettify(formatter="html5")
        return content.strip()

    @classmethod
    def hwp_parse_table(cls, table_tag):
        table_html = "<table>"
        for tr in table_tag.find_all('tr', recursive=False):
            table_html += "<tr>"
            for td in tr.find_all('td', recursive=False):
                col = td.get('colspan', 1)
                row = td.get('rowspan', 1)
                table_html += f'<td colspan="{col}" rowspan="{row}">'
                table_html += cls.hwp_process_children(td)
                table_html += "</td>"
            table_html += "</tr>"
        table_html += "</table>"
        for child in table_tag.children:
            if child.name == 'caption':
                table_html += f"<p>{cls.hwp_process_children(child)}</p>"
        return table_html

    @classmethod
    def parser_from_hwp_html(cls, file_path):
        temp_dir = tempfile.mkdtemp()
        pages = []
        try:
            subprocess.run(['hwp5html', '--output', temp_dir, file_path], check=True, timeout=60, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            xhtml_file_path = os.path.join(temp_dir, "index.xhtml")
            with open(xhtml_file_path, "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f.read(), 'html.parser').body

            for span_tag in soup.find_all('span'): span_tag.unwrap()
            
            all_pages = soup.find_all('div', class_='Page')
            if not all_pages: all_pages = [soup.body]
            
            for content_area in all_pages:
                current_page_text = []
                for element in content_area.children:
                    if element.name == 'table':
                        current_page_text.append(cls.hwp_parse_table(element).strip())
                    elif element.name == 'p':
                        if element.find(attrs={'class': 'GShapeObjectControl'}): continue 
                        table_in_p = element.find('table')
                        if table_in_p:
                            current_page_text.append(cls.hwp_parse_table(table_in_p).strip())
                        else:
                            processed_html = cls.hwp_process_children(element) + "<br>"
                            current_page_text.append(processed_html)
                    elif element.name == 'img':
                        src = element.get('src', "")
                        current_page_text.append(f"<img src='{src}' alt='이미지'/>")
                
                page_str = "<br>".join(current_page_text).replace('`', "'")
                page_str = re.sub(r"(<br> ){2,}", "<br>", page_str).strip()
                pages.append(re.sub(r"(<br>){2,}", "<br>", page_str).strip())

        except Exception as error:
            logger.warning(f"HTML fallback 파싱 에러: {error}")
        finally:
            shutil.rmtree(temp_dir)
        return pages

    @classmethod
    def get_text_from_hwpx(cls, element):
        text = []
        sentence_tags = element.findall('./{*}run')
        for sentence_tag in sentence_tags:
            for child in sentence_tag:
                child_tag = child.tag.split('}')[-1]
                if child_tag == "t": text.append(child.text)
                elif child_tag == "fwSpace": text.append(" ")
                elif child_tag in ["pic", "rect", "poly", "ellipse"]:
                    obj_id = child.get("id", "")
                    if child.find("./{*}drawText") is not None:
                        obj_text = ""
                        for t_tag in child.findall("./{*}drawText/{*}subList/{*}p"):
                            obj_text += cls.get_text_from_hwpx(t_tag)
                        text.append(f"<{child_tag} id='{obj_id}'> {obj_text} </{child_tag}>")
                    else:
                        text.append(f"<{child_tag} id='{obj_id}' alt='이미지'/>")
        text = [t for t in text if t is not None]
        return "".join(text).replace('`', "'")

    @classmethod
    def get_table_from_hwpx(cls, element):
        table = "<table>"
        rows = element.findall('./{*}tr')
        for row in rows:
            table += "<tr>"
            for cell in row.findall('./{*}tc'):
                span = cell.find('./{*}cellSpan')
                colspan = span.get('colSpan', '1') if span is not None else '1'
                rowspan = span.get('rowSpan', '1') if span is not None else '1'
                table_in_cell = cell.findall('./{*}subList/{*}p/{*}run/{*}tbl')
                table += f"<td colspan='{colspan}' rowspan='{rowspan}'>"
                if table_in_cell:
                    for inner_table in table_in_cell: table += cls.get_table_from_hwpx(inner_table)
                else:
                    for cell_text_tag in cell.findall('./{*}subList/{*}p'):
                        table += cls.get_text_from_hwpx(cell_text_tag)
                table += "</td>"
            table += "</tr>"
        return table + "</table>"

    @classmethod
    def parser_from_hwpx_xml(cls, file_path):
        pages = []
        try:
            all_p_elements = []
            with zipfile.ZipFile(file_path, 'r') as zf:
                section_files = sorted(
                    [n for n in zf.namelist() if re.match(r'Contents/section\d+\.xml', n)],
                    key=lambda x: int(re.search(r'\d+', x).group())
                )
                for sec_file in section_files:
                    xml_content = zf.read(sec_file).decode('utf-8')
                    root = ET.fromstring(xml_content)
                    all_p_elements.extend(root.findall('./{*}p'))

            current_page = []
            for p_element in all_p_elements:
                has_break = any(k.endswith('pageBreak') and v == "1"
                                for k, v in p_element.attrib.items())
                if has_break and current_page:
                    pages.append("<br>".join(current_page))
                    current_page = []

                table = p_element.find('./{*}run/{*}tbl')
                if table is not None:
                    current_page.append(cls.get_table_from_hwpx(table))
                else:
                    text = cls.get_text_from_hwpx(p_element)
                    if text: current_page.append(text)

            if current_page: pages.append("<br>".join(current_page))
        except Exception as e:
            logger.error(f"HWPX 파싱 에러: {e}")
        return [re.sub(r"(<br>){2,}", "<br>", p).strip() for p in pages]

    @classmethod
    def _process_hwp_child(cls, child, text, include_table=True):
        if child.tag == "Text":
            text.append(child.text)
        elif child.tag == "GShapeObjectControl":
            shp_comp = child.find("./ShapeComponent")
            if shp_comp is not None:
                text_tag = shp_comp.find("TextboxParagraphList")
                gso_type = shp_comp.get("chid")
                if gso_type == "$pic":
                    picture_id = child.get("instance-id", "")
                    if text_tag is not None:
                        pic_text = "".join(cls.get_text_from_hwp(p) for p in text_tag.findall("./Paragraph"))
                        text.append(f"<img id='{picture_id}'> {pic_text} </img>")
                    else:
                        text.append(f"<img id='{picture_id}' alt='이미지'/>")
                elif gso_type == "$rec":
                    rect_id = child.get("instance-id", "")
                    if text_tag is not None:
                        rect_text = "".join(cls.get_text_from_hwp(p) for p in text_tag.findall("./Paragraph"))
                        text.append(f"<rect id='{rect_id}'> {rect_text} </rect>")
        elif child.tag == "TableControl" and include_table:
            table_html = cls.get_table_from_hwp(child)
            if table_html: text.append(table_html)

    @classmethod
    def get_text_from_hwp(cls, element):
        text = []
        for sentence_tag in element.findall('./LineSeg'):
            fieldclickheres = sentence_tag.findall('./FieldClickHere')
            if fieldclickheres:
                for fch in fieldclickheres:
                    inner_fchs = fch.findall('./FieldClickHere')
                    if inner_fchs:
                        for inner_fch in inner_fchs:
                            for child in inner_fch:
                                cls._process_hwp_child(child, text)
                    else:
                        for child in fch:
                            cls._process_hwp_child(child, text)
            else:
                for child in sentence_tag:
                    cls._process_hwp_child(child, text)
        text = [t for t in text if t is not None]
        return "".join(text)

    @classmethod
    def _get_text_from_hwp_for_table(cls, element):
        text = []
        for sentence_tag in element.findall('./LineSeg'):
            fieldclickheres = sentence_tag.findall('./FieldClickHere')
            if fieldclickheres:
                for fch in fieldclickheres:
                    inner_fchs = fch.findall('./FieldClickHere')
                    if inner_fchs:
                        for inner_fch in inner_fchs:
                            for child in inner_fch:
                                cls._process_hwp_child(child, text, include_table=False)
                    else:
                        for child in fch:
                            cls._process_hwp_child(child, text, include_table=False)
            else:
                for child in sentence_tag:
                    cls._process_hwp_child(child, text, include_table=False)
        text = [t for t in text if t is not None]
        return "".join(text)

    @classmethod
    def get_table_from_hwp(cls, element):
        table_body = element.find('./TableBody')
        if table_body is None: return ""
        table = "<table>"
        for row in table_body.findall('./TableRow'):
            table += "<tr>"
            for cell in row.findall('./TableCell'):
                colspan, rowspan = cell.get('colspan', '1'), cell.get('rowspan', '1')
                table += f"<td colspan='{colspan}' rowspan='{rowspan}'>"
                table_in_cell = cell.findall('./Paragraph/LineSeg/TableControl')
                table_in_cell_ex = cell.findall('./ColumnSet/Paragraph/LineSeg/TableControl')
                if table_in_cell:
                    for it in table_in_cell: table += cls.get_table_from_hwp(it)
                elif table_in_cell_ex:
                    for it in table_in_cell_ex: table += cls.get_table_from_hwp(it)
                else:
                    columnsets = cell.findall('./ColumnSet')
                    if columnsets:
                        for cs in columnsets:
                            for p in cs.findall('./Paragraph'):
                                table += cls._get_text_from_hwp_for_table(p)
                    else:
                        for p in cell.findall('./Paragraph'):
                            table += cls._get_text_from_hwp_for_table(p)
                table += "</td>"
            table += "</tr>"
        return table + "</table>"

    @classmethod
    def parser_from_hwp_xml(cls, file_path):
        pre, _ = os.path.splitext(file_path)
        temp_xml = pre + f"_{uuid.uuid4().hex[:6]}_temp.xml"
        pages = []
        try:
            with open(temp_xml, 'w', encoding='utf-8') as f:
                subprocess.run(['hwp5proc', 'xml', file_path], stdout=f, stderr=subprocess.DEVNULL, check=True)
            
            root = ET.parse(temp_xml).getroot()
            current_page = []
            
            for p_element in root.findall('.//SectionDef/ColumnSet/Paragraph'):
                has_break = any(k.endswith('new-page') and v == "1"
                                for k, v in p_element.attrib.items())
                if has_break and current_page:
                    pages.append("<br>".join(current_page))
                    current_page = []
                text = cls.get_text_from_hwp(p_element)
                if text: current_page.append(text)
            
            if current_page: pages.append("<br>".join(current_page))
        except Exception as e:
            raise e
        finally:
            if os.path.exists(temp_xml): os.remove(temp_xml)
            
        return [re.sub(r"(<br>){2,}", "<br>", p).strip() for p in pages]

    @staticmethod
    def _clean_native(text):
        if not text: return ""
        text = unicodedata.normalize("NFC", text)
        text = re.sub(r'[-]', "", text)
        text = re.sub(" ", " ", text)
        text = re.sub(r'[ \t]+', ' ', text)
        text = re.sub(r'\n{2,}', '\n', text)
        return text.strip()

    @classmethod
    def extract_pages(cls, file_path):
        ext = os.path.splitext(file_path)[1].lower()

        try:
            from hwp_extract import extract as _native_extract
            text, status = _native_extract(file_path)
            if status == "ok":
                cleaned = cls._clean_native(text)
                if cleaned:
                    return [cleaned]
            else:
                logger.info(f"네이티브 추출 비정상({status}), 폴백 시도: {os.path.basename(file_path)}")
        except Exception as e:
            logger.warning(f"네이티브 추출 예외, 폴백: {e}")

        pages = []
        if ext == ".hwpx":
            pages = cls.parser_from_hwpx_xml(file_path)
        elif ext == ".hwp":
            pages = cls.parser_from_hwp_html(file_path)
            if not pages:
                try:
                    pages = cls.parser_from_hwp_xml(file_path)
                except Exception:
                    pages = []

        return [cls.text_preprocessing(p) for p in pages if cls.text_preprocessing(p)]


def _salvage_truncated_json(text: str):
    """max_tokens 로 잘려 파싱 불가한 JSON 에서 elements 배열의 완결 객체만 복구.
    닫히지 않은 마지막 객체는 버리고 그 앞까지의 요소로 유효 dict 를 재구성(불가 시 None)."""
    m = re.search(r'"elements"\s*:\s*\[', text)
    if not m:
        return None
    i = m.end()
    depth, in_str, esc, start, objs = 0, False, False, None, []
    while i < len(text):
        c = text[i]
        if in_str:
            if esc:
                esc = False
            elif c == '\\':
                esc = True
            elif c == '"':
                in_str = False
        elif c == '"':
            in_str = True
        elif c == '{':
            if depth == 0:
                start = i
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0 and start is not None:
                frag = text[start:i + 1]
                try:
                    objs.append(json.loads(frag))
                except Exception:
                    pass
                start = None
        elif c == ']' and depth == 0:
            break
        i += 1
    return {"elements": objs} if objs else None


def _sanitize_json_strings(text: str) -> str:
    _CTRL_ESCAPES = {'\n': '\\n', '\r': '\\r', '\t': '\\t'}
    _VALID_ESCAPES = {'"', '\\', '/', 'b', 'f', 'n', 'r', 't', 'u'}
    result = []
    in_string = False
    i = 0
    while i < len(text):
        c = text[i]
        if in_string:
            if c == '\\':
                i += 1
                if i < len(text):
                    nc = text[i]
                    if nc in _VALID_ESCAPES:
                        result.append('\\')
                        result.append(nc)
                    else:
                        result.append('\\\\')
                        result.append(nc)
                else:
                    result.append('\\\\')
            elif c == '"':
                result.append(c)
                in_string = False
            elif ord(c) < 0x20 or c == '\x7f':
                result.append(_CTRL_ESCAPES.get(c, ''))
            else:
                result.append(c)
        else:
            if c == '"':
                in_string = True
            result.append(c)
        i += 1
    return ''.join(result)


class VLMProcessor:

    @classmethod
    def encode_image(cls, image_path, max_width=None):
        max_width = max_width or VLM_IMG_MAXW
        try:
            from PIL import Image
            import io
            with Image.open(image_path) as img:
                if img.width > max_width:
                    ratio = max_width / img.width
                    new_size = (max_width, int(img.height * ratio))
                    img = img.resize(new_size, Image.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                return base64.b64encode(buf.getvalue()).decode('utf-8')
        except ImportError:
            with open(image_path, "rb") as f:
                return base64.b64encode(f.read()).decode('utf-8')

    @classmethod
    def extract_structure(cls, txt_path, img_path=None, api_key=None, output_format="json", model_name="Qwen/Qwen3-VL-30B-A3B-Instruct"):
        # 반환: (구조 문자열 또는 None, retryable). retryable=False 면 동일 입력 재시도가 무의미
        # (반복 루프·타임아웃·유효성 실패) → 호출자의 최종재시도 패스에서 제외.
        if not api_key or "여기에_" in api_key:
            return None, False

        try:
            from openai import OpenAI
            try:
                from openai import APITimeoutError as _APITimeout
            except Exception:
                class _APITimeout(Exception):
                    pass
            class _StreamDeadline(Exception):
                pass
            base_url = os.environ.get("VLM_BASE_URL", "http://localhost:8000/v1")
            client = OpenAI(base_url=base_url, api_key=api_key or "EMPTY", timeout=VLM_TIMEOUT, max_retries=0)
        except Exception as e:
            logger.error(f"VLM 클라이언트 초기화 오류: {e}")
            return None, False

        if txt_path and os.path.exists(txt_path):
            with open(txt_path, "r", encoding="utf-8") as f:
                extracted_text = f.read()
        else:
            extracted_text = ""        # 텍스트레이어 없음(예: 다단 열크롭) → 이미지만으로 추출, coverage 게이트 비활성

        if output_format.lower() == "xml":
            schema_instruction = """[FORMAT INSTRUCTION: XML ONLY]
CRITICAL: You MUST output ONLY valid XML. Do NOT output JSON format under any circumstances.
You MUST extract the contents of the page strictly in the sequential order they appear visually.
Use the following exact XML template:

<document>
    <page_number>int</page_number>
    <elements>
        <element type="heading_1">Top level heading (e.g., Ⅰ., 1.)</element>
        <element type="heading_2">Second level heading (e.g., 가., 1.1)</element>
        <element type="heading_3">Third level heading (e.g., 1), 가)</element>
        <element type="toc_entry">TOC item — content format: "title::page_number"</element>
        <element type="text">General paragraph text</element>
        <element type="table" caption="Table title or caption if available">Markdown table content</element>
        <element type="figure" caption="Figure title or caption if available" description="Detailed visual description in the document body language (Korean for a Korean doc) even if the figure labels are foreign: for charts/graphs describe what it shows, trends, key observations; for photographs/UI captures describe what is depicted. Never empty for a figure.">For charts/graphs: extracted visible text data here (axis labels, tick values, etc.). Empty for photographs.</element>
        <element type="footnote">Footnote or reference text at the bottom</element>
    </elements>
</document>

 Rules for XML:
 1. Reading order:
    - Single-column: read top to bottom.
    - Continuous newspaper/report columns: emit the shared page header first, then read the LEFT body column fully (top to bottom), then the RIGHT body column (top to bottom), and emit the shared footer last. Keep each heading immediately followed by its body. If the bottom-left sentence continues at the top-right, emit that top-right fragment immediately after the bottom-left fragment before the next complete heading. Never use raw-text order to relocate a visual block.
    - Bounded peer cards/tiles in a grid: read each row left to right, then move to the next row. Keep each card's title and bullets together; never move bullets into a neighboring card.
    - Use the layout image to determine the number of columns.
2. Repeat <element> tags as needed — multiple elements of the same type are allowed.
3. If an element does not exist, simply do not include that type.
4. [TABLE — CRITICAL] Any grid of data with rows and columns MUST be output as HTML table. Rules:
   a. Use ONLY <table>, <tr>, <td> tags (and <br> only inside a cell). NEVER use <th>. You MAY use colspan/rowspan (e.g. <td colspan="2">Header</td>).
   b. Always use the layout IMAGE to detect tables visually — raw text extraction often loses column alignment and merges cells.
   b-1. Equations, formulas, and variable-definition lists without visible cell borders are text, NOT tables.
   c. Include ALL rows: header rows, subheader rows, data rows, and any title row that spans all columns.
   d. Each row must have the same number of cells (accounting for colspan/rowspan). Empty cells must be included as empty <td></td>.
   e. If a table has a multi-row header (e.g. a "Training Datasets" row spanning 6 columns, then "Instruction / Alignment" row spanning 3 each), reproduce all those rows with correct colspan.
   f. NEVER output table data as flat text or bullet points — always use <table> structure.
5. Type rules:
   - heading_1/2/3: For actual section titles and prominent standalone headings. Use for: section titles in body text (e.g. "Introduction", "1.1 Methods"); on slide/presentation pages, EACH visually prominent section label and numbered step title MUST be a SEPARATE heading element — do NOT merge section labels with their body text (e.g. "Our Purpose" → heading_2, "Making AI Beneficial" → text; "01 - Find Resources" → heading_2, description → text); section labels that introduce a data block on a report/survey page (e.g. "Education Level", "Profession"). Do NOT use for: TOC entries; run-in paragraph starters — if a "heading" would contain a period in the MIDDLE followed by more text (e.g. "Business characteristics. Business size was determined by...", "Ablation on the SFT base models. We compare...", "Filtered task names. We present task names..."), it is a paragraph, NOT a heading — output as text with the opening label in **bold**; bold short labels ending with a period followed by body text (e.g. "Base model.", "Depthwise scaling." — use text element); lettered sub-items (a., b., c., d.) that introduce paragraphs.
     IMPORTANT — PRESERVE the leading numbering/marker at the START of the heading content exactly as printed (e.g. "제1장", "제2조", "①", "1.", "가.", "(1)", "□", "○", "Ⅰ.") — do NOT strip it. This marker encodes the heading's hierarchy level and is required for downstream structuring.
   - toc_entry: ONLY for items on a Table of Contents page. Format content as "title::page_number".
   - figure: ONLY for actual embedded images, photographs, charts, or diagrams — NOT for text formatted with symbols like ▴, ●, ○, etc. Put the caption in the caption attribute. Always fill the description attribute with a detailed visual description: for charts/graphs describe what the chart shows, the overall trend, and key observations; for photographs describe what is depicted. For charts/graphs/plots: scan the ENTIRE chart image and transcribe the legible title, legend entries, axis labels, major endpoint/tick values, explicitly printed data labels, and annotations needed to interpret it. Do NOT enumerate unlabeled plotted samples, infer intermediate values, or repeat dense micro-labels. List retained items once in visual reading order. A screenshot dominated by readable forms, messages, Q&A cards, or document text is structured as separate text/table elements for its visible regions; do not duplicate that text in a figure element. For photographs or decorative images with no data, leave the element body empty. The description MUST be written in the document body language (Korean for a Korean document) EVEN IF the figure's labels are in another language; for other screenshots/UI captures put non-structured on-screen text in the body but still write the description — NEVER leave description empty for a figure. Match detail to type: charts → chart type + concrete trend shape (계단식/선형/급증·급락) + peak position + series comparison; diagrams → main nodes and their flow/relationships; photos/maps/renderings → key subject in 1–2 sentences; screenshots → one sentence on purpose (do not repeat the body data). ANTI-HALLUCINATION: if no numbers are printed on the axes/legend, do NOT invent them — state "축에 수치 미표기" and describe only the qualitative shape.
   - A flowchart or architecture diagram made of boxes, arrows, actors, and free-positioned labels is a figure, NOT a table unless a genuine rectangular cell grid is visible. Emit one figure per visually distinct diagram, transcribe each node and connector label once in figure content, and do not repeat those labels in separate text elements.
   - footnote: ONLY for superscript-style reference notes at the very bottom margin (e.g. *, ①, ※ markers). Do NOT use for regular body paragraphs.
   - table caption: Put the table title in the caption attribute, NOT as a separate heading element.
 6. CRITICAL — NO DUPLICATION: Each piece of content must appear exactly ONCE. Never output the same text in multiple elements. Emit an identical figure element twice only when two distinct visual instances are actually present on the page."""
        else:
            schema_instruction = """[FORMAT INSTRUCTION: JSON ONLY]
CRITICAL: Output ONLY valid JSON.
You MUST extract the contents of the page strictly in the sequential order they appear visually.
Use the following JSON schema:

{
    "page_number": int,
    "elements": [
        {
            "type": "heading_1 | heading_2 | heading_3 | toc_entry | text | table | figure | footnote",
            "content": "Text content or HTML table. For toc_entry: 'title::page_number'. For figure (chart/graph): extracted visible text data (axis labels, tick values, legend, percentages). Empty for photographs.",
            "caption": "Optional: table title or figure caption if present",
            "description": "For figure only: detailed visual description in the DOCUMENT BODY LANGUAGE (Korean for a Korean document) even if the figure's labels are in another language — for charts/graphs describe what it shows, trends, and key observations; for photographs/UI captures describe what is depicted. Never empty for a figure. Leave empty for non-figure elements."
        }
    ]
}

 Rules for JSON:
 1. Reading order:
    - Single-column: read top to bottom.
    - Continuous newspaper/report columns: emit the shared page header first, then read the LEFT body column fully (top to bottom), then the RIGHT body column (top to bottom), and emit the shared footer last. Keep each heading immediately followed by its body. If the bottom-left sentence continues at the top-right, emit that top-right fragment immediately after the bottom-left fragment before the next complete heading. Never use raw-text order to relocate a visual block.
    - Bounded peer cards/tiles in a grid: read each row left to right, then move to the next row. Keep each card's title and bullets together; never move bullets into a neighboring card.
    - Use the layout image to determine the number of columns.
2. "type" MUST be one of the specified enum values:
   - heading_1/2/3: For actual section titles and prominent standalone headings. Use for: section titles in body text (e.g. "Introduction", "1.1 Methods"); on slide/presentation pages, EACH visually prominent section label and numbered step title MUST be a SEPARATE heading element — do NOT merge section labels with their body text (e.g. on a slide with "Our Purpose\nMaking AI Beneficial", "Our Purpose" is heading_2 and "Making AI Beneficial" is text; on a slide with "01 - Find Resources\nStart by searching...", "01 - Find Resources" is heading_2 and "Start by searching..." is text); section labels that introduce a data block on a report/survey page (e.g. "Education Level", "Profession"). Do NOT use for: TOC entries, bold inline labels within a paragraph sentence (e.g. "Base model. We trained..." where text continues on the same line).
     IMPORTANT — PRESERVE the leading numbering/marker at the START of the heading "content" exactly as printed (e.g. "제1장", "제2조", "①", "1.", "가.", "(1)", "□", "○", "Ⅰ.") — do NOT strip it. This marker encodes the heading's hierarchy level and is required for downstream structuring.
   - toc_entry: ONLY for items listed on a Table of Contents page. Set content to "title::page_number".
   - text: General body paragraphs, including lists, bullet points, and any text content that is NOT a heading. Use **bold** for bold text within paragraphs.
   - table: [CRITICAL] Any grid of data MUST be output as HTML in "content". Use ONLY <table>, <tr>, <td> (and <br> only inside a cell) — NEVER <th>. You MAY use colspan/rowspan. Always use the layout IMAGE to detect tables visually (raw text loses column structure). Include ALL rows: title rows, multi-row headers (with correct colspan), data rows, totals rows. Empty cells must be included as empty <td></td>. Preserve side-by-side repeated column groups from the image; never stack them vertically into a narrower table. NEVER output table data as flat text. Put table title in "caption". Equations, formulas, and variable-definition lists without visible cell borders are text, NOT tables.
   - figure: ONLY for actual embedded images, photographs, charts, or diagrams that are visual/graphical content — NOT for text formatted with symbols like ▴, ●, ○, ☞, etc. Put caption in "caption". For charts/graphs/plots: scan the ENTIRE chart and put the legible title, legend entries, axis labels, major endpoint/tick values, explicitly printed data labels, and annotations needed to interpret it into "content". Do NOT enumerate unlabeled plotted samples, infer intermediate values, or repeat dense micro-labels. List retained items once in visual reading order. A screenshot dominated by readable forms, messages, Q&A cards, or document text MUST be structured as separate text/table elements for its visible regions; do not duplicate that text in a figure element. For photographs or decorative images with no data, leave "content" empty. ALWAYS fill "description" with a detailed visual description, written in the DOCUMENT BODY LANGUAGE (Korean for a Korean document) EVEN IF the figure's own labels/axes/legend are in another language — transcribe the foreign labels verbatim into "content", but NARRATE "description" in the document language. Write "description" with detail matched to the figure TYPE:
     • Chart/graph: name the chart type (막대/선/원), then the CONCRETE shape of the trend (계단식/선형/급증·급락/평탄), WHERE the peak/trough sits, and compare the series/legend entries by name. Include printed numbers (peaks, totals, legend values) ONLY if they are actually printed on the figure.
     • Diagram/flowchart/structure: enumerate the main nodes/components and the relationships, flow direction, or contract/transaction names connecting them.
     • Photograph/aerial/map/rendering: 1–2 sentences on the key subject and composition, then STOP. Describe only what is in the frame; if you read a place-name, sign, or caption in the photo, do NOT append background facts, history, or guesses about it.
     • Screenshot/UI capture: ONE sentence on the screen's purpose — the on-screen data already goes in "content", so do NOT repeat it in the description.
     ANTI-HALLUCINATION: if the axes/legend/data-labels carry NO printed numbers, do NOT invent any — state "축에 수치 미표기" and describe only the qualitative shape. NEVER leave "description" empty for a figure. (한국어 차트 예: "분기별 매출 막대그래프. 9~12월 계단식 상승, 12월 1.2억으로 정점, A계열이 B계열을 상회.")
   - A flowchart or architecture diagram made of boxes, arrows, actors, and free-positioned labels is a figure, NOT a table unless a genuine rectangular cell grid is visible. Emit one figure per visually distinct diagram, transcribe each node and connector label once in figure content, and do not repeat those labels in separate text elements.
   - footnote: ONLY for superscript-style reference notes (e.g., *, ①, ※ markers at the very bottom margin of the page). Do NOT use for regular body text that happens to appear at the bottom.
 3. CRITICAL — NO DUPLICATION: Each piece of text content must appear exactly ONCE in the elements array. Never output the same content in multiple elements. Emit an identical figure element twice only when two distinct visual instances are actually present on the page.
4. If the page is empty, return an empty "elements" array."""

        system_prompt = f"""You are an expert document parsing AI.
Your task is to convert the provided document text (and layout image if available) into a structured format.

[CRITICAL OUTPUT RULES]
- Output ONLY the raw {output_format.upper()} — no explanation, no commentary, no markdown code fences.
- Your entire response must be valid, parseable {output_format.upper()}. If it cannot be parsed, it is wrong.
- Do not truncate or abbreviate content. Every element must be complete.
- Never use unescaped special characters inside JSON strings (e.g. use \\n for newlines, escape double quotes).
- For tables: every Markdown table row must have the same number of columns as the header row.
- LANGUAGE: Write every natural-language string you generate — figure "description", any summary, and captions you compose — in the SAME language as the document body. For a Korean document the description MUST be Korean; do NOT answer in English. (Transcribed content keeps the original language as printed.)
- FIGURE description LENGTH (hard limit): keep every figure "description" to AT MOST ~300 characters (about 2–4 sentences). Describe ONLY what is visibly in the frame. NEVER repeat a phrase, and do NOT drift into background knowledge, history, geography, or speculation triggered by a place-name, sign, or label you read in the image — if you catch yourself restating or elaborating beyond what is visible, STOP immediately.
- READING ORDER: emit any full-width page header first and footer last. For continuous newspaper/report body columns, finish the ENTIRE left column top-to-bottom before starting the right column, keeping every heading with its following body. If a sentence is cut at the bottom-left and resumes at the top-right, place the continuation immediately after that fragment before any new right-column heading. For bounded peer cards or tiles, read each row left-to-right and keep each card's title and body together. Use the image, never raw-text order, to place blocks.
- Transcribe ONLY what is visibly printed on THIS page. Do NOT repeat, re-emit, or duplicate any block of text — each piece appears exactly once.
- TABLE OF CONTENTS — STRUCTURE FIDELITY (critical):
  (1) If the page is a table of contents (e.g. "목차", "목 차", "Contents", dotted leader lines, right-aligned page numbers), preserve the page as TOC structure.
  (2) A TOC row with a visible page number MUST be one toc_entry whose content is exactly "title::page_number". The title is the printed text before the leader dots; the page_number is the visible right-aligned number. Do NOT include leader dots.
  (3) Section/group labels on a TOC page that do NOT have their own visible page number (e.g. "I. 시스템 개요 및 사용자 등록") are headings, not toc_entry. Do NOT copy a child row's page number onto a parent/group label.
  (4) Keep every visible TOC title and page number, including lower-page entries near the bottom. Never summarize the TOC.
- TABLES — STRUCTURE FIDELITY (critical):
  (1) ONE <table> = exactly ONE continuous bordered grid. If two or more grids sit SIDE BY SIDE (left/right), or are STACKED TOP/BOTTOM with a separate border/gap and different column structures, output EACH as its OWN separate <table> element. A single outer grid with repeated side-by-side header groups remains one full-width table. In either case, NEVER flatten side-by-side groups by stacking them vertically into one narrow table.
  (2) Count the column count ONCE from the vertical border lines; EVERY <tr> must sum (counting colspan) to exactly that count — pad missing cells with <td></td>, never invent or drop columns.
  (3) NO cell-boundary bleed: assign each printed value to EXACTLY ONE cell (the row and column whose borders bracket it). Two adjacent cells MUST NOT contain copied text unless it is genuinely printed twice. In particular, never append the next row's label to the preceding row's final cell. An empty cell is <td></td> — never fill it by copying a neighbour.
  (4) MERGED cells: a cell that visually spans N columns → colspan="N"; spanning N rows → rowspan="N". Reproduce multi-row headers exactly with colspan/rowspan; do not flatten or duplicate the spanned text into each cell.
  (5) NESTED table (a table inside a cell): keep it as a nested <table> INSIDE that <td>; do not splice its rows into the outer table.
  (6) HEADER WIDTH = BODY WIDTH: the header row's cells (counting colspan) MUST sum to the SAME column count as the body rows. When the body's left side is split into sub-columns by a rowspan category — e.g. body rows are [category | sub-item | content] (3 columns) where "category" spans several rows — the header label above must carry colspan to cover ALL the columns it sits over. Example: body is 3 columns [위험요인 | 세부요인 | 심사내용] → header MUST be <tr><td colspan="2">위험요인</td><td>심사내용</td></tr>, NEVER a 2-cell header over a 3-column body. A header narrower than the body is WRONG.
  (7) KEY-VALUE form tables: if one row is a single label-value pair while another row places TWO label-value pairs side by side, the table width is the WIDER row; on the 1-pair row give the value a colspan to fill the remaining columns (e.g. <td>항목</td><td colspan="3">값</td>) — never leave a row shorter than the table width.
  (8) RECTANGULAR-GRID SELF-CHECK — do this for EVERY <table> before you emit it. Expand all colspan/rowspan into a 2-D grid and confirm it is a PERFECT RECTANGLE: every row, after expansion, MUST resolve to the IDENTICAL number of columns C. Procedure: (a) fix C from the body data row that has the MOST cells; (b) make the header expand to exactly C — a group label sitting over k sub-columns gets colspan="k"; a full-height left label (category) gets rowspan and the rows beneath it MUST NOT be left one cell short; (c) NEVER insert an all-blank spacer column to pad width — if a column would be empty in EVERY row, it must NOT exist (do not emit a phantom <td></td> column); (d) finally sum colspans for the header and for each body row — if any total ≠ C, FIX it before output. A table whose header column-sum ≠ body column-sum, or that contains a column blank in every row, is WRONG.
     Worked example — every row resolves to exactly 7 columns:
     <table><tr><th rowspan="2">항목</th><th colspan="2">2024년</th><th colspan="2">2025년</th><th rowspan="2">증감</th><th rowspan="2">비고</th></tr><tr><th>상반기</th><th>하반기</th><th>상반기</th><th>하반기</th></tr><tr><td>매출</td><td>10</td><td>12</td><td>13</td><td>15</td><td>+2</td><td>-</td></tr></table>
     (header row-1 = 1+2+2+1+1 = 7; row-2 fills only the 4 sub-columns under the two colspan groups while 항목·증감·비고 carry down by rowspan; body = 7. All rows = 7. No blank padding column.)
  (8-1) A full-width section band inside a table is its OWN row with one cell whose colspan covers every outer column. Never attach that band to one data column or model it by splitting the following data row. A nested grid remains inside its actual parent cell below the section band.
  (9) MULTI-LINE CELL: when content is stacked on several lines INSIDE one bordered cell (e.g. a name with its phone number beneath it in the SAME cell, or a value with its unit below), transcribe EVERY line of that cell joined by <br> — never keep only the first line. This applies ONLY to lines within one cell's borders; genuinely separate rows stay separate <tr>.
- TABLE CAPTION + NOTES: put the FULL table title in "caption" INCLUDING any trailing parenthetical unit/as-of suffix exactly as printed (e.g. "현황(억원)", "생산량(단위:천톤)", "(’25.6.23 기준)"). A footnote or source line printed OUTSIDE the table border beneath it ("주) …", "* 출처: …", "* 주:") → emit as its OWN text element (type text, not footnote); do NOT drop it and do NOT duplicate an in-table note row.


{schema_instruction}"""

        user_prompt = f"""
Here is the raw text extracted from the document:
<raw_text>
{extracted_text}
</raw_text>

Based on the raw text and the visual layout in the image (if provided), structure the information strictly in {output_format.upper()} format.
"""
        def _payload(img_w):
            cp = [{"type": "text", "text": user_prompt}]
            if img_path and os.path.exists(img_path):
                b64 = cls.encode_image(img_path, max_width=img_w)
                cp.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}})
            return cp

        def _strip_fences(text):
            t = text.strip()
            if t.startswith("```json"): t = t[7:]
            elif t.startswith("```xml"): t = t[6:]
            elif t.startswith("```"): t = t[3:]
            if t.endswith("```"): t = t[:-3]
            return t.strip()

        def _clip_prompt_text(text, limit):
            text = text or ""
            if len(text) <= limit:
                return text
            half = max(1, limit // 2)
            return text[:half] + "\n...[middle omitted]...\n" + text[-half:]

        def _candidate_for_prompt(candidate, limit=12000):
            try:
                candidate_obj = json.loads(candidate)
                candidate_text = json.dumps(candidate_obj, ensure_ascii=False)
            except Exception:
                candidate_text = candidate or ""
            return _clip_prompt_text(candidate_text, limit)

        def _json_coverage(result_text):
            parsed = json.loads(result_text)
            elements = parsed if isinstance(parsed, list) else parsed.get("elements", [])
            in_clean = re.sub(r"\s", "", extracted_text)
            cap_clean = re.sub(r"\s", "", "".join((e.get("content") or "") for e in elements if isinstance(e, dict)))
            return (len(cap_clean) / len(in_clean)) if in_clean else 1.0

        def _verify_low_coverage(candidate, cov):
            if not VLM_COVERAGE_VERIFY or output_format.lower() != "json":
                return False, "coverage verifier disabled"
            if not img_path or not os.path.exists(img_path):
                return False, "coverage verifier needs page image"
            try:
                candidate_text = _candidate_for_prompt(candidate)

                verifier_system = """You are a strict visual document QA verifier.
Decide whether the structured extraction preserves the important VISIBLE text on the page image.
Use the page IMAGE as the source of truth. Ignore PDF text-layer noise such as duplicated characters,
leader dots, page-number artifacts, background watermarks, and repeated OCR fragments.
For TOC pages, a toc_entry value like "title::page_number" preserves both the visible title and page number;
missing dotted leaders are NOT missing content. Section/group headings may be heading elements.
Return ONLY valid JSON:
{"pass": true|false, "reason": "short Korean reason", "missing_visible_text": ["..."]}
pass=true means no important visible text is missing, even if formatting is simplified.
pass=false means important visible text from the image is absent or materially wrong in the extraction."""
                verifier_user = (
                    f"Coverage score against the noisy PDF text layer was {cov:.2f}.\n"
                    "Structured extraction candidate:\n"
                    "<candidate>\n"
                    f"{candidate_text}\n"
                    "</candidate>\n\n"
                    "Judge against the page image. Korean answer in JSON values."
                )
                b64 = cls.encode_image(img_path, max_width=VLM_COVERAGE_VERIFY_IMG_MAXW)
                with _VLM_SEMAPHORE:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": verifier_system},
                            {"role": "user", "content": [
                                {"type": "text", "text": verifier_user},
                                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                            ]},
                        ],
                        temperature=0,
                        max_tokens=VLM_COVERAGE_VERIFY_MAX_TOKENS,
                        timeout=VLM_COVERAGE_VERIFY_TIMEOUT,
                        extra_body={"repetition_penalty": 1.05},
                    )
                verdict_raw = _strip_fences(response.choices[0].message.content or "")
                verdict = json.loads(_sanitize_json_strings(verdict_raw))
                passed = bool(verdict.get("pass"))
                reason = verdict.get("reason") or ""
                missing = verdict.get("missing_visible_text") or []
                if missing:
                    reason = f"{reason} missing={missing}"
                return passed, reason
            except Exception as e:
                logger.warning(f"coverage verifier 실패: {e}")
                return False, f"coverage verifier error: {e}"

        def _repair_low_coverage(candidate, verify_reason):
            if not VLM_COVERAGE_REPAIR or output_format.lower() != "json":
                return None, "coverage repair disabled"
            if not img_path or not os.path.exists(img_path):
                return None, "coverage repair needs page image"
            try:
                repair_system = system_prompt + """

[REPAIR MODE]
The previous extraction failed visual QA or text coverage. Re-extract the page from scratch.
Use the IMAGE as the source of truth and the raw PDF text only as support for exact spelling.
Keep all important visible text. Do not summarize, do not invent, and do not copy PDF text-layer noise.
For TOC pages: output "목차/목 차" as a heading, every row with a visible page number as toc_entry
"title::page_number", omit dotted leaders, and do not assign a child page number to a parent/group label."""
                repair_user = (
                    f"Verifier failure reason: {verify_reason}\n\n"
                    "Previous extraction candidate:\n"
                    "<candidate>\n"
                    f"{_candidate_for_prompt(candidate)}\n"
                    "</candidate>\n\n"
                    "Raw PDF text layer:\n"
                    "<raw_text>\n"
                    f"{_clip_prompt_text(extracted_text, 20000)}\n"
                    "</raw_text>\n\n"
                    "Return the repaired extraction as valid JSON only."
                )
                b64 = cls.encode_image(img_path, max_width=VLM_COVERAGE_REPAIR_IMG_MAXW)
                with _VLM_SEMAPHORE:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": repair_system},
                            {"role": "user", "content": [
                                {"type": "text", "text": repair_user},
                                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                            ]},
                        ],
                        temperature=0,
                        max_tokens=VLM_COVERAGE_REPAIR_MAX_TOKENS,
                        timeout=VLM_COVERAGE_REPAIR_TIMEOUT,
                        extra_body={"repetition_penalty": max(VLM_REP_PENALTY, 1.08)},
                    )
                finish = _vlm_finish_reason(response)
                if finish == "length":
                    return None, "repair truncated"
                repaired = _strip_fences(response.choices[0].message.content or "")
                try:
                    parsed = json.loads(repaired)
                except json.JSONDecodeError:
                    repaired = _sanitize_json_strings(repaired)
                    parsed = json.loads(repaired)
                if isinstance(parsed, list):
                    parsed = {"page_number": 0, "elements": parsed}
                if not isinstance(parsed, dict) or not isinstance(parsed.get("elements"), list) or not parsed.get("elements"):
                    return None, "repair returned empty elements"
                return json.dumps(parsed, ensure_ascii=False), "repair generated"
            except Exception as e:
                logger.warning(f"coverage repair 실패: {e}")
                return None, f"coverage repair error: {e}"

        def _compact_visual_retry(reason):
            """Recovery path for chart/figure-heavy pages that trigger long generation."""
            if not VLM_COMPACT_RETRY or output_format.lower() != "json":
                return None, "compact retry disabled"
            if not img_path or not os.path.exists(img_path):
                return None, "compact retry needs page image"
            try:
                compact_system = """You are a document parsing recovery model.
Return ONLY compact valid JSON with this schema:
{"page_number": int, "elements": [{"type": "heading_1|heading_2|heading_3|text|table|figure|footnote|toc_entry", "content": "...", "caption": "", "description": ""}]}

Rules:
- Preserve the visible reading order.
- Extract headings and body/list text exactly enough for retrieval.
- For charts, graphs, maps, screenshots, or photos: output ONE figure element with a short Korean description.
- Do NOT transcribe every chart axis tick, every plotted point, or dense UI/chart micro-label. Put only key visible labels in content if useful.
- Use table only for real bordered tables. Do not invent table structure.
- Output no markdown fences and no commentary."""
                compact_user = (
                    f"Previous full extraction failed because: {reason}\n\n"
                    "Raw text layer, if useful for exact spelling:\n"
                    "<raw_text>\n"
                    f"{_clip_prompt_text(extracted_text, 12000)}\n"
                    "</raw_text>\n\n"
                    "Re-extract this page compactly from the image."
                )
                b64 = cls.encode_image(img_path, max_width=VLM_COMPACT_RETRY_IMG_MAXW)
                with _VLM_SEMAPHORE:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": compact_system},
                            {"role": "user", "content": [
                                {"type": "text", "text": compact_user},
                                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                            ]},
                        ],
                        temperature=0,
                        max_tokens=VLM_COMPACT_RETRY_MAX_TOKENS,
                        timeout=VLM_COMPACT_RETRY_TIMEOUT,
                        extra_body={"repetition_penalty": max(VLM_REP_PENALTY_HI, 1.08), "no_repeat_ngram_size": 16},
                    )
                finish = _vlm_finish_reason(response)
                if finish == "length":
                    return None, "compact retry truncated"
                text = _strip_fences(response.choices[0].message.content or "")
                try:
                    parsed, _ = json.JSONDecoder().raw_decode(text)
                except json.JSONDecodeError:
                    text = _sanitize_json_strings(text)
                    parsed, _ = json.JSONDecoder().raw_decode(text)
                if isinstance(parsed, list):
                    parsed = {"page_number": 0, "elements": parsed}
                if not isinstance(parsed, dict) or not isinstance(parsed.get("elements"), list) or not parsed.get("elements"):
                    return None, "compact retry returned empty elements"
                for elem in parsed.get("elements", []):
                    if isinstance(elem, dict) and elem.get("type") not in {
                        "heading_1", "heading_2", "heading_3", "text", "table", "figure", "footnote", "toc_entry"
                    }:
                        elem["type"] = "text"
                return json.dumps(parsed, ensure_ascii=False), "compact retry generated"
            except Exception as e:
                logger.warning(f"compact visual retry 실패: {e}")
                return None, f"compact retry error: {e}"

        N = VLM_MAX_ATTEMPTS
        best_result, best_cov = None, -1.0   # 최고 coverage 후보 — 재시도가 결과를 악화시키지 못하게
        escalated = False                     # 루프/타임아웃 의심 시 1회 한정 고온 탈출

        def _chat_completion(rep, img_w, temperature):
            req = {
                "model": model_name,
                "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": _payload(img_w)}],
                "temperature": temperature,
                "max_tokens": VLM_MAX_TOKENS,
                "timeout": VLM_TIMEOUT,
                "extra_body": {
                    "repetition_penalty": rep,
                    "no_repeat_ngram_size": (24 if escalated else 0),
                },
            }
            if not VLM_STREAM_ABORT:
                response = client.chat.completions.create(**req)
                return response.choices[0].message.content or "", _vlm_finish_reason(response)

            stream = None
            chunks = []
            finish = None
            char_count = 0
            last_check = 0
            started_at = time.monotonic()
            try:
                stream = client.chat.completions.create(**req, stream=True)
                for chunk in stream:
                    if time.monotonic() - started_at > VLM_TIMEOUT:
                        raise _StreamDeadline(
                            f"stream exceeded total timeout of {VLM_TIMEOUT}s"
                        )
                    choices = getattr(chunk, "choices", None) or []
                    if not choices:
                        continue
                    choice = choices[0]
                    delta = getattr(choice, "delta", None)
                    content = getattr(delta, "content", None) or ""
                    if content:
                        chunks.append(content)
                        char_count += len(content)
                        if char_count >= VLM_STREAM_ABORT_MIN_CHARS and char_count - last_check >= 2048:
                            last_check = char_count
                            current = "".join(chunks)
                            if _is_runaway_repeat(current, min_repeats=8, tail=8192):
                                logger.warning("스트리밍 반복 생성 감지 — 응답 조기 중단")
                                if hasattr(stream, "close"):
                                    stream.close()
                                return current, "stream_runaway"
                            if _stream_output_excessive(current, extracted_text):
                                logger.warning("원문 길이 대비 비정상 출력 감지 — 응답 조기 중단")
                                if hasattr(stream, "close"):
                                    stream.close()
                                return current, "stream_excessive"
                    fr = getattr(choice, "finish_reason", None)
                    if fr:
                        finish = fr
                return "".join(chunks), finish
            finally:
                if stream is not None and hasattr(stream, "close"):
                    try:
                        stream.close()
                    except Exception:
                        pass

        for attempt in range(N):
            last = attempt >= N - 1
            try:
                with _VLM_SEMAPHORE:
                    # 재시도(escalated) 시 rep_penalty↑·no_repeat_ngram·저해상도로 전환.
                    rep = VLM_REP_PENALTY_HI if escalated else VLM_REP_PENALTY
                    img_w = VLM_IMG_MAXW_FALLBACK if escalated else VLM_IMG_MAXW
                    result, finish = _chat_completion(rep, img_w, 0.5 if escalated else 0.1)   # 의심(반복/타임아웃) 시에만 고온으로 전환
                result = _strip_fences(result or "")

                # 반복 생성 차단: 1회 escalated 탈출 후에도 반복이면 폐기.
                if finish in {"stream_runaway", "stream_excessive"} or (result and _is_runaway_repeat(result)):
                    logger.warning(f"비정상 출력 감지 (시도 {attempt+1}/{N}, finish={finish}) — 토큰 런어웨이 차단")
                    if not escalated and not last:
                        escalated = True
                        time.sleep(2)
                        continue
                    compact, compact_reason = _compact_visual_retry(
                        f"abnormal streaming output ({finish})"
                    )
                    if compact:
                        logger.info(f"비정상 스트리밍 출력 — compact visual retry 통과: {compact_reason}")
                        return compact, True
                    logger.warning(f"비정상 스트리밍 출력 — compact visual retry 실패: {compact_reason}")
                    return None, False

                truncated = (finish == "length")   # max_tokens 도달 = 미완(반복이거나 초대형 페이지)

                if output_format.lower() == "json":
                    try:
                        parsed = json.loads(result)
                    except json.JSONDecodeError as je:
                        sanitized = _sanitize_json_strings(result)
                        try:
                            parsed = json.loads(sanitized)
                            result = sanitized
                        except json.JSONDecodeError:
                            logger.warning(f"JSON 파싱 오류 (시도 {attempt+1}/{N}): {je} — 재시도")
                            # 잘린 JSON → 1회 escalated 탈출.
                            if not escalated and not last:
                                escalated = True
                                time.sleep(2)
                                continue
                            if not last:
                                time.sleep(5 * (attempt + 1))
                                continue
                            if best_result is None:          # 최종 실패 → 잘린 JSON 부분 복구 시도
                                salv = _salvage_truncated_json(result)
                                if salv and salv.get("elements"):
                                    logger.warning(f"잘린 JSON 부분 복구: element {len(salv['elements'])}개 보존")
                                    return json.dumps(salv, ensure_ascii=False), True
                            logger.error("JSON 유효성 검증 실패, 건너뜀")
                            return None, False
                    # coverage 검사(공백 제거 후 캡처/입력 비율, 임계 0.7). 미완(length)도 재시도.
                    elements = parsed if isinstance(parsed, list) else parsed.get("elements", [])
                    in_clean = re.sub(r"\s", "", extracted_text)
                    cap_clean = re.sub(r"\s", "", "".join((e.get("content") or "") for e in elements))
                    cov = (len(cap_clean) / len(in_clean)) if in_clean else 1.0
                    if cov > best_cov:
                        best_cov, best_result = cov, result
                    low_coverage = bool(in_clean and cov < 0.7)
                    if truncated or low_coverage:
                        why = "출력 미완(length)" if truncated else f"내용 부족(cov={cov:.2f})"
                        if truncated and escalated:
                            compact, compact_reason = _compact_visual_retry(why)
                            if compact:
                                logger.info(f"{why} — compact visual retry 통과: {compact_reason}")
                                return compact, True
                            logger.warning(f"{why} — compact visual retry 실패: {compact_reason}")
                        if not last:
                            logger.warning(f"{why} (시도 {attempt+1}/{N}) — 재시도")
                            # 미완(length) → 1회 escalated 전환.
                            if truncated and not escalated:
                                escalated = True
                            time.sleep(5 * (attempt + 1))
                            continue
                        if low_coverage and not truncated:
                            candidate = best_result or result
                            passed, reason = _verify_low_coverage(candidate, best_cov if best_result else cov)
                            if passed:
                                logger.info(f"{why} — 이미지 verifier 통과: {reason}")
                                return candidate, True
                            logger.warning(f"{why} — 이미지 verifier 실패: {reason}")
                            repaired, repair_reason = _repair_low_coverage(candidate, reason)
                            if repaired:
                                try:
                                    repair_cov = _json_coverage(repaired)
                                except Exception:
                                    repair_cov = 0.0
                                repair_passed, repair_verify_reason = _verify_low_coverage(repaired, repair_cov)
                                if repair_passed:
                                    logger.info(f"{why} — coverage repair+이미지 verifier 통과(cov={repair_cov:.2f}): {repair_verify_reason}")
                                    return repaired, True
                                if repair_cov >= 0.7 and not VLM_COVERAGE_VERIFY:
                                    logger.info(f"{why} — coverage repair 통과(cov={repair_cov:.2f}): verifier disabled")
                                    return repaired, True
                                logger.warning(f"{why} — coverage repair 검증 실패(cov={repair_cov:.2f}): {repair_verify_reason}")
                            else:
                                logger.warning(f"{why} — coverage repair 실패: {repair_reason}")
                        logger.warning(f"{why} (최종 시도) — 구조추출 실패로 폴백 처리")
                        return None, True
                    return (best_result or result), True
                else:
                    # 비-JSON(xml 등)은 파싱 스키마가 달라 원시 길이 기반 검사 유지
                    input_len = len(extracted_text)
                    short_output = bool(input_len and len(result) < input_len * 0.9)
                    if truncated or short_output:
                        if not last:
                            logger.warning(f"출력 미완/부족 (시도 {attempt+1}/{N}, finish={finish}) — 재시도")
                            time.sleep(5 * (attempt + 1))
                            continue
                        logger.warning(f"출력 미완/부족 (최종 시도, finish={finish}) — 구조추출 실패로 폴백 처리")
                        return None, True
                    return result, True
            except (_APITimeout, _StreamDeadline) as e:
                # 타임아웃: 일시적 오류와 분리해 1회만 고온 재시도 후 중단
                logger.warning(f"VLM 타임아웃 (시도 {attempt+1}/{N}): {e} — 반복 생성 의심")
                if escalated:
                    compact, compact_reason = _compact_visual_retry(f"timeout after escalated retry: {e}")
                    if compact:
                        logger.info(f"VLM 타임아웃 — compact visual retry 통과: {compact_reason}")
                        return compact, True
                    logger.warning(f"VLM 타임아웃 — compact visual retry 실패: {compact_reason}")
                if not escalated and not last:
                    escalated = True
                    continue
                return None, False
            except Exception as e:
                # 연결/5xx/429 등 일시적 오류 → 백오프 후 재시도(최종재시도 패스 자격 유지)
                logger.warning(f"VLM API 에러 (시도 {attempt+1}/{N}): {e}")
                if not last:
                    time.sleep(5 * (attempt + 1))
                else:
                    logger.error("VLM API 재시도 소진, 건너뜀")
                    return None, True
        return None, False

    @classmethod
    def describe_image(cls, img_path, api_key, model_name="Qwen/Qwen3-VL-30B-A3B-Instruct"):
        if not api_key or "여기에_" in api_key or not os.path.exists(img_path):
            return ""
        try:
            from openai import OpenAI
            base_url = os.environ.get("VLM_BASE_URL", "http://localhost:8000/v1")
            client = OpenAI(base_url=base_url, api_key=api_key or "EMPTY", timeout=120, max_retries=0)
        except Exception as e:
            logger.error(f"VLM 클라이언트 초기화 오류: {e}")
            return ""
        prompt = ("이 이미지가 '기관 로고·CI·심볼마크·단순 아이콘·장식용 선/문양·도장(직인)'처럼 "
                  "정보 전달이 목적이 아닌 요소이면, 다른 말 없이 정확히 'LOGO' 한 단어만 출력하세요.\n"
                  "그 외 정보를 담은 시각자료 — 차트·그래프·사진·도식·표·인포그래픽·내용이 있는 화면 캡처 등 — "
                  "이면 한국어로 2~3문장으로 핵심(수치·추세·내용)을 설명하세요. 설명문만 출력하세요.")
        content = [{"type": "text", "text": prompt},
                   {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{cls.encode_image(img_path)}"}}]
        for attempt in range(2):
            try:
                with _VLM_SEMAPHORE:
                    r = client.chat.completions.create(
                        model=model_name,
                        messages=[{"role": "user", "content": content}],
                        temperature=0.2,
                        max_tokens=VLM_MAX_TOKENS,
                        timeout=120,
                    )
                return (r.choices[0].message.content or "").strip()
            except Exception as e:
                logger.warning(f"figure 설명 VLM 에러 (시도 {attempt+1}/2): {e}")
                if attempt < 1:
                    time.sleep(3)
        return ""

    @classmethod
    def extract_metadata(cls, txt_paths, img_paths, api_key, model_name):
        if not api_key or "여기에_" in api_key:
            return {}
        try:
            from openai import OpenAI
            base_url = os.environ.get("VLM_BASE_URL", "http://localhost:8000/v1")
            client = OpenAI(
                base_url=base_url,
                api_key=api_key or "EMPTY",
                timeout=VLM_METADATA_TIMEOUT,
                max_retries=0,
            )
        except Exception as e:
            logger.error(f"VLM 클라이언트 초기화 오류: {e}")
            return {}

        combined_text = ""
        for p in txt_paths[:2]:
            if p and os.path.exists(p):
                with open(p, "r", encoding="utf-8") as f:
                    combined_text += f.read() + "\n"

        system_prompt = """You are a document metadata extraction AI.
Extract global metadata from the provided document cover/table-of-contents pages.
Output ONLY valid JSON with this exact schema — no markdown, no extra text:
{
    "doc_title": "Full document title",
    "date": "Publication or effective date (YYYY-MM-DD or as written)",
    "organization": "Issuing organization or department",
    "author": "Author name(s) or null if not found",
    "keywords": ["keyword1", "keyword2"]
}
If a field cannot be found, use null.

[ANTI-HALLUCINATION — STRICT]
- Output a value ONLY if it is EXPLICITLY printed on the provided page(s). Never infer, guess, or invent.
- If author/date/organization is not literally written on the page, the value MUST be null.
- NEVER output placeholder/example values such as "John Doe", "Jane Doe", "2023-10-15", "Unknown", "N/A", or a sample organization. A placeholder is worse than null — use null.
- "organization" must be a real issuing body/department name printed in the document, NOT the document's title or a label phrase.
- Write field values in the document's own language."""

        content_payload = [{"type": "text", "text": f"<raw_text>\n{combined_text}\n</raw_text>\n\nExtract metadata as JSON."}]
        for ip in img_paths[:2]:
            if ip and os.path.exists(ip):
                content_payload.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{cls.encode_image(ip)}"}
                })

        for attempt in range(VLM_METADATA_ATTEMPTS):
            try:
                with _VLM_SEMAPHORE:
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": content_payload}],
                        temperature=0.0,
                        max_tokens=VLM_MAX_TOKENS,
                        timeout=VLM_METADATA_TIMEOUT,
                    )
                raw = (response.choices[0].message.content or "").strip()
                if raw.startswith("```"): raw = raw.split("\n", 1)[-1]
                if raw.endswith("```"): raw = raw.rsplit("```", 1)[0]
                parsed = json.loads(raw.strip())
                return parsed if isinstance(parsed, dict) else {}
            except Exception as e:
                logger.warning(
                    f"메타데이터 추출 실패 (시도 {attempt + 1}/{VLM_METADATA_ATTEMPTS}): {e}"
                )
                if attempt + 1 < VLM_METADATA_ATTEMPTS:
                    time.sleep(min(3, attempt + 1))
        return {}


def _xlsx_sheet_tables(path, batch_chars=3000):
    """각 시트를 HTML <table> 배치들로 추출(병합셀 colspan/rowspan, 빈영역 트림,
    행배치 분할+헤더반복, 차트/이미지 감지)."""
    import html as _html
    esc = lambda v: _html.escape("" if v is None else str(v))
    wb = openpyxl.load_workbook(path, data_only=True)            # 캐시(계산)값
    try:
        wb_raw = openpyxl.load_workbook(path, data_only=False)   # 수식/원본 — 캐시 없는 수식셀 복구용
    except Exception:
        wb_raw = None
    out = []
    for ws in wb.worksheets:
        ws_raw = wb_raw[ws.title] if (wb_raw and ws.title in wb_raw.sheetnames) else None
        def cval(r, c, _ws=ws, _wsr=ws_raw):
            v = _ws.cell(row=r, column=c).value
            if v in (None, "") and _wsr is not None:
                v = _wsr.cell(row=r, column=c).value            # 캐시 없으면 수식 문자열이라도 보존
            return v
        # 경계 박스: 수식셀도 포함되도록 raw 우선 스캔
        min_r = min_c = 10 ** 9; max_r = max_c = 0; nonempty = 0
        for row in (ws_raw or ws).iter_rows():
            for c in row:
                if c.value not in (None, ""):
                    nonempty += 1
                    min_r = min(min_r, c.row); max_r = max(max_r, c.row)
                    min_c = min(min_c, c.column); max_c = max(max_c, c.column)
        imgs = []
        for _im in getattr(ws, "_images", []):
            try:
                imgs.append(_im._data())   # 임베디드 이미지 원본 바이트
            except Exception as e:
                logger.warning(f"임베디드 이미지 추출 실패 ({ws.title}): {e}")
        rec = {"name": ws.title, "tables": [], "cells": nonempty,
               "has_chart": bool(getattr(ws, "_charts", [])), "images": imgs}
        if nonempty == 0:
            out.append(rec); continue
        # 병합셀: 경계 박스로 클램프(박스 밖으로 삐져나간 span 방지) + 세로병합이 가로지르는 행경계 표시
        span = {}; covered = set(); vspan_break = set()
        for rng in ws.merged_cells.ranges:
            r0 = max(rng.min_row, min_r); c0 = max(rng.min_col, min_c)
            r1 = min(rng.max_row, max_r); c1 = min(rng.max_col, max_c)
            if r0 > r1 or c0 > c1:
                continue
            span[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    if (r, c) != (r0, c0):
                        covered.add((r, c))
            for r in range(r0, r1):       # r→r+1 경계가 이 세로병합 내부 → 배치 분할 금지
                vspan_break.add(r)
        def row_html(r):
            tds = []
            for c in range(min_c, max_c + 1):
                if (r, c) in covered:
                    continue
                attr = ""
                if (r, c) in span:
                    rs, cs = span[(r, c)]
                    if rs > 1: attr += f' rowspan="{rs}"'
                    if cs > 1: attr += f' colspan="{cs}"'
                tds.append(f"<td{attr}>{esc(cval(r, c))}</td>")
            return "<tr>" + "".join(tds) + "</tr>"
        rows = {r: row_html(r) for r in range(min_r, max_r + 1)}
        full = "<table>" + "".join(rows[r] for r in range(min_r, max_r + 1)) + "</table>"
        if len(full) <= batch_chars:
            rec["tables"] = [full]
        else:
            # 헤더 반복은 첫 행이 고정창(freeze_panes)으로 '헤더'임이 분명할 때만 — 데이터행을 가짜
            # 헤더로 복제하지 않도록.
            has_header = False
            try:
                fp = ws.freeze_panes
                if fp:
                    from openpyxl.utils.cell import coordinate_to_tuple
                    if coordinate_to_tuple(fp)[0] == min_r + 1:
                        has_header = True
            except Exception:
                pass
            hdr = rows[min_r] if has_header else ""
            body0 = min_r + 1 if has_header else min_r
            batches = []; cur = []; clen = len(hdr) + 16
            for r in range(body0, max_r + 1):
                rh = rows[r]
                can_break = (r - 1) not in vspan_break   # 세로병합 내부에서는 분할 금지(rowspan 깨짐 방지)
                if cur and clen + len(rh) > batch_chars and can_break:
                    batches.append("<table>" + hdr + "".join(cur) + "</table>"); cur = []; clen = len(hdr) + 16
                cur.append(rh); clen += len(rh)
            if cur:
                batches.append("<table>" + hdr + "".join(cur) + "</table>")
            rec["tables"] = batches
        out.append(rec)
    wb.close()
    if wb_raw is not None:
        try: wb_raw.close()
        except Exception: pass
    return out


class DocumentProcessor:

    @staticmethod
    def get_libreoffice_cmd():
        if platform.system() == "Linux": return "libreoffice"
        candidates = [r"C:\Program Files\LibreOffice\program\soffice.exe", r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"]
        if shutil.which("soffice"): return "soffice"
        for p in candidates:
            if os.path.exists(p): return p
        return None

    @staticmethod
    def _optimize_excel_layout(input_path, output_path):
        try:
            wb = openpyxl.load_workbook(input_path)
            for ws in wb.worksheets:
                if not isinstance(ws, Worksheet):
                    continue
                try:
                    ws.print_area = None
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col[:100]:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except Exception:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                    # A4 페이지네이션
                    ws.page_setup.paperSize = 9
                    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                    ws.page_setup.fitToPage = False
                    ws.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
                except Exception:
                    continue
            wb.save(output_path)
            return True
        except InvalidFileException:
            shutil.copy2(input_path, output_path)
            return False
        except Exception as e:
            logger.warning(f"엑셀 레이아웃 최적화 실패: {e}")
            shutil.copy2(input_path, output_path)
            return False

    @classmethod
    def _convert_excel_to_pdf_with_sheet_map(cls, input_path, output_dir):
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        cmd_exe = cls.get_libreoffice_cmd()
        if not cmd_exe: return None, []

        try:
            wb_check = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            sheet_names = wb_check.sheetnames
            wb_check.close()
        except Exception:
            return cls.convert_to_pdf(input_path, output_dir), []

        abs_output_dir = os.path.abspath(output_dir)
        temp_id = uuid.uuid4().hex
        sheet_pdfs = []
        sheet_map = []
        current_page = 1

        for sheet_name in sheet_names:
            try:
                wb = openpyxl.load_workbook(input_path)
                for sn in [s for s in wb.sheetnames if s != sheet_name]:
                    del wb[sn]
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', sheet_name)
                temp_xlsx = os.path.join(abs_output_dir, f"temp_{temp_id}_{safe_name}.xlsx")
                opt_xlsx = os.path.join(abs_output_dir, f"opt_{temp_id}_{safe_name}.xlsx")
                wb.save(temp_xlsx)
                wb.close()
            except Exception as e:
                logger.warning(f"시트 {sheet_name} xlsx 생성 실패: {e}")
                continue

            cls._optimize_excel_layout(temp_xlsx, opt_xlsx)

            sheet_pdf = os.path.join(abs_output_dir, f"sheet_{temp_id}_{safe_name}.pdf")
            generated = os.path.join(abs_output_dir, f"opt_{temp_id}_{safe_name}.pdf")
            cmd = [cmd_exe, "--headless", "--convert-to", "pdf", "--outdir", abs_output_dir, opt_xlsx]
            try:
                startupinfo = None
                if platform.system() == "Windows":
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                with _SOFFICE_LOCK:
                    subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                   startupinfo=startupinfo, timeout=90)
                if os.path.exists(generated):
                    os.rename(generated, sheet_pdf)
                    doc = fitz.open(sheet_pdf)
                    n_pages = len(doc)
                    doc.close()
                    sheet_map.append({"name": sheet_name, "page_start": current_page, "page_end": current_page + n_pages - 1})
                    current_page += n_pages
                    sheet_pdfs.append(sheet_pdf)
            except Exception as e:
                logger.warning(f"시트 {sheet_name} PDF 변환 실패: {e}")
            finally:
                for f in [temp_xlsx, opt_xlsx]:
                    if os.path.exists(f): os.remove(f)

        if not sheet_pdfs:
            return None, []

        name = os.path.splitext(os.path.basename(input_path))[0]
        merged_pdf_path = os.path.join(abs_output_dir, f"{name}.pdf")
        try:
            merged = fitz.open()
            for sp in sheet_pdfs:
                src = fitz.open(sp)
                merged.insert_pdf(src)
                src.close()
            merged.save(merged_pdf_path)
            merged.close()
        except Exception as e:
            logger.error(f"PDF 합치기 실패: {e}")
            return None, []
        finally:
            for sp in sheet_pdfs:
                if os.path.exists(sp): os.remove(sp)

        return merged_pdf_path, sheet_map

    @classmethod
    def _xlsx_hybrid_write(cls, doc_output_dir, src, sheet_map, api_key, model_name, output_format="json"):
        """시트 데이터를 추출하고, 차트/이미지가 있는 시트만 VLM 설명을 붙여
        시트별 structured.json(+ markdown/xml)을 기록한다. 시트 첫 페이지에
        [시트명 heading + 표 배치들 (+차트/이미지 figure)], 나머지 연속 페이지는 빈 structured.
        내용이 없는 시트는 기록하지 않는다(일반 VLM 경로가 처리)."""
        tables = _xlsx_sheet_tables(src)   # 실패 시 예외 → 호출자가 일반 VLM 으로 폴백
        by_name = {t["name"]: t for t in tables}
        fmt = output_format.lower()

        def write_page(pg, elements):
            stem = os.path.join(doc_output_dir, f"page_{pg:04d}_structured")
            payload = {"page_number": pg, "elements": elements}
            with open(stem + ".json", "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            if fmt == "markdown":     # 다른 포맷과 동일하게 .md 도 생성(출력 계약 유지)
                try:
                    with open(stem + ".md", "w", encoding="utf-8") as f:
                        f.write(cls.json_to_markdown(json.dumps(payload, ensure_ascii=False)))
                except Exception as e:
                    logger.warning(f"XLSX markdown 생성 실패 (page {pg}): {e}")
            elif fmt == "xml":
                try:
                    import html as _h
                    parts = ["<document>", f"<page_number>{pg}</page_number>", "<elements>"]
                    for e in elements:
                        cap = _h.escape(e.get("caption", "") or "")
                        parts.append(f'<element type="{e.get("type","text")}"'
                                     + (f' caption="{cap}"' if cap else "")
                                     + f'>{_h.escape(e.get("content","") or "")}</element>')
                    parts += ["</elements>", "</document>"]
                    with open(stem + ".xml", "w", encoding="utf-8") as f:
                        f.write("\n".join(parts))
                except Exception as e:
                    logger.warning(f"XLSX xml 생성 실패 (page {pg}): {e}")

        for sm in sheet_map:
            ps, pe = sm["page_start"], sm["page_end"]
            if ps > pe:
                logger.warning(f"XLSX 시트 '{sm.get('name')}' 페이지범위 역전(ps={ps}>pe={pe}) — 보정")
                pe = ps
            rec = by_name.get(sm["name"])
            if rec is None:
                logger.warning(f"XLSX 시트명 불일치: sheet_map '{sm.get('name')}' 가 추출 결과에 없음")
            elements = []
            try:
                if rec is not None:
                    has_visual = rec.get("has_chart") or rec.get("images")
                    if rec["cells"] > 0 or has_visual:
                        # 각 시트 = 독립 최상위 섹션
                        elements.append({"type": "heading_1", "content": sm["name"], "_section_root": True})
                    for tb in rec.get("tables", []):
                        elements.append({"type": "table", "content": tb, "caption": sm["name"]})
                    # 차트 설명(상단 페이지)
                    if rec.get("has_chart"):
                        for pg in range(ps, min(pe, ps + 2) + 1):
                            img = os.path.join(doc_output_dir, f"page_{pg:04d}.png")
                            if not os.path.exists(img):
                                continue
                            try:
                                desc = VLMProcessor.describe_image(img, api_key, model_name)
                            except Exception as e:
                                logger.warning(f"XLSX 차트 설명 실패 (page {pg}): {e}"); desc = ""
                            if desc and desc.strip().upper() != "LOGO":
                                elements.append({"type": "figure", "content": desc, "page_number": pg})
                    # 임베디드 이미지 설명(로고 제외)
                    for k, data in enumerate(rec.get("images", [])[:5]):
                        tmp = os.path.join(doc_output_dir, f"_xlimg_{ps}_{k}.png")
                        try:
                            with open(tmp, "wb") as f:
                                f.write(data)
                            desc = VLMProcessor.describe_image(tmp, api_key, model_name)
                            if desc and desc.strip().upper() != "LOGO":
                                elements.append({"type": "figure", "content": desc, "caption": f"{sm['name']} 임베디드 이미지"})
                        except Exception as e:
                            logger.warning(f"XLSX 임베디드 이미지 설명 실패 ({sm.get('name')}): {e}")
                        finally:
                            if os.path.exists(tmp):
                                os.remove(tmp)
            except Exception as e:
                # 한 시트 실패가 다른 시트·이미 추출된 표 데이터를 버리지 않게 격리
                logger.warning(f"XLSX 시트 '{sm.get('name')}' 처리 실패(건너뜀): {e}")
                elements = [e for e in elements if e.get("type") in ("heading_1", "table")]
            # 내용이 전혀 없는 시트는 기록 안 함 → 일반 VLM 경로가 그 페이지들을 처리(누락 방지)
            if not elements:
                continue
            for pg in range(ps, pe + 1):
                write_page(pg, elements if pg == ps else [])
        return []

    @staticmethod
    def _extract_hwp_memos(input_path, doc=None):
        """HWP/HWPX 메모 텍스트 수집(파일 파싱 + rhwp IR memo 병합, 중복 제거)."""
        memos = []
        try:
            import hwp_memo
            memos.extend(hwp_memo.extract_hwp_memos(input_path))
        except Exception as e:
            logger.warning(f"HWP 메모 직접파싱 실패: {e}")
        # 보강: rhwp IR 의 FieldBlock(field_kind='memo')
        try:
            ir = doc.to_ir() if (doc is not None and hasattr(doc, "to_ir")) else None
            if ir is not None:
                for b in ir.iter_blocks(scope="all", recurse=True):
                    if getattr(b, "kind", "") == "field" and getattr(b, "field_kind", "") == "memo":
                        v = (getattr(b, "cached_value", None) or getattr(b, "raw_instruction", None) or "").strip()
                        if v:
                            memos.append(v)
        except Exception:
            pass
        # 중복 제거(순서 보존)
        seen, out = set(), []
        for m in memos:
            k = re.sub(r"\s+", "", m)
            if k and k not in seen:
                seen.add(k); out.append(m)
        return out

    @classmethod
    def _emit_hwp_memo_page(cls, input_path, doc, doc_output_dir, page_no):
        """메모가 있으면 본문 뒤에 '메모' 페이지 1장(png+txt)을 편입. 메모 없으면 0 반환(무동작)."""
        memos = cls._extract_hwp_memos(input_path, doc)
        if not memos:
            return 0
        lines = ["[문서 메모 — 본문 렌더에는 표시되지 않는 주석]"] + [f"· {m}" for m in memos]
        body = "\n".join(lines)
        pnum = str(page_no).zfill(4)
        try:
            from PIL import Image as _I, ImageDraw as _D, ImageFont as _F
            W, H = int(8.27 * RENDER_DPI), int(11.69 * RENDER_DPI)
            im = _I.new("RGB", (W, H), "white")
            dr = _D.Draw(im)
            try:
                ft = _F.truetype("/usr/share/fonts/truetype/nanum/NanumGothic.ttf", 30)
                fb = _F.truetype("/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf", 34)
            except Exception:
                ft = fb = _F.load_default()
            y = 90
            dr.text((90, 40), "📝 메모", font=fb, fill=(150, 30, 30))
            for ln in lines:
                for seg in [ln[i:i + 46] for i in range(0, len(ln), 46)] or [""]:
                    dr.text((90, y), seg, font=ft, fill=(30, 30, 30)); y += 44
            im.save(os.path.join(doc_output_dir, f"page_{pnum}.png"), "PNG")
        except Exception:
            open(os.path.join(doc_output_dir, f"page_{pnum}.png"), "wb").close()
        with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as f:
            f.write(HWPTextExtractor.text_preprocessing(body))
        logger.info(f"HWP 메모 {len(memos)}건 추출 → 메모 페이지 편입(p.{page_no})")
        return 1

    @classmethod
    def _render_hwp_with_rhwp(cls, input_path, doc_output_dir):
        import re as _re, html as _html
        rhwp = _load_rhwp()
        doc = rhwp.parse(input_path)
        n = doc.page_count
        if not n:
            return 0
        scale = RENDER_DPI / 96.0
        out_page = 0
        rendered_ok = 0            # 실제 렌더 성공한 본문 페이지 수(전부 실패 시 폴백 판단용)
        failed_pages = []
        for i in range(n):
            out_page += 1          # 실패해도 번호를 소비해 페이지 정합 유지(조용한 시프트 방지)
            pnum = str(out_page).zfill(4)
            try:
                png = doc.render_png(i, scale=scale)
            except Exception as e:
                logger.warning(f"rhwp page {i} render 실패(placeholder 대체): {e}")
                failed_pages.append(out_page)
                try:                # 번호 정합용 placeholder(흰 페이지) + 실패 마커 텍스트
                    from PIL import Image as _PILImg
                    import io as _io2
                    _buf = _io2.BytesIO()
                    _PILImg.new("RGB", (int(8.27 * RENDER_DPI), int(11.69 * RENDER_DPI)), "white").save(_buf, format="PNG")
                    with open(os.path.join(doc_output_dir, f"page_{pnum}.png"), "wb") as f:
                        f.write(_buf.getvalue())
                except Exception:
                    pass
                with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as f:
                    f.write("")
                continue
            rendered_ok += 1
            with open(os.path.join(doc_output_dir, f"page_{pnum}.png"), "wb") as f:
                f.write(png)
            try:
                svg = doc.render_svg(i)
                runs = _re.findall(r"<text[^>]*>(.*?)</text>", svg, flags=_re.S)
                txt = " ".join(_html.unescape(_re.sub(r"<[^>]+>", "", t)).strip() for t in runs if t.strip())
                txt = _re.sub(r"\s+", " ", txt).strip()
            except Exception:
                txt = ""
            with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as f:
                f.write(HWPTextExtractor.text_preprocessing(txt) if txt else "")
        # 렌더 실패가 많으면(전부 or 30%↑) placeholder 대신 전체 폴백(LibreOffice)에 맡김.
        if rendered_ok == 0 or len(failed_pages) > n * 0.3:
            logger.warning(f"rhwp 렌더 실패 과다({len(failed_pages)}/{n}) — 폴백 사용")
            for _pp in range(1, out_page + 1):
                for _ext in (".png", ".txt"):
                    _f = os.path.join(doc_output_dir, f"page_{str(_pp).zfill(4)}{_ext}")
                    if os.path.exists(_f):
                        try: os.remove(_f)
                        except OSError: pass
            return 0
        if failed_pages:
            logger.warning(f"rhwp 렌더 실패 페이지(placeholder 대체): {failed_pages}")
        # 메모(주석)는 렌더에 안 나오므로 파일에서 직접 추출해 본문 뒤 페이지로 편입(있을 때만).
        try:
            out_page += cls._emit_hwp_memo_page(input_path, doc, doc_output_dir, out_page + 1)
        except Exception as e:
            logger.warning(f"HWP 메모 페이지 편입 실패: {e}")
        # 네이티브 표 HTML 을 사이드카로 저장 → _persist_page 가 내용매칭으로 치환.
        try:
            blocks = list(doc.to_ir().iter_blocks())
            native = [{"html": b.html, "rows": b.rows or 0, "cols": b.cols or 0}
                      for b in blocks
                      if type(b).__name__ == "TableBlock" and (getattr(b, "html", "") or "").strip()]
            if native:
                with open(os.path.join(doc_output_dir, "_native_tables.json"), "w", encoding="utf-8") as f:
                    json.dump(native, f, ensure_ascii=False)
            # figure 앵커: 각 그림의 직전·직후 문단 텍스트(_persist_page 의 순서 교정용).
            anchors = []
            for i, b in enumerate(blocks):
                if type(b).__name__ != "PictureBlock":
                    continue
                prev = nxt = ""
                for j in range(i - 1, -1, -1):
                    t = (getattr(blocks[j], "text", "") or "").strip()
                    if t:
                        prev = t; break
                for j in range(i + 1, len(blocks)):
                    t = (getattr(blocks[j], "text", "") or "").strip()
                    if t:
                        nxt = t; break
                anchors.append({"prev": prev, "next": nxt})
            if anchors:
                with open(os.path.join(doc_output_dir, "_figure_anchors.json"), "w", encoding="utf-8") as f:
                    json.dump(anchors, f, ensure_ascii=False)
        except Exception as e:
            logger.warning(f"HWP 네이티브 표/figure 앵커 추출 실패: {e}")
        return out_page

    @classmethod
    def convert_to_pdf(cls, input_path, output_dir):
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        cmd_exe = cls.get_libreoffice_cmd()
        if not cmd_exe: return None

        filename = os.path.basename(input_path)
        name, ext = os.path.splitext(filename)
        temp_id = uuid.uuid4().hex
        abs_output_dir = os.path.abspath(output_dir)
        safe_input_path = os.path.join(abs_output_dir, f"temp_{temp_id}{ext}")

        if ext.lower() == '.xlsx':
            cls._optimize_excel_layout(input_path, safe_input_path)
        else:
            shutil.copy2(input_path, safe_input_path)

        cmd = [cmd_exe, "--headless", "--convert-to", "pdf", "--outdir", abs_output_dir, safe_input_path]
        generated_temp_pdf = os.path.join(abs_output_dir, f"temp_{temp_id}.pdf")
        final_pdf_path = os.path.join(abs_output_dir, f"{name}.pdf")

        try:
            startupinfo = None
            if platform.system() == "Windows":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            with _SOFFICE_LOCK:
                result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, startupinfo=startupinfo, timeout=90)
            if result.returncode != 0:
                logger.error(f"LibreOffice 변환 실패 (code {result.returncode}): {result.stderr.decode('utf-8', errors='replace')}")
                return None

            if os.path.exists(generated_temp_pdf):
                if os.path.exists(final_pdf_path): os.remove(final_pdf_path)
                os.rename(generated_temp_pdf, final_pdf_path)
                return final_pdf_path

            logger.error(f"LibreOffice 변환 후 PDF 없음: {generated_temp_pdf}")
            return None
        except Exception as e:
            logger.error(f"LibreOffice 변환 예외: {e}")
            return None
        finally:
            if os.path.exists(safe_input_path): os.remove(safe_input_path)
            if os.path.exists(generated_temp_pdf): os.remove(generated_temp_pdf)

    @classmethod
    def _render_eml(cls, file_path, doc_output_dir, temp_pdf_dir, api_key, model_name, set_progress=None,
                    output_format="json"):
        """EML → page_NNNN.png/.txt 연속 시퀀스.

        1) 헤더와 디코딩된 본문을 결정적 텍스트 페이지로 렌더.
        2) 첨부를 종류별로 재귀 렌더해 같은 페이지 시퀀스에 이어붙임
           (HWP/HWPX→rhwp, PDF/DOCX/XLSX 등→PDF 변환, 이미지→PNG 1페이지).
        반환: 생성된 총 페이지 수.
        """
        import email, email.policy, html as _html, io as _io
        def _say(m, p):
            if set_progress: set_progress(m, p)

        with open(file_path, "rb") as f:
            msg = email.message_from_binary_file(f, policy=email.policy.default)

        os.makedirs(temp_pdf_dir, exist_ok=True)
        state = {"n": 0}
        mat = fitz.Matrix(RENDER_DPI / 72, RENDER_DPI / 72)

        def _page_number(stem):
            try:
                return int(stem.split("_")[1])
            except (IndexError, ValueError):
                return 0

        def _write_text_structured(stem, text, low_confidence=False, warning=None):
            """Write deterministic structure for generated EML text/notice pages.

            These pages are created by this renderer from trusted text, so sending
            them back through VLM only adds latency and can trigger long repeats.
            """
            text = HWPTextExtractor.text_preprocessing((text or "").strip())
            paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
            if not paras and text:
                paras = [text]
            elem_source = "eml_notice" if warning else "eml_text"
            elem_conf = 0.35 if low_confidence else 1.0
            payload = {
                "page_number": _page_number(stem),
                "elements": [
                    _mark_element({"type": "text", "content": p}, elem_source, elem_conf)
                    for p in paras
                ],
            }
            if low_confidence:
                payload["low_confidence"] = True
            if warning:
                payload["warning"] = warning
            raw_json = json.dumps(payload, ensure_ascii=False, indent=4)
            with open(os.path.join(doc_output_dir, f"{stem}_structured.json"), "w", encoding="utf-8") as fh:
                fh.write(raw_json)
            fmt = (output_format or "json").lower()
            if fmt == "markdown":
                with open(os.path.join(doc_output_dir, f"{stem}_structured.md"), "w", encoding="utf-8") as fh:
                    fh.write(cls.json_to_markdown(raw_json))
            elif fmt == "xml":
                import html as _html
                parts = [
                    "<document>",
                    f"  <page_number>{payload['page_number']}</page_number>",
                    "  <elements>",
                ]
                for elem in payload["elements"]:
                    parts.append(f"    <element type=\"text\">{_html.escape(elem.get('content') or '')}</element>")
                parts.append("  </elements>")
                if low_confidence:
                    parts.append("  <low_confidence>true</low_confidence>")
                if warning:
                    parts.append(f"  <warning>{_html.escape(warning)}</warning>")
                parts.append("</document>")
                with open(os.path.join(doc_output_dir, f"{stem}_structured.xml"), "w", encoding="utf-8") as fh:
                    fh.write("\n".join(parts))

        def _emit_pdf(pdf_path, direct_structured=False):
            try:
                d = fitz.open(pdf_path)
            except Exception as e:
                logger.warning(f"EML PDF 열기 실패: {e}")
                return
            try:
                for i in range(len(d)):
                    state["n"] += 1
                    pnum = str(state["n"]).zfill(4)
                    d[i].get_pixmap(matrix=mat).save(os.path.join(doc_output_dir, f"page_{pnum}.png"))
                    raw = d[i].get_text().strip()
                    cleaned = HWPTextExtractor.text_preprocessing(raw)
                    with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as fh:
                        fh.write(cleaned)
                    if direct_structured:
                        _write_text_structured(f"page_{pnum}", cleaned)
            finally:
                d.close()

        def _emit_image(im, hint="", direct_structured=False, low_confidence=False, warning=None):
            state["n"] += 1
            pnum = str(state["n"]).zfill(4)
            im.save(os.path.join(doc_output_dir, f"page_{pnum}.png"), "PNG")
            with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as fh:
                fh.write(hint)
            if direct_structured:
                _write_text_structured(
                    f"page_{pnum}",
                    hint,
                    low_confidence=low_confidence,
                    warning=warning,
                )

        def _emit_notice(lines):
            """첨부를 렌더할 수 없을 때(DRM·실패) 가시적 마커 페이지를 한 장 생성.
            본문 텍스트레이어에도 동일 안내를 남겨 청크/검색에 누락이 아닌 '보호됨'으로 드러나게 한다."""
            try:
                from PIL import Image as _PILImage, ImageDraw as _ImageDraw, ImageFont as _ImageFont
                W, H = int(8.27 * RENDER_DPI), int(11.69 * RENDER_DPI)
                im = _PILImage.new("RGB", (W, H), "white")
                dr = _ImageDraw.Draw(im)
                try:
                    fb = _ImageFont.truetype("/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf", 40)
                    fr = _ImageFont.truetype("/usr/share/fonts/truetype/nanum/NanumGothic.ttf", 30)
                except Exception:
                    fb = fr = _ImageFont.load_default()
                dr.rectangle([60, 60, W - 60, 60 + 70 * len(lines) + 80], outline=(180, 60, 60), width=4)
                y = 110
                for i, ln in enumerate(lines):
                    dr.text((110, y), ln, font=(fb if i == 0 else fr), fill=(150, 30, 30) if i == 0 else (40, 40, 40))
                    y += 70
                _emit_image(
                    im,
                    hint="\n".join(lines),
                    direct_structured=True,
                    low_confidence=True,
                    warning="렌더 불가 첨부 — 원본 확인 필요.",
                )
            except Exception:
                # PIL 실패시 텍스트 페이지라도 남김
                state["n"] += 1
                pnum = str(state["n"]).zfill(4)
                open(os.path.join(doc_output_dir, f"page_{pnum}.png"), "wb").close()
                with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as fh:
                    fh.write("\n".join(lines))
                _write_text_structured(
                    f"page_{pnum}",
                    "\n".join(lines),
                    low_confidence=True,
                    warning="렌더 불가 첨부 — 원본 확인 필요.",
                )

        def _is_drm_hwp(buf):
            """HWP 페이로드가 DRM/문서보안 래퍼(DOCUMENTSAFER·Fasoo·MarkAny 등)로 잠겼는지 시그니처로 판별."""
            try:
                head = buf[:8192]
                return any(sig in head for sig in
                           (b"DOCUMENTSAFER", b"Fasoo", b"MarkAny", b"\x00D\x00R\x00M", b"SoftCamp"))
            except Exception:
                return False

        # --- 1) 헤더 + 본문 → 결정적 텍스트 페이지 ---
        body_text = ""
        try:
            body_obj = msg.get_body(preferencelist=("plain", "html"))
        except Exception:
            body_obj = None
        if body_obj is not None:
            try:
                body_text = body_obj.get_content() or ""
            except Exception:
                body_text = ""
            if body_obj.get_content_type() == "text/html":
                body_text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", body_text)
                body_text = re.sub(r"(?i)<br\s*/?>", "\n", body_text)
                body_text = re.sub(r"(?i)</(p|div|tr|li|h[1-6])>", "\n", body_text)
                body_text = re.sub(r"<[^>]+>", " ", body_text)
                body_text = _html.unescape(body_text)
        body_text = re.sub(r"[ \t]+\n", "\n", (body_text or "").strip())
        # 더블스페이싱 제거 → 단일 간격.
        body_text = re.sub(r"\n[ \t]*(?:\n[ \t]*)+", "\n", body_text)
        _say("EML 본문 렌더 중...", 8)
        header_lines = []
        for label, field in (
            ("From", "From"),
            ("To", "To"),
            ("Cc", "Cc"),
            ("Subject", "Subject"),
            ("Date", "Date"),
        ):
            value = msg.get(field)
            if value:
                header_lines.append(f"{label}: {value}")
        cover_text = "\n".join(header_lines)
        if body_text:
            cover_text += ("\n\n" if cover_text else "") + body_text

        plain_pages = _paginate_plain_text(cover_text)
        if plain_pages:
            from PIL import (
                Image as _PILImage,
                ImageDraw as _ImageDraw,
                ImageFont as _ImageFont,
            )

            def _load_text_font(size):
                for candidate in (
                    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
                    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                ):
                    try:
                        return _ImageFont.truetype(candidate, size)
                    except OSError:
                        continue
                return _ImageFont.load_default()

            width, height = int(8.27 * RENDER_DPI), int(11.69 * RENDER_DPI)
            font = _load_text_font(26)
            for page_text in plain_pages:
                image = _PILImage.new("RGB", (width, height), "white")
                draw = _ImageDraw.Draw(image)
                y = 100
                for line in page_text.splitlines():
                    draw.text((100, y), line, font=font, fill=(25, 25, 25))
                    y += 36
                _emit_image(image, hint=page_text, direct_structured=True)

        # --- 2) 첨부 재귀 처리 ---
        IMG_EXT = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}
        DOC_EXT = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".odt", ".rtf", ".txt", ".csv"}
        for part in msg.walk():
            if part.is_multipart():
                continue
            fname = part.get_filename()
            disp = part.get_content_disposition()
            ctype = (part.get_content_type() or "").lower()
            if disp != "attachment" and not fname:
                continue   # 본문 파트(이미 커버에 반영)
            try:
                payload = part.get_payload(decode=True)
            except Exception:
                payload = None
            if not payload:
                continue
            aext = os.path.splitext(fname or "")[1].lower()
            if not aext and ctype.startswith("image/"):
                aext = "." + ctype.split("/")[-1]

            if aext in IMG_EXT or ctype.startswith("image/"):
                try:
                    from PIL import Image
                    im = Image.open(_io.BytesIO(payload)).convert("RGB")
                except Exception as e:
                    logger.warning(f"EML 첨부 이미지 열기 실패({fname}): {e}")
                    _emit_notice(["⚠ 첨부 이미지 열기 실패 — 원본 필요",
                                  f"파일명: {fname or '(이름없음)'}",
                                  f"오류: {str(e)[:120]}"])
                    continue
                if max(im.size) < 32:        # 1x1 트래킹 픽셀·스페이서 — VLM 낭비 없이 제외
                    continue
                # 바이트 임계값 대신 VLM 으로 로고/장식 판별(XLSX·HWP figure 경로와 동일 규약)
                desc = ""
                if api_key:
                    tmp = os.path.join(temp_pdf_dir, f"_emlimg_{uuid.uuid4().hex}.png")
                    try:
                        im.save(tmp, "PNG")
                        desc = VLMProcessor.describe_image(tmp, api_key, model_name) or ""
                    except Exception as e:
                        logger.warning(f"EML 첨부 이미지 설명 실패({fname}): {e}")
                    finally:
                        if os.path.exists(tmp):
                            os.remove(tmp)
                    if desc.strip().upper().startswith("LOGO"):
                        logger.info(f"EML 첨부 이미지 로고로 판별 → 스킵: {fname}")
                        continue
                _say(f"EML 첨부 이미지 편입: {fname}", 12)
                # The description is useful only for logo filtering. It is generated
                # text, not an OCR layer, so feeding it back as source text can turn
                # paraphrases into extraction errors.
                _emit_image(im, hint=f"[첨부 이미지: {fname or ''}]".strip())
                continue

            if aext in (".hwp", ".hwpx"):
                # DRM/문서보안 잠금 첨부는 어떤 엔진으로도 복호화 불가 → 렌더 시도 대신 마커로 가시화.
                if _is_drm_hwp(payload):
                    logger.info(f"EML 첨부 HWP DRM 잠금 감지 → 마커 페이지: {fname}")
                    _say(f"EML 첨부 HWP(DRM 보호): {fname}", 12)
                    _emit_notice(["🔒 DRM 보호 첨부 — 원본 필요",
                                  f"파일명: {fname or '(이름없음)'}",
                                  "문서보안(DOCUMENTSAFER 등)으로 잠겨 자동 파싱이 불가합니다.",
                                  "내용 확인은 원본 열람 권한이 필요합니다."])
                    continue
                _say(f"EML 첨부 HWP 렌더: {fname}", 12)
                sub = os.path.join(temp_pdf_dir, f"_att_{uuid.uuid4().hex}")
                os.makedirs(sub, exist_ok=True)
                ap = os.path.join(sub, os.path.basename(fname) or f"att{aext}")
                with open(ap, "wb") as fh:
                    fh.write(payload)
                emitted = 0
                try:
                    cnt = cls._render_hwp_with_rhwp(ap, sub)
                    for k in range(1, cnt + 1):
                        src_png = os.path.join(sub, f"page_{str(k).zfill(4)}.png")
                        if not os.path.exists(src_png):
                            continue
                        src_txt = os.path.join(sub, f"page_{str(k).zfill(4)}.txt")
                        state["n"] += 1
                        emitted += 1
                        dn = str(state["n"]).zfill(4)
                        shutil.move(src_png, os.path.join(doc_output_dir, f"page_{dn}.png"))
                        dst_txt = os.path.join(doc_output_dir, f"page_{dn}.txt")
                        if os.path.exists(src_txt):
                            shutil.move(src_txt, dst_txt)
                        else:
                            open(dst_txt, "w", encoding="utf-8").close()
                except Exception as e:
                    logger.warning(f"EML 첨부 HWP 렌더 실패({fname}): {e}")
                finally:
                    shutil.rmtree(sub, ignore_errors=True)
                # 렌더 결과가 0쪽이면(DRM 미탐지·손상 등) 누락 대신 마커로 가시화.
                if emitted == 0:
                    _emit_notice(["⚠ 첨부 HWP 파싱 실패 — 원본 필요",
                                  f"파일명: {fname or '(이름없음)'}",
                                  "문서보안 잠금 또는 손상으로 자동 렌더에 실패했습니다."])
                continue

            if aext in DOC_EXT:
                _say(f"EML 첨부 문서 렌더: {fname}", 12)
                ap = os.path.join(temp_pdf_dir, f"_att_{uuid.uuid4().hex}{aext}")
                with open(ap, "wb") as fh:
                    fh.write(payload)
                before_pages = state["n"]
                try:
                    if aext == ".pdf":
                        _emit_pdf(ap)
                    elif aext == ".xlsx":
                        pdfp, _sm = cls._convert_excel_to_pdf_with_sheet_map(ap, temp_pdf_dir)
                        if pdfp: _emit_pdf(pdfp)
                    else:
                        pdfp = cls.convert_to_pdf(ap, temp_pdf_dir)
                        if pdfp: _emit_pdf(pdfp)
                except Exception as e:
                    logger.warning(f"EML 첨부 변환 실패({fname}): {e}")
                finally:
                    if os.path.exists(ap): os.remove(ap)
                if state["n"] == before_pages:
                    _emit_notice(["⚠ 첨부 문서 변환 실패 — 원본 필요",
                                  f"파일명: {fname or '(이름없음)'}",
                                  "자동 렌더 결과 페이지가 생성되지 않았습니다."])
                continue

            logger.info(f"EML 첨부 미지원 건너뜀: {fname} ({ctype})")
            _emit_notice(["⚠ 미지원 첨부 — 원본 필요",
                          f"파일명: {fname or '(이름없음)'}",
                          f"Content-Type: {ctype or 'unknown'}"])

        if state["n"] == 0:
            _emit_notice(["⚠ EML 렌더 결과 없음",
                          f"파일명: {os.path.basename(file_path)}",
                          "본문 또는 첨부를 페이지로 생성하지 못했습니다. 원본 확인이 필요합니다."])
        return state["n"]

    @classmethod
    def convert_hwp_to_pdf_via_odt(cls, input_path, output_dir):
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        cmd_exe = cls.get_libreoffice_cmd()
        if not cmd_exe: return None

        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        temp_id = uuid.uuid4().hex
        abs_output_dir = os.path.abspath(output_dir)
        temp_odt_path = os.path.join(abs_output_dir, f"temp_{temp_id}.odt")
        final_pdf_path = os.path.join(abs_output_dir, f"{name}.pdf")

        try:
            result = subprocess.run(
                ['hwp5odt', '--output', temp_odt_path, input_path],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60
            )
            if result.returncode != 0 or not os.path.exists(temp_odt_path):
                logger.error(f"hwp5odt 변환 실패: {result.stderr.decode('utf-8', errors='replace')}")
                return None

            startupinfo = None
            if platform.system() == "Windows":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            result2 = subprocess.run(
                [cmd_exe, "--headless", "--convert-to", "pdf", "--outdir", abs_output_dir, temp_odt_path],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                startupinfo=startupinfo, timeout=90
            )
            if result2.returncode != 0:
                logger.error(f"ODT→PDF 변환 실패: {result2.stderr.decode('utf-8', errors='replace')}")
                return None

            odt_pdf = os.path.join(abs_output_dir, f"temp_{temp_id}.pdf")
            if os.path.exists(odt_pdf):
                if os.path.exists(final_pdf_path): os.remove(final_pdf_path)
                os.rename(odt_pdf, final_pdf_path)
                return final_pdf_path
            return None
        except Exception as e:
            logger.error(f"HWP→ODT→PDF 변환 예외: {e}")
            return None
        finally:
            if os.path.exists(temp_odt_path): os.remove(temp_odt_path)

    @staticmethod
    def _auto_click_hwp_popup():
        try:
            from pywinauto import Desktop
            import threading
            import time

            def clicker():
                for _ in range(30):
                    try:
                        dlg = Desktop(backend="win32").window(title_re=".*한글.*|.*보안.*", visible_only=True)
                        if dlg.exists():
                            btn = dlg.child_window(title_re=".*모두 허용.*|.*허용.*", class_name="Button")
                            if btn.exists():
                                btn.click()
                                break
                    except Exception: pass
                    time.sleep(0.5)

            t = threading.Thread(target=clicker)
            t.daemon = True
            t.start()
        except ImportError:
            logger.warning("pywinauto 패키지가 없습니다.")

    @classmethod
    def convert_hwp_to_pdf_win32(cls, input_path, output_dir):
        if platform.system() != "Windows": return None

        filename = os.path.basename(input_path)
        name, _ = os.path.splitext(filename)
        temp_id = uuid.uuid4().hex
        abs_output_dir = os.path.abspath(output_dir)
        temp_pdf_path = os.path.join(abs_output_dir, f"temp_{temp_id}.pdf")
        final_pdf_path = os.path.join(abs_output_dir, f"{name}.pdf")

        helper_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hwp_to_pdf.py")

        try:
            result = subprocess.run(
                [sys.executable, helper_script, input_path, temp_pdf_path],
                timeout=120, capture_output=True, text=True
            )
            if result.returncode != 0:
                logger.error(f"HWP 변환 subprocess 실패: {result.stderr.strip()}")
                return None

            if os.path.exists(temp_pdf_path):
                if os.path.exists(final_pdf_path): os.remove(final_pdf_path)
                os.rename(temp_pdf_path, final_pdf_path)
                return final_pdf_path
            return None
        except Exception as e:
            logger.error(f"HWP 변환 subprocess 예외: {e}")
            return None

    @staticmethod
    def json_to_markdown(structured_json_str: str) -> str:
        try:
            data = json.loads(structured_json_str)
        except json.JSONDecodeError:
            return structured_json_str
        elements = data.get("elements", [])
        types = [e.get("type", "text") for e in elements]
        has_toc = "toc_entry" in types
        has_heading = any(t in ("heading_1", "heading_2", "heading_3") for t in types)
        lines = []
        if has_toc and not has_heading:
            lines.append("# Table of Contents")
            lines.append("")
        for elem in elements:
            etype = elem.get("type", "text")
            content = (elem.get("content") or "").strip()
            caption = (elem.get("caption") or "").strip()
            description = (elem.get("description") or "").strip()

            if etype in ("heading_2", "heading_3"):
                mid_period = re.search(r'[a-zA-Z]\.\s+[A-Z]', content)
                if mid_period and len(content.split()) > 10:
                    split_pos = content.index(mid_period.group()[2:], mid_period.start())
                    bold_part = content[:mid_period.start() + 1]
                    rest_part = content[mid_period.start() + 2:].strip()
                    content = f"**{bold_part}** {rest_part}"
                    etype = "text"
                elif (content.endswith('.') and len(content.split()) <= 8
                      and not re.match(r'^[\d.]+\s', content)
                      and not re.match(r'^(Figure|Fig|Table|Eq)\b', content, re.IGNORECASE)):
                    content = f"**{content}**"
                    etype = "text"

            if etype == "heading_1":
                lines.append(f"# {content}")
            elif etype == "heading_2":
                lines.append(f"## {content}")
            elif etype == "heading_3":
                lines.append(f"### {content}")
            elif etype == "toc_entry":
                parts = content.split("::", 1)
                if len(parts) == 2:
                    lines.append(f"- {parts[0].strip()} ··· {parts[1].strip()}")
                else:
                    lines.append(f"- {content}")
            elif etype == "table":
                if caption:
                    lines.append(caption)
                if content:
                    lines.append(content)
            elif etype == "figure":
                if caption:
                    lines.append(f"*{caption}*")
                if description:
                    lines.append(f"> {description}")
                if content:
                    lines.append(content)
            elif etype == "footnote":
                if content:
                    lines.append(f"---\n> {content}")
            else:
                if content:
                    lines.append(content)
            lines.append("")
        return "\n".join(lines).strip()

    @classmethod
    def process_and_save(cls, file_path, base_output_dir, api_key=None, output_format="json", model_name="Qwen/Qwen3-VL-30B-A3B-Instruct", progress_callback=None, chunk_strategies=None, output_name=None, source_filename=None):
        filename = os.path.basename(source_filename or file_path)
        storage_filename = os.path.basename(file_path)
        name, _ = os.path.splitext(filename)
        ext = os.path.splitext(storage_filename)[1].lower()

        os.makedirs(base_output_dir, exist_ok=True)
        base_name = output_name or name or "document"
        base_name = re.sub(r'[\/\\\:\*\?\"\<\>\|\x00-\x1f\x7f]', '_', base_name).strip(' ._') or "document"
        # 같은 stem을 병렬 처리해도 서로의 출력 폴더를 지우지 않도록 새 폴더를 원자적으로 점유한다.
        for attempt in range(1000):
            candidate = base_name if attempt == 0 else f"{base_name}__{attempt + 1}"
            doc_output_dir = os.path.join(base_output_dir, candidate)
            try:
                os.makedirs(doc_output_dir)
                break
            except FileExistsError:
                continue
        else:
            candidate = f"{base_name}__{uuid.uuid4().hex[:8]}"
            doc_output_dir = os.path.join(base_output_dir, candidate)
            os.makedirs(doc_output_dir)

        def set_progress(msg, percent):
            logger.info(msg)
            if progress_callback: progress_callback({"msg": msg, "percent": percent})

        set_progress("문서 파싱 준비 중...", 2)

        temp_pdf_dir = os.path.join(doc_output_dir, "temp_pdf")
        if not os.path.exists(temp_pdf_dir): os.makedirs(temp_pdf_dir)
        excel_sheet_map = []

        if ext in ['.hwp', '.hwpx']:
            pdf_path = None
            rhwp_done = False
            if USE_RHWP and platform.system() != "Windows":
                try:
                    set_progress(f"{ext.upper()} rhwp 렌더 중...", 5)
                    op = cls._render_hwp_with_rhwp(file_path, doc_output_dir)
                    if op:
                        rhwp_done = True
                        set_progress(f"{ext.upper()} rhwp 렌더 완료 ({op}쪽)", 20)
                except Exception as e:
                    logger.warning(f"rhwp 렌더 실패, 폴백 사용: {e}")
            if rhwp_done:
                pass
            elif platform.system() == "Windows":
                set_progress(f"{ext.upper()} PDF 변환 중 (한글 COM)...", 5)
                pdf_path = cls.convert_hwp_to_pdf_win32(file_path, temp_pdf_dir)
            else:
                set_progress(f"{ext.upper()} PDF 변환 중 (LibreOffice+H2Orestart)...", 5)
                pdf_path = cls.convert_to_pdf(file_path, temp_pdf_dir)
            two_up_split = platform.system() == "Windows"

            if rhwp_done:
                pass
            elif pdf_path:
                try:
                    doc = fitz.open(pdf_path)
                    out_page = 0
                    mat = fitz.Matrix(RENDER_DPI / 72, RENDER_DPI / 72)
                    for i in range(len(doc)):
                        page = doc[i]
                        w, h = page.rect.width, page.rect.height
                        is_two_up = two_up_split and w > h
                        halves = (
                            [(0, 0, w / 2, h), (w / 2, 0, w, h)] if is_two_up
                            else [(0, 0, w, h)]
                        )
                        for x0, y0, x1, y1 in halves:
                            out_page += 1
                            pnum = str(out_page).zfill(4)
                            clip = fitz.Rect(x0, y0, x1, y1)
                            pix = page.get_pixmap(matrix=mat, clip=clip)
                            pix.save(os.path.join(doc_output_dir, f"page_{pnum}.png"))
                            words = page.get_text("words")
                            region_words = [
                                wd for wd in words
                                if wd[0] >= x0 - 5 and wd[2] <= x1 + 5
                            ]
                            text = " ".join(
                                wd[4] for wd in sorted(region_words, key=lambda r: (r[1], r[0]))
                            )
                            cleaned_text = HWPTextExtractor.text_preprocessing(text.strip())
                            with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as f:
                                f.write(cleaned_text)
                    doc.close()
                    set_progress(f"{ext.upper()} 파싱 완료 ({out_page}쪽)", 20)
                except Exception as e:
                    raise Exception(f"{ext.upper()} PDF 변환 후 처리 실패: {e}")
            else:
                raise Exception(f"{ext.upper()} → PDF 변환 실패. (Windows: 한글 프로그램 / Linux: LibreOffice+H2Orestart 필요)")

        elif ext == '.eml':
            set_progress("EML 파싱 중 (본문 디코딩 + 첨부 렌더)...", 5)
            try:
                n_eml = cls._render_eml(
                    file_path, doc_output_dir, temp_pdf_dir, api_key, model_name, set_progress,
                    output_format=output_format)
            except Exception as e:
                raise Exception(f"EML 처리 실패: {e}")
            set_progress(f"EML 파싱 완료 ({n_eml}쪽)", 20)

        else:
            set_progress("일반 문서 시각 레이아웃 생성 중...", 10)
            if ext == '.pdf':
                pdf_path = file_path
            elif ext == '.xlsx':
                pdf_path, excel_sheet_map = cls._convert_excel_to_pdf_with_sheet_map(file_path, temp_pdf_dir)
            else:
                pdf_path = cls.convert_to_pdf(file_path, temp_pdf_dir)

            if pdf_path:
                try:
                    doc = fitz.open(pdf_path)
                    total_pages = len(doc)
                    for i in range(total_pages):
                        pnum = str(i + 1).zfill(4)
                        pix = doc[i].get_pixmap(dpi=RENDER_DPI)
                        pix.save(os.path.join(doc_output_dir, f"page_{pnum}.png"))

                        raw_text = doc[i].get_text().strip()
                        cleaned_text = HWPTextExtractor.text_preprocessing(raw_text)
                        with open(os.path.join(doc_output_dir, f"page_{pnum}.txt"), "w", encoding="utf-8") as f:
                            f.write(cleaned_text)
                    set_progress(f"문서 파싱 완료 ({total_pages}쪽)", 20)
                    doc.close()
                except Exception as e:
                    raise Exception(f"문서 파싱 에러: {e}")

        page_files = sorted(f for f in os.listdir(doc_output_dir) if f.startswith("page_") and f.endswith(".txt"))
        total_vlm_pages = len(page_files)

        if not page_files:
            set_progress("❌ 분석할 페이지를 생성하지 못했습니다.", 100)
            raise Exception("분석할 페이지를 생성하지 못했습니다. 입력 파일 변환 또는 렌더링에 실패했습니다.")
        if not api_key or "여기에" in api_key: raise Exception("API 키가 누락되었거나 유효하지 않습니다.")

        set_progress("문서 메타데이터 추출 중...", 22)
        meta_txts = [os.path.join(doc_output_dir, f"page_{str(i).zfill(4)}.txt") for i in range(1, 3)]
        meta_imgs = [os.path.join(doc_output_dir, f"page_{str(i).zfill(4)}.png") for i in range(1, 3)]
        metadata = VLMProcessor.extract_metadata(meta_txts, meta_imgs, api_key, model_name)
        if not isinstance(metadata, dict):
            metadata = {}
        metadata["source_file"] = filename
        if excel_sheet_map:
            metadata["sheets"] = excel_sheet_map
        with open(os.path.join(doc_output_dir, "metadata.json"), "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        set_progress(f"🤖 AI 분석 대기 중 (총 {total_vlm_pages}페이지)...", 25)

        # XLSX: 시트 데이터를 직접 추출해 structured.json 기록(아래 VLM 루프가 건너뜀).
        if ext == '.xlsx' and excel_sheet_map:   # .xls 는 일반 VLM 경로
            try:
                set_progress("📊 엑셀 시트 직접 추출 중...", 30)
                cls._xlsx_hybrid_write(doc_output_dir, file_path, excel_sheet_map, api_key, model_name,
                                       output_format=output_format)
            except Exception as e:
                logger.warning(f"XLSX 하이브리드 실패 → 일반 VLM 폴백: {e}")

        failed_pages = []   # VLM 구조추출이 (재시도 후에도) 실패한 페이지 — 부분 실패 가시화용
        low_conf_pages = []  # 이미지엔 내용이 있으나 추출이 빈약(저신뢰)한 페이지 — silent-drop 가시화용
        vlm_fmt = "json" if output_format.lower() == "markdown" else output_format

        def _persist_page(stem, structured_data):
            """추출 구조를 후처리(page_number·중복제거·저신뢰탐지·PII) 후 파일로 기록(메인/최종재시도 공용)."""
            if vlm_fmt == "json":
                try:
                    parsed = json.loads(structured_data)
                    try:
                        parsed["page_number"] = int(stem.split("_")[1])
                    except (IndexError, ValueError):
                        pass
                    # 중복 제거(완전일치 + substring 포함). 본문성(text/footnote)만 대상.
                    _SUBSTR_DROPPABLE = {"text", "footnote"}
                    elems = [e for e in parsed.get("elements", []) if isinstance(e, dict)]
                    elems = _resolve_flattened_table_duplicates(elems)
                    try:
                        with open(
                            os.path.join(doc_output_dir, f"{stem}.txt"),
                            "r",
                            encoding="utf-8",
                        ) as text_layer_file:
                            source_text_layer = text_layer_file.read()
                    except OSError:
                        source_text_layer = ""
                    elems = _restore_leading_source_blocks(
                        elems, source_text_layer
                    )
                    elems = _dedupe_tables_supported_by_text_layer(
                        elems, source_text_layer
                    )
                    elems = _dedupe_figures_supported_by_text_layer(
                        elems, source_text_layer
                    )
                    elems = _dedupe_headings_supported_by_text_layer(
                        elems, source_text_layer
                    )
                    elems = _drop_prose_duplicated_by_nearby_figures(
                        elems, source_text_layer
                    )
                    norm = [re.sub(r"\s+", " ", (e.get("content") or "")).strip() for e in elems]
                    drop = [False] * len(elems)
                    for i in range(len(elems)):
                        if drop[i] or len(norm[i]) < 12:
                            continue
                        for j in range(len(elems)):
                            if i == j or drop[j] or len(norm[j]) < 12:
                                continue
                            if elems[j].get("type", "text") not in _SUBSTR_DROPPABLE:
                                continue   # heading/table/figure/toc 는 substring 제거 면제
                            # j 가 i 에 포함되고 i 가 더 길면(또는 동일·앞선 블록) j 를 중복으로 제거
                            if norm[j] in norm[i] and (len(norm[j]) < len(norm[i]) or j > i):
                                drop[j] = True
                    # 근접중복 제거: 서식 문자는 무시하되 주소·URL의 영숫자
                    # 정체성은 보존해 서로 다른 연락처를 합치지 않는다.
                    nnorm = [_dedupe_comparison_text(n) for n in norm]
                    for i in range(len(elems)):
                        if drop[i] or elems[i].get("type", "text") not in _SUBSTR_DROPPABLE or len(nnorm[i]) < 16:
                            continue
                        for j in range(i + 1, len(elems)):
                            if drop[j] or elems[j].get("type", "text") not in _SUBSTR_DROPPABLE or len(nnorm[j]) < 16:
                                continue
                            a, b = nnorm[i], nnorm[j]
                            if a == b or a in b or b in a:
                                if len(norm[j]) <= len(norm[i]):   # 더 완전한 쪽(원문 긴 쪽) 유지
                                    drop[j] = True
                                else:
                                    drop[i] = True
                                    break
                    deduped = _dedupe_exact_text_elements(
                        [e for k, e in enumerate(elems) if not drop[k]]
                    )
                    deduped = _drop_trailing_duplicate_heading_cluster(deduped)
                    # TOC sanity cleanup: VLM sometimes assigns the first child page number to a parent/group
                    # label ("I. ...::7") even when no page number is printed on that line. Keep only page
                    # numbers supported by the text-layer leader/page pattern; demote unsupported group labels.
                    try:
                        if any(e.get("type") == "toc_entry" for e in deduped):
                            def _toc_norm(s):
                                s = re.sub(r"\s+", " ", (s or "")).strip()
                                return s.rstrip(".·•∙⋯…").strip()

                            def _leader_line(s):
                                compact = re.sub(r"\s+", "", s or "")
                                return bool(compact) and bool(re.fullmatch(r"[.\u2026·•∙⋯…_\-]{3,}", compact))

                            def _page_line(s):
                                return bool(re.fullmatch(r"\d{1,4}", (s or "").strip()))

                            def _looks_toc_group(title):
                                title = (title or "").strip()
                                return bool(
                                    re.match(r"^(?:[IVXLCDM]+|[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+)[.)]?\s+", title, re.I) or
                                    re.match(r"^제\s*\d+\s*[장부편]\b", title)
                                )

                            raw_lines = []
                            tp = os.path.join(doc_output_dir, f"{stem}.txt")
                            if os.path.exists(tp):
                                with open(tp, "r", encoding="utf-8") as _tf:
                                    raw_lines = [ln.strip() for ln in _tf.read().splitlines() if ln.strip()]

                            visible_pairs = set()
                            inline_re = re.compile(r"^(?P<title>.+?)[.\u2026·•∙⋯…_ \t-]{3,}(?P<page>\d{1,4})$")
                            for idx, line in enumerate(raw_lines):
                                m = inline_re.match(line)
                                if m:
                                    visible_pairs.add((_toc_norm(m.group("title")), m.group("page")))
                                if idx + 2 < len(raw_lines) and _leader_line(raw_lines[idx + 1]) and _page_line(raw_lines[idx + 2]):
                                    visible_pairs.add((_toc_norm(line), raw_lines[idx + 2].strip()))

                            if visible_pairs:
                                cleaned_toc = []
                                for e in deduped:
                                    if e.get("type") != "toc_entry":
                                        cleaned_toc.append(e)
                                        continue
                                    raw = (e.get("content") or "").strip()
                                    title, sep, pg = raw.rpartition("::")
                                    title = title.strip() if sep else raw
                                    pg = pg.strip() if sep else ""
                                    if _looks_toc_group(title) and (not sep or (_toc_norm(title), pg) not in visible_pairs):
                                        fixed = {k: v for k, v in e.items() if k not in ("caption", "description")}
                                        fixed["type"] = "heading_2"
                                        fixed["content"] = title
                                        cleaned_toc.append(fixed)
                                    else:
                                        cleaned_toc.append(e)
                                deduped = cleaned_toc
                    except Exception as te:
                        logger.warning(f"TOC sanity cleanup 실패({stem}): {te}")
                    try:
                        _page_no_for_artifacts = parsed.get("page_number")
                        deduped = _drop_page_artifact_elements(deduped, _page_no_for_artifacts)
                        deduped = _backfill_table_captions(deduped)
                        deduped = _attach_adjacent_table_unit_captions(deduped)
                    except Exception as ae:
                        logger.warning(f"페이지 노이즈/표 캡션 보정 실패({stem}): {ae}")
                    # silent-drop 탐지: 이미지엔 잉크가 있는데 추출 본문이 페이지번호/푸터 수준이면 저신뢰 표기.
                    alltext = "".join((e.get("content") or "") + (e.get("caption") or "") +
                                      (e.get("description") or "") for e in deduped)
                    core = re.sub(r"\d+", "", re.sub(r"[\s\-–—·.~_=|]", "", alltext))
                    if len(core) < 8:
                        png = os.path.join(doc_output_dir, f"{stem}.png")
                        if _image_has_ink(png):
                            parsed["warning"] = "자동 추출 불완전 — 이 페이지는 이미지(표·도해·캡처)로 구성되어 본문이 거의 추출되지 않았습니다. 원본 확인 필요."
                            parsed["low_confidence"] = True
                            try: low_conf_pages.append(int(stem.split("_")[1]))
                            except (IndexError, ValueError): low_conf_pages.append(stem)
                    # 표 검증·수리(나란한표 분할·column-bleed 병합·열수 패딩). element 1개가 2개로 분할될 수 있음.
                    try:
                        import table_validate
                        # HWP 네이티브 표 로드(있으면 VLM 이미지표를 정확한 colspan/rowspan 으로 교체)
                        _native_prep = []
                        _np = os.path.join(doc_output_dir, "_native_tables.json")
                        if os.path.exists(_np):
                            try:
                                with open(_np, encoding="utf-8") as _nf:
                                    _native_prep = table_validate.prepare_native(json.load(_nf))
                            except Exception:
                                _native_prep = []
                        deduped = table_validate.restore_uniquely_supported_native_parents(
                            deduped, _native_prep, source_text_layer
                        )
                        deduped = table_validate.merge_adjacent_native_table_fragments(
                            deduped, _native_prep
                        )
                        rebuilt = []
                        for e in deduped:
                            if e.get("type") == "table" and e.get("content"):
                                if e.get("_native"):
                                    q = table_validate.assess_table_quality(
                                        e["content"], e.get("caption"), allow_nested=True
                                    )
                                    rebuilt.append(_mark_element(
                                        e, "native_table",
                                        min(0.99, q.get("confidence", 0.99)),
                                        q.get("issues")))
                                    continue
                                # 네이티브 우선: 내용일치하면 정답구조로 교체(검증/수리 생략 — 이미 정확).
                                _nh = (
                                    table_validate.native_substitute_for_source_page(
                                        e["content"], _native_prep, source_text_layer
                                    )
                                    if _native_prep else None
                                )
                                if _nh and _nh.strip() and "</table>" in _nh:
                                    _nh = table_validate.strip_caption_duplicate_metadata_row(
                                        _nh, e.get("caption")
                                    )
                                    q = table_validate.assess_table_quality(
                                        _nh, e.get("caption"), allow_nested=True
                                    )
                                    rebuilt.append(_mark_element(
                                        {**e, "content": _nh, "_native": True},
                                        "native_table", min(0.99, q.get("confidence", 0.99)),
                                        q.get("issues")))
                                    continue
                                new_els, _retry, _iss = table_validate.validate_and_repair_table(
                                    e["content"], e.get("caption"))
                                for ne in new_els:
                                    m = {**e, "content": ne["content"]}
                                    if ne.get("caption") is not None:
                                        m["caption"] = ne["caption"]
                                    q = table_validate.assess_table_quality(m.get("content"), m.get("caption"))
                                    src = "vlm_table_repaired" if _iss else "vlm_table"
                                    conf = min(q.get("confidence", 0.85), 0.80 if _iss else 0.88)
                                    issues = list(_iss or []) + list(q.get("issues") or [])
                                    _mark_element(m, src, conf, issues)
                                    rebuilt.append(m)
                            else:
                                rebuilt.append(e)
                        deduped = _drop_prose_duplicated_by_nearby_tables(rebuilt)
                        deduped = _dedupe_exact_text_elements(deduped)
                        # Early VLM text can temporarily duplicate native outer-cell
                        # labels and block the conservative parent restore. Retry only
                        # after table/prose deduplication sees the final representation.
                        deduped = table_validate.restore_uniquely_supported_native_parents(
                            deduped, _native_prep, source_text_layer
                        )
                    except Exception as te:
                        logger.warning(f"표 검증 실패({stem}): {te}")
                    # figure description 폴백(결정적): 빈 description 으로 RAG 앵커가 사라지지 않게 보정.
                    #   - 산문이 content 에 잘못 들어간 경우 → description 으로 승급.
                    #   - HTML/캡처면 caption 또는 ⚠ 마커로 폴백.
                    #   - 한글문서인데 description 이 외국어(인용 라벨 제외 한글비율 낮음)면 ⚠ 표기(프롬프트 1차 + 코드 하한).
                    try:
                        _page_kr = _korean_ratio("".join((e.get("content") or "")
                                    for e in deduped if e.get("type") != "figure"))
                        for e in deduped:
                            if e.get("type") != "figure":
                                continue
                            desc = (e.get("description") or "").strip()
                            # figure 설명 길이 상한 — 초과 시 앞부분만 남기고 절단.
                            if len(desc) > 1200:
                                _cap = (e.get("caption") or "").strip()
                                _head = re.split(r"[.\n]", desc, 1)[0][:160].strip()
                                desc = ((_cap + " — ") if _cap else "") + (_head or "이미지") + " ⚠(설명 비정상 생성으로 절단)"
                                e["description"] = desc
                            if not desc:
                                content = (e.get("content") or "").strip()
                                cap = (e.get("caption") or "").strip()
                                if content and "<" not in content and len(content) >= 12:
                                    desc = content            # 산문 시각설명이 content 로 잘못 간 경우 구제
                                elif cap:
                                    desc = cap
                                else:
                                    desc = "⚠ 설명 누락 — 원본 캡처 확인 필요"
                                e["description"] = desc
                            if _page_kr >= 0.30 and desc and not desc.startswith("⚠"):
                                _core = re.sub(r"[\"'].*?[\"']", "", desc)   # 인용된 외국어 라벨 제외
                                if _korean_ratio(_core) < 0.20:
                                    e["description"] = "⚠ 외국어 설명 — " + desc
                    except Exception as fe:
                        logger.warning(f"figure description 폴백 실패({stem}): {fe}")
                    # figure 캡션 교정(경계 분리) + 순서 교정(IR 앵커 기반).
                    try:
                        _fa = os.path.join(doc_output_dir, "_figure_anchors.json")
                        if os.path.exists(_fa) and any(e.get("type") == "figure" for e in deduped):
                            with open(_fa, encoding="utf-8") as _af:
                                _anchors = json.load(_af)
                            deduped = _correct_figure_captions(deduped, _anchors)
                            deduped = _reposition_figures_by_anchor(deduped, _anchors)
                    except Exception as re_e:
                        logger.warning(f"figure 교정 실패({stem}): {re_e}")
                    # Short VLM title/header duplicates such as "참고참고1717" are
                    # mechanical repeats, not meaningful document text.
                    for e in deduped:
                        et = e.get("type", "text")
                        for key in ("content", "caption"):
                            val = e.get(key)
                            if val and (key == "caption" or et.startswith(("text", "heading", "toc"))):
                                e[key] = _collapse_compound_repeat_text(val)
                        if not e.get("_source"):
                            if e.get("salvaged"):
                                _mark_element(e, "hwp_embedded_figure", 0.78)
                            elif e.get("_zoom"):
                                _mark_element(e, "zoom_table", 0.88)
                            elif et == "figure":
                                _mark_element(e, "vlm_figure", 0.82)
                            elif et == "table":
                                _mark_element(e, "vlm_table", 0.85)
                            else:
                                _mark_element(e, "vlm_page", 0.88)
                    # PII 마스킹(사용자 지정 타입만; 기본 OFF)
                    if PII_MASK_TYPES:
                        for e in deduped:
                            for key in ("content", "caption", "description"):
                                if e.get(key):
                                    e[key] = mask_pii(e[key])
                    parsed["elements"] = deduped
                    structured_data = json.dumps(parsed, ensure_ascii=False, indent=4)
                except Exception as e:
                    logger.warning(f"JSON 후처리 실패 ({stem}): {e}")
            if output_format.lower() == "markdown":
                with open(os.path.join(doc_output_dir, f"{stem}_structured.json"), "w", encoding="utf-8") as f:
                    f.write(structured_data)
                md_data = DocumentProcessor.json_to_markdown(structured_data)
                with open(os.path.join(doc_output_dir, f"{stem}_structured.md"), "w", encoding="utf-8") as f:
                    f.write(md_data)
            else:
                with open(os.path.join(doc_output_dir, f"{stem}_structured.{output_format.lower()}"), "w", encoding="utf-8") as f:
                    f.write(structured_data)

        def _extract_multicol(stem, img_path, frac):
            """다단 페이지: 좌/우 열을 따로 추출해 순서대로(좌→우) 합쳐 읽기순서 보장. 실패 시 None."""
            try:
                from PIL import Image
                with Image.open(img_path) as im:
                    W, H = im.size
                    gx = int(W * frac); pad = int(W * 0.012)
                    crops = [im.crop((0, 0, min(gx + pad, W), H)), im.crop((max(gx - pad, 0), 0, W, H))]
                columns = []
                for ci, crop in enumerate(crops):
                    tmp = os.path.join(doc_output_dir, f".{stem}_col{ci}_{uuid.uuid4().hex}.png")
                    crop.save(tmp, "PNG")
                    try:
                        sd, _ = VLMProcessor.extract_structure(
                            txt_path=None, img_path=tmp, api_key=api_key,
                            output_format=vlm_fmt, model_name=model_name)
                    finally:
                        try: os.remove(tmp)
                        except OSError: pass
                    if not sd:
                        return None
                    pj = json.loads(sd)
                    columns.append(
                        pj if isinstance(pj, list) else pj.get("elements", [])
                    )
                try:
                    with open(
                        os.path.join(doc_output_dir, f"{stem}.txt"),
                        encoding="utf-8",
                        errors="replace",
                    ) as source_file:
                        source_text = source_file.read()
                except OSError:
                    source_text = ""
                columns = _drop_clipped_multicolumn_header_fragments(
                    columns, source_text
                )
                merged = _merge_multicolumn_elements(columns)
                if not merged:
                    return None
                pn = int(stem.split("_")[1]) if "_" in stem else 0
                return json.dumps({"page_number": pn, "elements": merged}, ensure_ascii=False)
            except Exception as e:
                logger.warning(f"{stem} 다단 분리추출 실패 → 전체페이지 폴백: {e}")
                return None

        def _extract_page(stem):
            """한 페이지 추출 + 성공 시 기록. (성공여부, 재시도가치) 반환.
            재시도가치=False 면 일시적 오류가 아니라(반복 생성·타임아웃·유효성 실패) 최종재시도 무의미."""
            img_path = os.path.join(doc_output_dir, f"{stem}.png")
            has_img = os.path.exists(img_path)
            # 다단 페이지면 열별 분리추출 우선(읽기순서 보장). 실패하면 전체페이지 추출로 폴백.
            if has_img and VLM_MULTICOL:
                frac = _detect_column_split(img_path)
                if frac:
                    sd = _extract_multicol(stem, img_path, frac)
                    if sd:
                        _persist_page(stem, sd); return True, True
            sd, retryable = VLMProcessor.extract_structure(
                txt_path=os.path.join(doc_output_dir, f"{stem}.txt"),
                img_path=img_path if has_img else None,
                api_key=api_key, output_format=vlm_fmt, model_name=model_name)
            if sd:
                _persist_page(stem, sd)
                return True, True
            return False, retryable

        from concurrent.futures import ThreadPoolExecutor, as_completed
        retry_eligible = set()        # 일시적 실패라 최종재시도 가치가 있는 페이지 번호
        budget_exceeded = False       # 문서 VLM 시간예산 초과(이후 페이지는 텍스트 폴백)
        _vlm_start = time.monotonic()

        # 처리 대상(이미 처리된 XLSX 하이브리드 페이지 제외)
        pending = [tf[:-4] for tf in page_files
                   if not os.path.exists(os.path.join(doc_output_dir, f"{tf[:-4]}_structured.json"))]
        total_vlm_analyzed_pages = len(pending)

        def _run_page(stem):
            # 예산 초과면(제출됐어도) 즉시 폴백 처리
            if DOC_VLM_BUDGET_SEC > 0 and (time.monotonic() - _vlm_start) > DOC_VLM_BUDGET_SEC:
                return False, False, True
            ok, retryable = _extract_page(stem)
            return ok, retryable, False

        # 페이지 동시 추출 → vLLM 배칭으로 GPU 포화(순차 대비 수 배 가속).
        # 집계(failed_pages·retry_eligible)는 메인 스레드의 as_completed 루프에서만 수행 → 락 불필요.
        # 각 워커는 자기 페이지의 {stem}_structured.json 만 기록(파일 비공유) → 스레드 안전.
        n_pending = len(pending)
        workers = max(1, min(VLM_PAGE_CONCURRENCY, n_pending or 1))
        done_cnt = 0
        if n_pending == 0:
            set_progress("⏭️ VLM 구조추출 대상 페이지 없음", 95)
        else:
            set_progress(f"⏳ AI 모델 분석 중... (0/{n_pending} 쪽, 동시 {workers})", 25)
            with ThreadPoolExecutor(max_workers=workers) as _ex:
                _futs = {_ex.submit(_run_page, s): s for s in pending}
                for fut in as_completed(_futs):
                    stem = _futs[fut]
                    done_cnt += 1
                    try:
                        ok, retryable, skipped = fut.result()
                    except Exception as e:   # 워커 예외(파일쓰기 등)가 문서 전체를 날리지 않게 페이지 단위 격리
                        logger.warning(f"{stem} 페이지 워커 예외: {e}")
                        ok, retryable, skipped = False, False, False
                    done_pct = 25 + int((done_cnt / max(n_pending, 1)) * 75)
                    if skipped:
                        if not budget_exceeded:
                            logger.error(f"문서 VLM 시간예산({DOC_VLM_BUDGET_SEC}s) 초과 — 남은 페이지는 텍스트 폴백")
                        budget_exceeded = True
                        try: failed_pages.append(int(stem.split("_")[-1]))
                        except ValueError: failed_pages.append(stem)
                        continue
                    if ok:
                        set_progress(f"✅ {done_cnt}/{n_pending}쪽 분석 완료", done_pct)
                    else:
                        try:
                            pn = int(stem.split("_")[-1]); failed_pages.append(pn)
                            if retryable: retry_eligible.add(pn)
                        except ValueError:
                            failed_pages.append(stem)
                        logger.warning(f"{stem} AI 분석 실패 ({'최종 재시도 대상' if retryable else '재시도 무의미'})")
                        set_progress(f"⚠️ {stem} 분석 실패", done_pct)

        # 최종 재시도: 부하가 빠진 뒤 '일시적' 실패 페이지만 1회 더 시도해 복구한다.
        # 반복 생성·타임아웃·유효성 실패(retry_eligible 아님)는 동일 입력 재시도가 무의미하므로 제외(런어웨이 방지).
        retry = [pn for pn in failed_pages if isinstance(pn, int) and pn in retry_eligible]
        if retry and not budget_exceeded:
            set_progress(f"🔁 실패 {len(retry)}쪽 최종 재시도 중...", 96)
            rworkers = max(1, min(VLM_PAGE_CONCURRENCY, len(retry)))
            recovered = []
            with ThreadPoolExecutor(max_workers=rworkers) as _ex:
                _rf = {_ex.submit(_extract_page, f"page_{pn:04d}"): pn for pn in retry}
                for fut in as_completed(_rf):
                    try:
                        if fut.result()[0]:
                            recovered.append(_rf[fut])
                    except Exception:
                        pass
            if recovered:
                logger.info(f"최종 재시도로 {len(recovered)}쪽 구조 복구: {sorted(recovered)}")
            failed_pages = [pn for pn in failed_pages if pn not in recovered]

        def _page_num_from_stem(stem):
            try:
                return int(stem.split("_")[1])
            except (IndexError, ValueError):
                return 0

        def _clean_fallback_text(raw):
            def _collapse_unit(s):
                t = s.strip()
                if not t:
                    return s
                max_unit = min(40, len(t) // 2)
                for n in range(1, max_unit + 1):
                    if len(t) % n == 0:
                        unit = t[:n]
                        if unit and unit * (len(t) // n) == t:
                            return s[:len(s) - len(s.lstrip())] + unit + s[len(s.rstrip()):]
                return s

            cleaned = []
            prev = None
            for line in (raw or "").splitlines():
                line = _collapse_unit(line)
                key = re.sub(r"\s+", "", line)
                if key and key == prev:
                    continue
                cleaned.append(line)
                prev = key if key else None
            return "\n".join(cleaned).strip()

        def _fallback_payload(stem):
            tp = os.path.join(doc_output_dir, f"{stem}.txt")
            raw = ""
            if os.path.exists(tp):
                with open(tp, "r", encoding="utf-8") as f:
                    raw = f.read().strip()
            raw = _clean_fallback_text(raw)
            paras = [p.strip() for p in re.split(r"\n\s*\n", raw) if p.strip()] or ([raw] if raw else [])
            els = [_mark_element({"type": "text", "content": p}, "text_layer_fallback", 0.45)
                   for p in paras]
            return raw, {
                "page_number": _page_num_from_stem(stem),
                "elements": els,
                "low_confidence": True,
                "warning": "구조추출 미완 — 텍스트레이어 기반 폴백(레이아웃·표 구조 없음). 원본 확인 권장.",
            }

        def _write_fallback_page(stem):
            raw, payload = _fallback_payload(stem)
            pn = payload["page_number"]
            if vlm_fmt == "json":
                _persist_page(stem, json.dumps(payload, ensure_ascii=False))
            elif vlm_fmt == "xml":
                import html as _html
                parts = [
                    "<document>",
                    f"  <page_number>{pn}</page_number>",
                    "  <elements>",
                ]
                paras = [p.strip() for p in re.split(r"\n\s*\n", raw) if p.strip()] or ([raw] if raw else [])
                for para in paras:
                    parts.append(f"    <element type=\"text\">{_html.escape(para)}</element>")
                parts.extend([
                    "  </elements>",
                    "  <low_confidence>true</low_confidence>",
                    "  <warning>구조추출 미완 — 텍스트레이어 기반 폴백(레이아웃·표 구조 없음). 원본 확인 권장.</warning>",
                    "</document>",
                ])
                with open(os.path.join(doc_output_dir, f"{stem}_structured.xml"), "w", encoding="utf-8") as f:
                    f.write("\n".join(parts))
            else:
                with open(os.path.join(doc_output_dir, f"{stem}_structured.{output_format.lower()}"), "w", encoding="utf-8") as f:
                    f.write(raw)
            if isinstance(pn, int):
                if pn not in low_conf_pages:
                    low_conf_pages.append(pn)
                if pn and pn not in failed_pages:
                    failed_pages.append(pn)
            logger.info(f"{stem}: 텍스트레이어 폴백으로 structured 합성")

        # 재시도 후에도 structured 결과가 없는 페이지는 포맷별 fallback을 생성해 누락을 막되,
        # metadata/status에는 VLM 구조추출 실패 페이지로 남긴다.
        for stem in [tf[:-4] for tf in page_files]:
            if output_format.lower() == "markdown":
                exists = (
                    os.path.exists(os.path.join(doc_output_dir, f"{stem}_structured.json")) and
                    os.path.exists(os.path.join(doc_output_dir, f"{stem}_structured.md"))
                )
            else:
                exists = os.path.exists(os.path.join(doc_output_dir, f"{stem}_structured.{output_format.lower()}"))
            if not exists:
                _write_fallback_page(stem)

        # Objective table failures (ragged/collapsed grouped headers) get one
        # image-grounded repair pass. Strict coverage and table-count gates in
        # table_quality_repair prevent a prettier but lossy candidate replacing
        # the original page.
        if api_key and output_format.lower() in ("json", "markdown"):
            try:
                from table_quality_repair import repair_low_quality_pages

                quality_repairs = repair_low_quality_pages(
                    doc_output_dir, api_key, model_name, _persist_page
                )
                accepted_repairs = [item["page"] for item in quality_repairs if item.get("accepted")]
                rejected_repairs = [item["page"] for item in quality_repairs if not item.get("accepted")]
                if accepted_repairs:
                    logger.info(f"표 품질 재추출 적용: {accepted_repairs}")
                if rejected_repairs:
                    logger.warning(f"표 품질 재추출 보류(무손실 게이트 미통과): {rejected_repairs}")
            except Exception as e:
                logger.warning(f"표 품질 재추출 실패: {e}")

        # Dense card/panel pages can preserve every character while assigning a
        # bullet to the wrong visual block. Review only a small raster-selected
        # subset and accept inventory-preserving text reassignment exclusively.
        if api_key and output_format.lower() in ("json", "markdown"):
            try:
                from layout_consistency_repair import repair_layout_consistency

                layout_repairs = repair_layout_consistency(
                    doc_output_dir, api_key, model_name, _persist_page
                )
                accepted_layout = [
                    item["page"] for item in layout_repairs if item.get("accepted")
                ]
                rejected_layout = [
                    item["page"]
                    for item in layout_repairs
                    if (item.get("review") or {}).get("pass") is False
                    and not item.get("accepted")
                ]
                if accepted_layout:
                    logger.info(f"패널 텍스트 재배치 적용: {accepted_layout}")
                if rejected_layout:
                    logger.warning(
                        f"패널 텍스트 재배치 보류(보존 게이트 미통과): {rejected_layout}"
                    )
            except Exception as e:
                logger.warning(f"패널 레이아웃 재검증 실패: {e}")

        if output_format.lower() in ("json", "markdown"):
            toc_entries = []
            toc_ended = False
            for fname in sorted(f for f in os.listdir(doc_output_dir) if f.endswith("_structured.json")):
                if toc_ended:
                    break
                fpath = os.path.join(doc_output_dir, fname)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    page_has_toc = False
                    for elem in data.get("elements", []):
                        if elem.get("type") == "toc_entry":
                            page_has_toc = True
                            raw = elem.get("content", "")
                            if "::" in raw:
                                title, _, pg = raw.rpartition("::")
                                try:
                                    toc_entries.append({"title": title.strip(), "page": int(pg.strip())})
                                except ValueError:
                                    toc_entries.append({"title": title.strip(), "page": pg.strip()})
                            else:
                                toc_entries.append({"title": raw.strip(), "page": None})
                    if not page_has_toc and toc_entries:
                        toc_ended = True
                except Exception as e:
                    logger.warning(f"TOC 수집 중 {fname} 읽기 실패: {e}")

            if toc_entries:
                meta_path = os.path.join(doc_output_dir, "metadata.json")
                try:
                    with open(meta_path, "r", encoding="utf-8") as f:
                        metadata = json.load(f)
                except Exception:
                    metadata = {}
                metadata["toc"] = toc_entries
                with open(meta_path, "w", encoding="utf-8") as f:
                    json.dump(metadata, f, ensure_ascii=False, indent=2)
                logger.info(f"TOC {len(toc_entries)}개 항목 수집 완료")

        # 시각자료 salvage: 렌더가 잘라먹을 수 있는 임베디드 이미지를 BinData 원본에서
        # reading-order 위치/크기와 함께 복원 + VLM 설명 → figures.json.
        # rhwp 렌더(rhwp_done)는 이미지를 페이지에 온전히 그려 VLM 이 직접 잡으므로 salvage 불필요(중복+VLM 낭비) → 폴백(LibreOffice)일 때만 수행.
        if ext in ('.hwp', '.hwpx') and api_key and not rhwp_done:
            try:
                from hwp_figures import extract_figures, significant
                from PIL import Image
                import io as _io
                figs = significant(extract_figures(file_path))
                if figs:
                    set_progress(f"🖼️ 임베디드 시각자료 {len(figs)}건 복원·분석 중...", 99)
                    fig_dir = os.path.join(doc_output_dir, "figures")
                    os.makedirs(fig_dir, exist_ok=True)
                    records = []; page_inserts = []
                    for fobj in figs:
                        try:
                            im = Image.open(_io.BytesIO(fobj["data"])).convert("RGB")
                        except Exception:
                            continue
                        fname = f"figure_{fobj['order']:04d}.png"
                        fpath = os.path.join(fig_dir, fname)
                        im.save(fpath)
                        desc = VLMProcessor.describe_image(fpath, api_key, model_name)
                        if not desc or desc.strip().upper().startswith("LOGO"):
                            try:
                                os.remove(fpath)   # 로고/장식 — VLM이 걸러냄
                            except OSError:
                                pass
                            continue
                        caption = fobj.get("context", "")[:80]
                        records.append({
                            "order": fobj["order"], "section": fobj["section"],
                            "para_index": fobj["para_index"], "ref": fobj["ref"],
                            "size_inch": [fobj["w_in"], fobj["h_in"]],
                            "context": caption, "anchor": fobj.get("anchor", "")[:80],
                            "image": f"figures/{fname}", "description": desc,
                        })
                        page_inserts.append({
                            "anchor": fobj.get("anchor", ""),
                            "element": {"type": "figure", "content": "", "caption": "",
                                        "description": desc, "image": f"figures/{fname}",
                                        "salvaged": True, "para_index": fobj["para_index"],
                                        "_source": "hwp_embedded_figure", "_confidence": 0.78},
                        })
                    with open(os.path.join(doc_output_dir, "figures.json"), "w", encoding="utf-8") as ff:
                        json.dump(records, ff, ensure_ascii=False, indent=2)
                    placed = 0
                    if output_format.lower() in ("json", "markdown") and page_inserts:
                        from hwp_figures import insert_figures_into_pages
                        placed = insert_figures_into_pages(doc_output_dir, page_inserts)
                        # markdown 모드: 페이지 .md 는 salvage 이전에 생성됐으므로 갱신된 json 으로 재생성
                        if placed and output_format.lower() == "markdown":
                            import glob as _glob
                            for _pj in _glob.glob(os.path.join(doc_output_dir, "page_*_structured.json")):
                                _stem = os.path.basename(_pj)[:-len("_structured.json")]
                                try:
                                    with open(_pj, encoding="utf-8") as _f:
                                        _md = DocumentProcessor.json_to_markdown(_f.read())
                                    with open(os.path.join(doc_output_dir, f"{_stem}_structured.md"), "w", encoding="utf-8") as _f:
                                        _f.write(_md)
                                except Exception as _e:
                                    logger.warning(f"salvage 후 {_stem} 마크다운 재생성 실패: {_e}")
                    logger.info(f"시각자료 salvage: {len(records)}건 → figures.json, 페이지 삽입 {placed}건")
            except Exception as e:
                logger.warning(f"시각자료 salvage 실패: {e}")

        # 페이지 분할된 네이티브 표의 행을 페이지에 1회씩 재분배(seam 손실·중복 방지).
        if output_format.lower() in ("json", "markdown"):
            try:
                import table_validate
                table_validate.repartition_native_tables(doc_output_dir)
            except Exception as e:
                logger.warning(f"네이티브 표 재분배 실패: {e}")

        # 임베디드 raster 표: 격리·확대 재추출로 오구조(열 붕괴·병합)를 수리(VLM 심판 승인 시만 교체).
        if ext in ('.hwp', '.hwpx') and api_key and output_format.lower() in ("json", "markdown"):
            try:
                from zoom_tables import zoom_raster_tables
                nz = zoom_raster_tables(doc_output_dir, file_path, api_key, model_name)
                if nz:
                    logger.info(f"raster 표 zoom-pass: {nz}건 교체")
            except Exception as e:
                logger.warning(f"raster 표 zoom-pass 실패: {e}")

        if chunk_strategies and output_format.lower() in ("json", "markdown"):
            try:
                from chunker import chunk_document
                set_progress("📦 청킹 중...", 98)
                chunk_document(doc_output_dir, strategies=chunk_strategies)
            except Exception as e:
                logger.error(f"청킹 오류: {e}")

        def _unique_page_list(values):
            ints = sorted({p for p in values if isinstance(p, int)})
            others = sorted({str(p) for p in values if not isinstance(p, int)})
            return ints + others

        # 부분 실패 가시화: VLM 구조추출 실패 페이지를 metadata 에 기록(프로그램 접근용).
        # 실패 페이지는 raw 텍스트로 폴백되어 내용은 보존되나 구조(heading/표)는 빠진다.
        try:
            meta_path = os.path.join(doc_output_dir, "metadata.json")
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    metadata = json.load(f)
            except Exception:
                metadata = {}
            metadata["vlm_pages_total"] = total_vlm_pages
            metadata["vlm_pages_analyzed"] = total_vlm_analyzed_pages
            provenance_totals = {"sources": {}, "issues": {}, "low_confidence_elements": 0}
            if output_format.lower() in ("json", "markdown"):
                for fname in sorted(f for f in os.listdir(doc_output_dir) if f.endswith("_structured.json")):
                    fpath = os.path.join(doc_output_dir, fname)
                    try:
                        with open(fpath, "r", encoding="utf-8") as f:
                            pdata = json.load(f)
                    except Exception:
                        continue
                    if pdata.get("low_confidence"):
                        stem = fname[:-len("_structured.json")]
                        try:
                            low_conf_pages.append(int(stem.split("_")[1]))
                        except (IndexError, ValueError):
                            low_conf_pages.append(stem)
                    ps = _collect_provenance_summary(pdata.get("elements", []))
                    provenance_totals["low_confidence_elements"] += ps["low_confidence_elements"]
                    for k, v in ps["sources"].items():
                        provenance_totals["sources"][k] = provenance_totals["sources"].get(k, 0) + v
                    for k, v in ps["issues"].items():
                        provenance_totals["issues"][k] = provenance_totals["issues"].get(k, 0) + v
            metadata["provenance_summary"] = provenance_totals
            failed_pages = _unique_page_list(failed_pages)
            low_conf_pages = _unique_page_list(low_conf_pages)
            metadata["vlm_failed_pages"] = failed_pages
            metadata["low_confidence_pages"] = low_conf_pages
            with open(meta_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"실패 페이지 메타 기록 실패: {e}")

        if os.path.exists(temp_pdf_dir):
            shutil.rmtree(temp_pdf_dir, ignore_errors=True)

        lc = _unique_page_list(low_conf_pages)
        if failed_pages or lc:
            msg = "⚠️ 완료"
            if failed_pages:
                msg += f" — 구조추출 실패 {len(failed_pages)}/{total_vlm_pages}쪽(폴백: {failed_pages})"
            if lc:
                msg += f" — 저신뢰(이미지 추출 빈약) {len(lc)}쪽: {lc} 원본 확인 필요"
            set_progress(msg, 100)
        else:
            set_progress("🎉 모든 구조화 완료!", 100)
        return doc_output_dir
